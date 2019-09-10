# -*- coding: utf-8 -*-
"""
Created on Tue Sep 25 12:46:03 2018

@author: hlan


This script is used for running the functions in other scripts that either create the data we need to run TSEB, 
or runs TSEB using the data we created. Also used for converting .npy arrays to excel files.
"""


import threading  #use this later maybe
import createStationArrays
from createStationArrays import runMultiTSEB_stations
from RunTSEB_Multi_Station_Arrays import runMultiTSEB
import xlsxwriter
from openpyxl import workbook
import numpy as np




def createAllStationArrays(Year): #This script creates the station arrays that we use to run TSEB later. Takes a while to run.
    
    
    
    for year in range(Year, 2016):
        
        for day in range(1, 366):
            
            
            try:
                runMultiTSEB_stations(year, day)
            
            except:
                print("-------------------Something went wrong--------------")
                continue
            
def runTSEBAllStationArrays(YearFrom, YearTo, DayFrom, DayTo): #runs TSEB using the data in the station arrays that we created in "createAllStationArrays"
    """
    YearFrom = includes that year, YearTo = up until, but not including that year.
    
    """

    for year in range(YearFrom, YearTo):
        
        for day in range(DayFrom, DayTo):
            
            
            try:
                runMultiTSEB(year, day) #this method will run TSEB and also create the .npy arrays currently.
            
            except:
                print("-------------------Something went wrong--------------")  
                continue            
     
    for year in range(YearFrom, YearTo):
        createExcelFromResultArrays(year) #now we takes the data from the .npy arrays and put it into excel sheets for each year.    
        
        
         #just keeping this here incase i need to manually write to the excel sheets.
#        createExcelFromResultArrays(2003)
#        createExcelFromResultArrays(2004)
#        createExcelFromResultArrays(2005)
#        createExcelFromResultArrays(2006)
#        createExcelFromResultArrays(2007)
#        createExcelFromResultArrays(2008)
#        createExcelFromResultArrays(2009)
#        createExcelFromResultArrays(2010)
#        createExcelFromResultArrays(2011)
#        createExcelFromResultArrays(2012)
#        createExcelFromResultArrays(2013)
#        createExcelFromResultArrays(2014)
#        createExcelFromResultArrays(2015)
#        createExcelFromResultArrays(2016)
            
#Path_ET = 'E:\\HenrikTSEB\\Station_Result_Arrays_ET\\'            
#Results_Array = np.load(Path_ET +str(2004) + '.npy')

            
def createExcelFromResultArrays(Year):  #creates the excel sheets based on the arrays.
    
        from openpyxl.reader.excel import load_workbook
        workbookr = load_workbook(filename="E:\\HenrikTSEB\\Station_Result_excel\\TSEB_Station_Output_UsingSmallArrays"+str(Year)+".xlsx")    
            
        #-------------------------------------------------------------------------------------------------------------- 
            
        #writing ET into excel in sheet 1.
        sheet = workbookr.get_sheet_by_name('ET')
        
        Path_ET = 'E:\\HenrikTSEB\\Station_Result_Arrays_ET\\'  #the path of the 7*10 array created in RunTSEB_Multi_Station_Arrays.py
    
        
        Results_Array = np.load(Path_ET +str(Year) + '.npy')    #load the .npy array into local array.
        
        

        wolo=-1
        count=-0
        for column in range(0,365): #iterate through days.
                    
            wolo=wolo+1
            for row in range(1,70): #iterate through stations
                count=count+1
                sheet.cell(row=1, column=column+1).value = str(column)
                if(Results_Array[row, column] != 0):
                    sheet.cell(row=row+1, column=column+1).value = Results_Array[row, column] 
                else:
                    sheet.cell(row=row+1, column=column+1).value = -999
                
         #----------------------------------------------------------------------------------------------------------        

        #writing LST into excel in sheet 2.        
        sheet = workbookr.get_sheet_by_name('LST')
        
        Path_LST = 'E:\\HenrikTSEB\\Station_Result_Arrays_LST\\'
    
        
        Results_Array = np.load(Path_LST +str(Year) + '.npy')
        
        

            
        wolo=-1
        count=-0
        for column in range(0,365):
                    
            wolo=wolo+1
            for row in range(1,70):
                count=count+1
                sheet.cell(row=1, column=column+1).value = str(column)
                
                if(Results_Array[row, column] != 0):
                    sheet.cell(row=row+1, column=column+1).value = Results_Array[row, column] 
                else:
                    sheet.cell(row=row+1, column=column+1).value = -999    
                
        #--------------------------------------------------------------------------------------------------------        

        #writing AirTemp into excel in sheet 3.        
        sheet = workbookr.get_sheet_by_name('AirTemp')
        
        Path_AirTemp = 'E:\\HenrikTSEB\\Station_Result_Arrays_AirTemp\\'
    
        
        Results_Array = np.load(Path_AirTemp +str(Year) + '.npy')

        wolo=-1
        count=-0
        for column in range(0,365):
                    
            wolo=wolo+1
            for row in range(1,70):
                count=count+1
                sheet.cell(row=1, column=column+1).value = str(column)
                if(Results_Array[row, column] != 0):
                    sheet.cell(row=row+1, column=column+1).value = Results_Array[row, column] 
                else:
                    sheet.cell(row=row+1, column=column+1).value = -999   
                
         #--------------------------------------------------------------------------------------------------------       
                
        
        #writing HC into excel sheet 4
        sheet = workbookr.get_sheet_by_name('HC')
        
        Path_HC = 'E:\\HenrikTSEB\\Station_Result_Arrays_HC\\'
    
        
        Results_Array = np.load(Path_HC +str(Year) + '.npy')

        wolo=-1
        count=-0
        for column in range(0,365):
                    
            wolo=wolo+1
            for row in range(1,70):
                count=count+1
                sheet.cell(row=1, column=column+1).value = str(column)
                if(Results_Array[row, column] != 0):
                    sheet.cell(row=row+1, column=column+1).value = Results_Array[row, column] 
                else:
                    sheet.cell(row=row+1, column=column+1).value = -999 
        
         #-------------------------------------------------------------------------------------------------------------- 
        
        #writing Ln into excel sheet 5        
        sheet = workbookr.get_sheet_by_name('LN')
        
        Path_HC = 'E:\\HenrikTSEB\\Station_Result_Arrays_Ln\\'
    
        
        Results_Array = np.load(Path_HC +str(Year) + '.npy')

        wolo=-1
        count=-0
        for column in range(0,365):
                    
            wolo=wolo+1
            for row in range(1,70):
                count=count+1
                sheet.cell(row=1, column=column+1).value = str(column)
                if(Results_Array[row, column] != 0):
                    sheet.cell(row=row+1, column=column+1).value = Results_Array[row, column] 
                else:
                    sheet.cell(row=row+1, column=column+1).value = -999
        
        #-------------------------------------------------------------------------------------------------------------- 
        
        #writing Sn into excel sheet 6        
        sheet = workbookr.get_sheet_by_name('SN')
        
        Path_Sn = 'E:\\HenrikTSEB\\Station_Result_Arrays_Sn\\'
    
        
        Results_Array = np.load(Path_Sn +str(Year) + '.npy')
        
        

        wolo=-1
        count=-0
        for column in range(0,365):
                    
            wolo=wolo+1
            for row in range(1,70):
                count=count+1
                sheet.cell(row=1, column=column+1).value = str(column)
                if(Results_Array[row, column] != 0):
                    sheet.cell(row=row+1, column=column+1).value = Results_Array[row, column] 
                else:
                    sheet.cell(row=row+1, column=column+1).value = -999
                    
        #-------------------------------------------------------------------------------------------------------------- 
                    
        #writing Ra into excel sheet7        
        sheet = workbookr.get_sheet_by_name('Ra')
        
        Path_R_a = 'E:\\HenrikTSEB\\Station_Result_Arrays_R_a\\'
    
        
        Results_Array = np.load(Path_R_a +str(2004) + '.npy')

        wolo=-1
        count=-0
        for column in range(0,365):
                    
            wolo=wolo+1
            for row in range(1,70):
                count=count+1
                sheet.cell(row=1, column=column+1).value = str(column)
                if(Results_Array[row, column] != 0):
                    sheet.cell(row=row+1, column=column+1).value = Results_Array[row, column] 
                else:
                    sheet.cell(row=row+1, column=column+1).value = -999
          
        #-------------------------------------------------------------------------------------------------------------- 
          
        #writing Rs into excel sheet 8        
        sheet = workbookr.get_sheet_by_name('Rs')
        
        Path_R_s = 'E:\\HenrikTSEB\\Station_Result_Arrays_R_s\\'
    
        
        Results_Array = np.load(Path_R_s +str(Year) + '.npy')

        wolo=-1
        count=-0
        for column in range(0,365):
                    
            wolo=wolo+1
            for row in range(1,70):
                count=count+1
                sheet.cell(row=1, column=column+1).value = str(column)
                if(Results_Array[row, column] != 0):
                    sheet.cell(row=row+1, column=column+1).value = Results_Array[row, column] 
                else:
                    sheet.cell(row=row+1, column=column+1).value = -999
                    
        #--------------------------------------------------------------------------------------------------------------   
                    
        #writing Rx into excel sheet 9        
        sheet = workbookr.get_sheet_by_name('Rx')
        
        Path_R_x = 'E:\\HenrikTSEB\\Station_Result_Arrays_R_x\\'
    
        
        Results_Array = np.load(Path_R_x +str(Year) + '.npy')

        wolo=-1
        count=-0
        for column in range(0,365):
                    
            wolo=wolo+1
            for row in range(1,70):
                count=count+1
                sheet.cell(row=1, column=column+1).value = str(column)
                if(Results_Array[row, column] != 0):
                    sheet.cell(row=row+1, column=column+1).value = Results_Array[row, column] 
                else:
                    sheet.cell(row=row+1, column=column+1).value = -999
          
        #-------------------------------------------------------------------------------------------------------------- 
          
        #writing G into excel sheet 10      
        sheet = workbookr.get_sheet_by_name('G')
        
        Path_G = 'E:\\HenrikTSEB\\Station_Result_Arrays_G\\'
    
        
        Results_Array = np.load(Path_G +str(Year) + '.npy')
        
        #Results_Array = np.load(Path_G +str(2003) + '.npy')

        wolo=-1
        count=-0
        for column in range(0,365):
                    
            wolo=wolo+1
            for row in range(1,70):
                count=count+1
                sheet.cell(row=1, column=column+1).value = str(column)
                if(Results_Array[row, column] != 0):
                    sheet.cell(row=row+1, column=column+1).value = Results_Array[row, column] 
                else:
                    sheet.cell(row=row+1, column=column+1).value = -999
        #--------------------------------------------------------------------------------------------------------------   
                    
        #writing Tc into excel sheet 11      
        sheet = workbookr.get_sheet_by_name('Tc')
        
        Path_Tc = 'E:\\HenrikTSEB\\Station_Result_Arrays_Tc\\'
    
        Results_Array = np.load(Path_Tc +str(Year) + '.npy')
 
        wolo=-1
        count=-0
        for column in range(0,365):
                    
            wolo=wolo+1
            for row in range(1,70):
                count=count+1
                sheet.cell(row=1, column=column+1).value = str(column)
                if(Results_Array[row, column] != 0):
                    sheet.cell(row=row+1, column=column+1).value = Results_Array[row, column] 
                else:
                    sheet.cell(row=row+1, column=column+1).value = -999
       
        #-------------------------------------------------------------------------------------------------------------- 
       
        #writing Ts into excel sheet 20      
        sheet = workbookr.get_sheet_by_name('Ts')
        
        Path_Ts = 'E:\\HenrikTSEB\\Station_Result_Arrays_Ts\\'
    
        Results_Array = np.load(Path_Ts +str(Year) + '.npy')

        wolo=-1
        count=-0
        for column in range(0,365):
                    
            wolo=wolo+1
            for row in range(1,70):
                count=count+1
                sheet.cell(row=1, column=column+1).value = str(column)
                if(Results_Array[row, column] != 0):
                    sheet.cell(row=row+1, column=column+1).value = Results_Array[row, column] 
                else:
                    sheet.cell(row=row+1, column=column+1).value = -999
                
        #-------------------------------------------------------------------------------------------------------------- 
                
        #saving the excel sheet.    
        workbookr.save("E:\\HenrikTSEB\\Station_Result_excel\\TSEB_Station_Output_UsingSmallArrays_alpha126"+str(Year)+".xlsx")    


def CombineExcelSheets(YearFrom, YearTo): #combines the excel sheets for each year into one big excel sheet with data from 2003-2015.
    from openpyxl.reader.excel import load_workbook
    workbookall = load_workbook(filename="E:\\HenrikTSEB\\Station_Result_excel\\TSEB_Station_Output_UsingSmallArrays_ALLYEARS2_Alpha2.xlsx")
    
    StationArray = np.load("E:\\HenrikTSEB\\StationNames\\StaionID.npy")
    StationArrayString = StationArray.astype('U') #convert the array to string so we can write the data into excel cells.
                
    for year in range(YearFrom, YearTo):
        workbookyears = load_workbook(filename="E:\\HenrikTSEB\\Station_Result_excel\\TSEB_Station_Output_UsingSmallArrays_alpha126"+str(year)+".xlsx")
            
        sheet_ET = workbookyears.get_sheet_by_name('ET')
        sheet_LST = workbookyears.get_sheet_by_name('LST')
        sheet_AirTemp = workbookyears.get_sheet_by_name('AirTemp')
        sheet_HC = workbookyears.get_sheet_by_name('HC')
        sheet_LN = workbookyears.get_sheet_by_name('LN')
        sheet_SN = workbookyears.get_sheet_by_name('SN')
        sheet_Ra = workbookyears.get_sheet_by_name('Ra')
        sheet_Rs = workbookyears.get_sheet_by_name('Rs')
        sheet_Rx = workbookyears.get_sheet_by_name('Rx')
        sheet_Ts = workbookyears.get_sheet_by_name('Ts')
        sheet_Tc = workbookyears.get_sheet_by_name('Tc')
        sheet_G = workbookyears.get_sheet_by_name('G')
            
        sheet_workbook_ET = workbookall.get_sheet_by_name('ET')
        sheet_workbook_LST = workbookall.get_sheet_by_name('LST')
        sheet_workbook_AirTemp = workbookall.get_sheet_by_name('AirTemp')
        sheet_workbook_HC = workbookall.get_sheet_by_name('HC')
        sheet_workbook_LN = workbookall.get_sheet_by_name('LN')
        sheet_workbook_SN = workbookall.get_sheet_by_name('SN')
        sheet_workbook_Ra = workbookall.get_sheet_by_name('Ra')
        sheet_workbook_Rs = workbookall.get_sheet_by_name('Rs')
        sheet_workbook_Rx = workbookall.get_sheet_by_name('Rx')
        sheet_workbook_Ts = workbookall.get_sheet_by_name('Ts')
        sheet_workbook_Tc = workbookall.get_sheet_by_name('Tc')
        sheet_workbook_G = workbookall.get_sheet_by_name('G')
            
        for column in range(1,365):
            for row in range(1,70):
                real_column = (year-2003)*365+column+1
                
                sheet_workbook_ET.cell(row=1, column=real_column).value = real_column
                sheet_workbook_LST.cell(row=1, column=real_column).value = real_column
                sheet_workbook_AirTemp.cell(row=1, column=real_column).value = real_column
                sheet_workbook_HC.cell(row=1, column=real_column).value = real_column
                sheet_workbook_LN.cell(row=1, column=real_column).value = real_column
                sheet_workbook_SN.cell(row=1, column=real_column).value = real_column
                sheet_workbook_Ra.cell(row=1, column=real_column).value = real_column
                sheet_workbook_Rs.cell(row=1, column=real_column).value = real_column
                sheet_workbook_Rx.cell(row=1, column=real_column).value = real_column
                sheet_workbook_Ts.cell(row=1, column=real_column).value = real_column
                sheet_workbook_Tc.cell(row=1, column=real_column).value = real_column
                sheet_workbook_G.cell(row=1, column=real_column).value = real_column
                
                
                            
                
                
                if(row<62):
                    sheet_workbook_ET.cell(row=row+1, column=1).value = StationArrayString[row-1]
                    sheet_workbook_LST.cell(row=row+1, column=1).value = StationArrayString[row-1]
                    sheet_workbook_AirTemp.cell(row=row+1, column=1).value = StationArrayString[row-1]
                    sheet_workbook_HC.cell(row=row+1, column=1).value = StationArrayString[row-1]
                    sheet_workbook_LN.cell(row=row+1, column=1).value = StationArrayString[row-1]
                    sheet_workbook_SN.cell(row=row+1, column=1).value = StationArrayString[row-1]
                    sheet_workbook_Ra.cell(row=row+1, column=1).value = StationArrayString[row-1]
                    sheet_workbook_Rs.cell(row=row+1, column=1).value = StationArrayString[row-1]
                    sheet_workbook_Rx.cell(row=row+1, column=1).value = StationArrayString[row-1]
                    sheet_workbook_Ts.cell(row=row+1, column=1).value = StationArrayString[row-1]
                    sheet_workbook_Tc.cell(row=row+1, column=1).value = StationArrayString[row-1]
                    sheet_workbook_G.cell(row=row+1, column=1).value = StationArrayString[row-1]
                
                    
                sheet_workbook_ET.cell(row=row+1, column=real_column).value = sheet_ET.cell(row=row+1, column=column).value
                sheet_workbook_LST.cell(row=row+1, column=real_column).value = sheet_LST.cell(row=row+1, column=column).value
                sheet_workbook_AirTemp.cell(row=row+1, column=real_column).value = sheet_AirTemp.cell(row=row+1, column=column).value
                sheet_workbook_HC.cell(row=row+1, column=real_column).value = sheet_HC.cell(row=row+1, column=column).value
                sheet_workbook_LN.cell(row=row+1, column=real_column).value = sheet_LN.cell(row=row+1, column=column).value
                sheet_workbook_SN.cell(row=row+1, column=real_column).value = sheet_SN.cell(row=row+1, column=column).value
                sheet_workbook_Ra.cell(row=row+1, column=real_column).value = sheet_Ra.cell(row=row+1, column=column).value
                sheet_workbook_Rs.cell(row=row+1, column=real_column).value = sheet_Rs.cell(row=row+1, column=column).value
                sheet_workbook_Rx.cell(row=row+1, column=real_column).value = sheet_Rx.cell(row=row+1, column=column).value
                sheet_workbook_Ts.cell(row=row+1, column=real_column).value = sheet_Ts.cell(row=row+1, column=column).value
                sheet_workbook_Tc.cell(row=row+1, column=real_column).value = sheet_Tc.cell(row=row+1, column=column).value
                sheet_workbook_G.cell(row=row+1, column=real_column).value = sheet_G.cell(row=row+1, column=column).value
                    
            
            
    workbookall.save("E:\\HenrikTSEB\\Station_Result_excel\\TSEB_Station_Output_UsingSmallArrays_ALLYEARS2_alpha126.xlsx")





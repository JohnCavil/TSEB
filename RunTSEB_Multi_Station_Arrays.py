# -*- coding: utf-8 -*-
"""
Created on Fri May 19 12:50:38 2017

@author: gmg

T_R1
theta1
T_A1
u1
ea1
p1
Rn_sw_veg
Rn_sw_soil
LWin
Fg
hc
emisVeg
emisGrd
z_0M
d_0M
zu
zt
leaf_width=leaf_width
z0_soil=z0_soil
alpha_PT=alpha_PT
x_LAD=x_L

--------------------------------------------------------------------------------------------
This script is used for running TSEB for the stations using the 7*10 arrays created in the "createStationArrays.py script
--------------------------------------------------------------------------------------------


"""
from __future__ import print_function
import numpy as np
import numpy.ma as ma
from osgeo import gdal
import pyTSEB
import os
import multiprocessing
import pandas
from joblib import Parallel, delayed
import xlsxwriter
from openpyxl import workbook
from PIL import Image
from skimage import io

fileName='Y:\\Gorka\\EUROPA\\FluxnetSitesCoordinates\\FluxnetSitesCoordinates.xlsx'

ExcelFile=pandas.read_excel(fileName,sheetname=0)
Rows=ExcelFile.RowPos
Cols=ExcelFile.ColPos
IGBP=ExcelFile.IGBP
StationID=ExcelFile.SITE_ID
StationName=ExcelFile.SITE_NAME





def GetFGDataFromTiff(Dictionary, JulDay):  #Gets the FG data from the tiff file (created in another script)
    
   im = io.imread('E:\\HenrikTSEB\\Fg_stations_only_output\\Fg_Image.mosaic.tif')           #reading the tiff file created in the FG_stations script.
   FgDoyArray = np.zeros([7, 10])                                                       
   TimeStep = JulDay/8    #THIS NEEDS TO ROUND UP!!                                         #finding the correct timestep to read.
   FgDoyArray = im[:, :, TimeStep]                                                          #reading one of the 46 time steps into the array for this specific DoY.
   
   Dictionary ['Fg'] = FgDoyArray
   
   return Dictionary
   
    
    


def getSceneMetadata(reference_image):
    import gdal
    metadata = {'xsize':0 ,'ysize': 0, 'bands':0, 'gt':[], 'proj':''}
    # Get reprojection parameters from the first dataset
#    scene = '"'+scene+'":Grid:band1'
    inputImage = gdal.Open(reference_image,gdal.GA_ReadOnly)
    if inputImage:
        metadata['xsize'] = inputImage.RasterXSize
        metadata['ysize'] = inputImage.RasterYSize
        metadata['bands'] = inputImage.RasterCount
        metadata['gt'] = inputImage.GetGeoTransform()
        metadata['proj'] = inputImage.GetProjection()
        inputImage = None
    return metadata
    
def readValueArray(array,row,col,ID):           #reads data from an array by row and column. Used for reading stations.
    
    import netCDF4
    import numpy as np

    value = array[row,col]   
    
    return value    
    
def extractStationsFromMap(Map):                        #gets the data for just the stations out of a 3600x3600 array
    station_list_data=np.zeros([len(StationID),46])
    globalMaxLaiArray_Out_put=np.zeros([len(StationID)])
    LandCoverVector=np.zeros([len(StationID)])         
    ID_all= [''  for x in xrange(len(Rows))]
    IGBP_all= [''  for x in xrange(len(Rows))]
    i=0
    


    for Station in range(0,len(Rows)):

        Row=Rows[Station]
        Col=Cols[Station]
        ID=StationID[Station]
        IGBP_Class=IGBP[Station]
        TestRead=readValueArray(Map,Row,Col,ID)
        
        station_list_data[i,:]=TestRead
        #globalMaxLaiArray_Out_put[i]=globalMaxLaiArray[Row,Col]
        #print(Row,Col,LAI,ID)
        ID_all[i]=ID
        IGBP_all[i]=IGBP_Class
        
        
        i=i+1
    
    return station_list_data
      

def MakeListIntoMap(mlist):          #takes a list of stations and makes it into a 10x60 array like a map. Usefull since we use the station data from a grid like map.
    Map_array = np.zeros([7,10])
    countRow = 0
    
    for n in range(0,10):
        
        for o in range(0,7):
            
            if(countRow<61):
                Map_array[o,n]=mlist[countRow, 0]
                countRow=countRow+1

    return Map_array    

 
def savefilesTSEB(folder,year,doy,geotransform,proj,flag, Ts, Tc, T_AC,L_nS,L_nC, LE_C,H_C,LE_S,H_S,G,R_s,R_x,R_a,u_friction, L,n_iterations):
    import numpy as np
    
    stryear=str(year)
    if doy < 100:
        if doy<10:
            strdoy='00'+str(doy)
        else:
            strdoy='0'+str(doy)
    else:
         strdoy=str(doy)   
    yeardoy= stryear+ strdoy  
    outputfile=folder+'flag\\Flag_'+yeardoy+'.tif'
    Save_array_tiff(flag, outputfile,geotransform, proj)
    outputfile=folder+'Ts\\Ts_'+yeardoy+'.tif'
    Ts[flag==255]=np.nan
    #Ts[Ts==0]=np.nan
    Save_array_tiff(Ts, outputfile,geotransform, proj)
    outputfile=folder+'Tc\\Tc_'+yeardoy+'.tif'
    Tc[flag==255]=np.nan
    #Tc[Tc==0]=np.nan
    Save_array_tiff(Tc, outputfile,geotransform, proj)
    outputfile=folder+'T_AC\\T_AC_'+yeardoy+'.tif'
    T_AC[flag==255]=np.nan
    #T_AC[T_AC==0]=np.nan
    Save_array_tiff(T_AC, outputfile,geotransform, proj)
    outputfile=folder+'L_nS\\L_nS_'+yeardoy+'.tif'
    L_nS[flag==255]=np.nan
    #L_nS[L_nS==0]=np.nan
    Save_array_tiff(L_nS, outputfile,geotransform, proj)
    outputfile=folder+'L_nC\\L_nC_'+yeardoy+'.tif'
    L_nC[flag==255]=np.nan
    #L_nC[L_nC==0]=np.nan
    Save_array_tiff(L_nC, outputfile,geotransform, proj)
    outputfile=folder+'LE_C\\LE_C_'+yeardoy+'.tif'
    LE_C[flag==255]=np.nan
    LE_C[LE_C<=0]=np.nan
    Save_array_tiff(LE_C, outputfile,geotransform, proj)
    outputfile=folder+'H_C\\H_C_'+yeardoy+'.tif'
    H_C[flag==255]=np.nan
    H_C[H_C<=0]=np.nan
    Save_array_tiff(H_C, outputfile,geotransform, proj)
    outputfile=folder+'LE_S\\LE_S_'+yeardoy+'.tif'
    LE_S[flag==255]=np.nan
    LE_S[LE_S<=0]=np.nan
    Save_array_tiff(LE_S, outputfile,geotransform, proj)
    outputfile=folder+'H_S\\H_S_'+yeardoy+'.tif'
    H_S[flag==255]=np.nan
    H_S[H_S<=0]=np.nan
    Save_array_tiff(H_S, outputfile,geotransform, proj)
    outputfile=folder+'G\\G_'+yeardoy+'.tif'
    G[flag==255]=np.nan
    G[G<=0]=np.nan
    Save_array_tiff(G, outputfile,geotransform, proj)
    outputfile=folder+'R_s\\R_s_'+yeardoy+'.tif'
    R_s[flag==255]=np.nan
    #R_s[R_s==0]=np.nan
    Save_array_tiff(R_s, outputfile,geotransform, proj)
    outputfile=folder+'R_x\\R_x_'+yeardoy+'.tif'
    R_x[flag==255]=np.nan
    #R_x[R_x==0]=np.nan
    Save_array_tiff(R_x, outputfile,geotransform, proj)
    outputfile=folder+'R_a\\R_a_'+yeardoy+'.tif'
    R_a[flag==255]=np.nan
    #R_a[R_a==0]=np.nan
    Save_array_tiff(R_a, outputfile,geotransform, proj)
    outputfile=folder+'u_friction\\u_frictions_'+yeardoy+'.tif'
    u_friction[flag==255]=np.nan
    u_friction[u_friction==0]=np.nan
    Save_array_tiff(u_friction, outputfile,geotransform, proj)
    outputfile=folder+'L\\L_'+yeardoy+'.tif'
    L[flag==255]=np.nan
    #L[L<=0]=np.nan
    Save_array_tiff(L, outputfile,geotransform, proj)
#    outputfile=folder+'n_iterations\\n_iterations_'+yeardoy+'.tif'
#    Save_array_tiff(n_iterations, outputfile,geotransform, proj)
    LE_Total=LE_C+LE_S
    LE_Total[LE_Total<=0]=np.nan
    outputfile=folder+'LE_Total\\'+yeardoy+'_LE_EUROPE.tif'
    Save_array_tiff(LE_Total, outputfile,geotransform, proj)
 
 
#    outputfile=folder+'Landcover\\Landcover_Map.tif'
#    Save_array_tiff(landcover, outputfile,geotransform, proj)
#    landcover
def Save_array_tiff(array, outputfile,geotransform, proj):
    
    from osgeo import gdal, osr
 
    # Start the gdal driver for GeoTIFF
    driver = gdal.GetDriverByName("GTiff")
 
    # Write array
    shape=array.shape
    print(len(shape))
 
    if len(shape) > 2:
        ds = driver.Create(outputfile, shape[1], shape[0], shape[2], gdal.GDT_Float32)
        ds.SetProjection(proj)
        ds.SetGeoTransform(geotransform)
 
        for i in range(shape[2]):
            print(i)
            ds.GetRasterBand(i+1).WriteArray(array[:,:,i])
 
    else:
        ds = driver.Create(outputfile, shape[1], shape[0], 1, gdal.GDT_Float32)
        ds.SetProjection(proj)
        ds.SetGeoTransform(geotransform)
        ds.GetRasterBand(1).WriteArray(array)
 
    ds=None
 
    print('Saved ' +outputfile )
 
    return


def createoutputfolders(folder):
    if not os.path.exists(folder+'outputs\\flag\\'):
        newpath=folder+'outputs\\flag'
        os.makedirs(newpath)
    if not os.path.exists(folder+'outputs\\G'):
        newpath=folder+'outputs\\G'
        os.makedirs(newpath)
    if not os.path.exists(folder+'outputs\\H_C'):
        newpath=folder+'outputs\\H_C'
        os.makedirs(newpath)
    if not os.path.exists(folder+'outputs\\H_S'):
        newpath=folder+'outputs\\H_S'
        os.makedirs(newpath)
    if not os.path.exists(folder+'outputs\\L'):
        newpath=folder+'outputs\\L'
        os.makedirs(newpath)
    if not os.path.exists(folder+'outputs\\L_nC'):
        newpath=folder+'outputs\\L_nC'
        os.makedirs(newpath)
    if not os.path.exists(folder+'outputs\\L_nS'):
        newpath=folder+'outputs\\L_nS'
        os.makedirs(newpath)
    if not os.path.exists(folder+'outputs\\Landcover'):
        newpath=folder+'outputs\\Landcover'
        os.makedirs(newpath)
    if not os.path.exists(folder+'outputs\\LE_C'):
        newpath=folder+'outputs\\LE_C'
        os.makedirs(newpath)
    if not os.path.exists(folder+'outputs\\LE_S'):
        newpath=folder+'outputs\\LE_S'
        os.makedirs(newpath)
    if not os.path.exists(folder+'outputs\\LE_Total'):
        newpath=folder+'outputs\\LE_Total'
        os.makedirs(newpath)
    if not os.path.exists(folder+'outputs\\n_iterations'):
        newpath=folder+'outputs\\n_iterations'
        os.makedirs(newpath)
    if not os.path.exists(folder+'outputs\\R_a'):
        newpath=folder+'outputs\\R_a'
        os.makedirs(newpath)
    if not os.path.exists(folder+'outputs\\R_s'):
        newpath=folder+'outputs\\R_s'
        os.makedirs(newpath)
    if not os.path.exists(folder+'outputs\\R_x'):
        newpath=folder+'outputs\\R_x'
        os.makedirs(newpath)
    if not os.path.exists(folder+'outputs\\T_AC'):
        newpath=folder+'outputs\\T_AC'
        os.makedirs(newpath)
    if not os.path.exists(folder+'outputs\\Tc'):
        newpath=folder+'outputs\\Tc'
        os.makedirs(newpath)
    if not os.path.exists(folder+'outputs\\Ts'):
        newpath=folder+'outputs\\Ts'
        os.makedirs(newpath)
    if not os.path.exists(folder+'outputs\\u_friction'):
        newpath=folder+'outputs\\u_friction'
        os.makedirs(newpath)
        
def runTSEB(Dictionary):
    from TSEB import TSEB_PT
    Tr_K=Dictionary['T_R'][:]
    vza=Dictionary['theta1'][:] 
    Ta_K=Dictionary['T_A'][:] 
    vza=Dictionary['theta1'][:]
    u=Dictionary['u'][:]
    ea=Dictionary ['ea'][:]
    p=Dictionary ['p1'][:]
    Sn_C=Dictionary['Rn_sw_veg'][:] 
    Sn_S=Dictionary['Rn_sw_soil'][:] 
    Lsky=Dictionary ['LWin'][:]
    LAI=(Dictionary ['LAI'][:])
    #LAI[LAI==25.5]=0.0
    hc=Dictionary ['hc'][:]
    emisVeg=(Dictionary['emisVeg'][:])
    emisGrd=(Dictionary['emisGrd'][:])
    z_0M=Dictionary ['z_0M'][:]
    d_0=Dictionary ['d_0M'][:]
    zu=Dictionary ['zu'][:]
    zt=Dictionary ['zt'][:]
    alpha_PT=(Dictionary['alpha_PT'][:])
    leaf_width=(Dictionary['leaf_width'][:])
    z0_soil=(Dictionary['z0_soil'][:])
    wc=Dictionary['wc'][:]
    x_LAD=Dictionary['x_LAD'][:]
    fg=Dictionary['Fg'][:]
    
    #FG set to 1:
    #flag, Ts, Tc, T_AC,L_nS,L_nC, LE_C,H_C,LE_S,H_S,G,R_s,R_x,R_a,u_friction, L,n_iterations=TSEB_PT(Tr_K,vza,Ta_K,u,ea,p,Sn_C,Sn_S,Lsky,LAI,hc,emisVeg,emisGrd,z_0M,d_0,zu,zt,leaf_width=leaf_width,z0_soil=z0_soil,alpha_PT=alpha_PT,x_LAD=x_LAD,f_c=0.999990,f_g=0.999990,wc=wc,Resistance_flag=0,calcG_params=[[1],0.35], UseL=False)
    
    #using real FG:
    flag, Ts, Tc, T_AC,L_nS,L_nC, LE_C,H_C,LE_S,H_S,G,R_s,R_x,R_a,u_friction, L,n_iterations=TSEB_PT(Tr_K,vza,Ta_K,u,ea,p,Sn_C,Sn_S,Lsky,LAI,hc,emisVeg,emisGrd,z_0M,d_0,zu,zt,leaf_width=leaf_width,z0_soil=z0_soil,alpha_PT=alpha_PT,x_LAD=x_LAD,f_c=0.999990,f_g=fg,wc=wc,Resistance_flag=0,calcG_params=[[1],0.35], UseL=False)
    return flag, Ts, Tc, T_AC,L_nS,L_nC, LE_C,H_C,LE_S,H_S,G,R_s,R_x,R_a,u_friction, L,n_iterations
    
    
def GetShortWaveRatiation(Dictionary, Year, JulDay):
       
    
    from netRadiation import CalcDifuseRatio,CalcSnCampbell
    import glob
    import gdal
    import os
    import numpy as np
    
    Path_soil ='E:\\HenrikTSEB\\Station_Data_Arrays\\Rn_sw_soil\\'
    Path_veg ='E:\\HenrikTSEB\\Station_Data_Arrays\\Rn_sw_veg\\'
    
    
        
    Rn_sw_veg_Map_array = np.load(Path_veg +str(Year) + '-' + str(JulDay) + '.npy')
        
    
        
    
        
    Rn_sw_soil_Map_array = np.load(Path_soil +str(Year) + '-' + str(JulDay) + '.npy')
        
    Dictionary['Rn_sw_veg']=  Rn_sw_veg_Map_array  #return DiffuseSwRad,DirectSwRad,Sdn
    Dictionary ['Rn_sw_soil']=   Rn_sw_soil_Map_array
    return Dictionary
    
    



def GetLancover_IGBP_Gorka(Dictionary,Year, JulDay):
    
    
    
    
     
    import netRadiation
    import resistances
    import numpy as np
    import numpy.ma as ma
    from osgeo import gdal
    Path = 'E:\\HenrikTSEB\\Station_Data_Arrays\\lc_IGBP\\'
    
    
        
    landcover_Map_array = np.load(Path +str(Year) + '-' + str(JulDay) + '.npy')
    
    landcover=np.zeros([7,10])
    landcover=landcover_Map_array            
            
       
    
    LAI=Dictionary['LAI'][:]
    
    #manually overriding landcover classes
    real_landcover = np.zeros([7,10])
    real_landcover[0,0] = 12
    real_landcover[1,0] = 12
    real_landcover[2,0] = 12
    real_landcover[3,0] = 12
    real_landcover[4,0] = 12
    real_landcover[5,0] = 12
    real_landcover[6,0] = 12
    real_landcover[0,1] = 12
    real_landcover[1,1] = 12
    real_landcover[2,1] = 12 
    real_landcover[3,1] = 6
    real_landcover[4,1] = 4
    real_landcover[5,1] =4
    real_landcover[6,1] =4
    real_landcover[0,2] =4
    real_landcover[1,2] =4
    real_landcover[2,2] =4
    real_landcover[3,2] =4
    real_landcover[4,2] =4
    real_landcover[5,2] =4
    real_landcover[6,2] =4
    real_landcover[0,3] = 4
    real_landcover[1,3] = 2
    real_landcover[2,3] =2
    real_landcover[3,3] =2
    real_landcover[4,3] =1
    real_landcover[5,3] =1
    real_landcover[6,3] =1
    real_landcover[0,4] =1
    real_landcover[1,4] =1
    real_landcover[2,4] =1
    real_landcover[3,4] = 1
    real_landcover[4,4] =1
    real_landcover[5,4] =1
    real_landcover[6,4] =1
    real_landcover[0,5] =1
    real_landcover[1,5] =1
    real_landcover[2,5] =1
    real_landcover[3,5] =10
    real_landcover[4,5] =10
    real_landcover[5,5] =10
    real_landcover[6,5] = 10
    real_landcover[0,6] =10
    real_landcover[1,6] =10
    real_landcover[2,6] =10
    real_landcover[3,6] =10
    real_landcover[4,6] =10
    real_landcover[5,6] =10
    real_landcover[6,6] =10
    real_landcover[0,7] =5
    real_landcover[1,7] =5
    real_landcover[2,7] = 5
    real_landcover[3,7] =7
    real_landcover[4,7] =7
    real_landcover[5,7] =7
    real_landcover[6,7] =7
    real_landcover[0,8] =11
    real_landcover[1,8] =11
    real_landcover[2,8] =11
    real_landcover[3,8] =11
    real_landcover[4,8] =11
    
    Dictionary['lc_IGBP'] =np.float64(real_landcover)
        
    return Dictionary



def GetInterpolatedLAI(Folder,Year,JulDay,Dictionary):
    
    
    
    print('Getting the LAI data....!!!!')
    import glob
    import numpy as np
    Path = 'E:\\HenrikTSEB\\Station_Data_Arrays\\LAI\\'
        
    Interpolated_Map_array = np.load(Path +str(Year) + '-' + str(JulDay) + '.npy')           
            
    Dictionary['LAI']=Interpolated_Map_array
    
    return Dictionary
    
    
        
def Calc_VaporPressure(T_K):
    
    # Calculates the vapor pressure in hPa, required temperature in Kelvin
    from numpy import exp,zeros,shape
    ea=zeros(shape(T_K))-1
    T_C=T_K-273.15
    ea = 6.112 * exp((17.67*T_C)/(T_C + 243.5))
    return ea
    
def openEuropeanLSTfile(Filename):
    import os
    import gdal
    import numpy as np
    filename=os.path.split(Filename)
    print ('The file '+filename[1]+' was found. Extracting the information from the file...')
    fid = gdal.Open(Filename,gdal.GA_ReadOnly)
    LST_Dataset=np.zeros([fid.RasterXSize,fid.RasterYSize,fid.RasterCount])
    LST_Dataset[LST_Dataset==0]=np.nan
    for i in range(fid.RasterCount):
        print(filename[1]+' succesfully opened. Reading band '+str(i))
        LST_Dataset[:,:,i]=fid.GetRasterBand(i+1).ReadAsArray()
    
    return LST_Dataset
def MakeTSEBDictionary():
    Dictionary              =   {}
    Dictionary2 = {}
    Dictionary ['T_R']       =   [] #Done
    Dictionary ['theta1']    =   [] #Done
    Dictionary ['T_A']      =   [] #Done
    Dictionary ['u']        =   [] #Done
    Dictionary ['ea1']        =   []
    Dictionary ['p1']        =   []
    Dictionary ['Sdn'] =         []
    Dictionary ['Rn_sw_veg'] =   []
    Dictionary ['Rn_sw_soil']=   []
    Dictionary ['LWin']      =   []
    Dictionary ['Fg']        =   []
    Dictionary ['hc']        =   []
    Dictionary ['emisVeg']   =   []
    Dictionary ['emisGrd']   =   []
    Dictionary ['z_0M']      =   []
    Dictionary ['d_0M']      =   []
    Dictionary ['zu']        =   []
    Dictionary ['zt']        =   []
    Dictionary ['sza']        =   []
    Dictionary ['lc_IGBP']        =   []
    Dictionary ['LAI']        =   []  
    Dictionary ['FCover']        =   [] 
    
    
    
    Dictionary2=Dictionary                          #we make a copy of the Dictionary because we need a variables to be in the 3600x3600 format. So some functions will put the stations map array of 10x6 into Dictionary, and the full picture of 3600x3600 into Dictionary2.
    return Dictionary

def MakeTSEBDictionary2():
    Dictionary              =   {}
    Dictionary2 = {}
    Dictionary ['T_R']       =   [] #Done
    Dictionary ['theta1']    =   [] #Done
    Dictionary ['T_A']      =   [] #Done
    Dictionary ['u']        =   [] #Done
    Dictionary ['ea1']        =   []
    Dictionary ['p1']        =   []
    Dictionary ['Sdn'] =         []
    Dictionary ['Rn_sw_veg'] =   []
    Dictionary ['Rn_sw_soil']=   []
    Dictionary ['LWin']      =   []
    Dictionary ['Fg']        =   []
    Dictionary ['hc']        =   []
    Dictionary ['emisVeg']   =   []
    Dictionary ['emisGrd']   =   []
    Dictionary ['z_0M']      =   []
    Dictionary ['d_0M']      =   []
    Dictionary ['zu']        =   []
    Dictionary ['zt']        =   []
    Dictionary ['sza']        =   []
    Dictionary ['lc_IGBP']        =   []
    Dictionary ['LAI']        =   []  
    Dictionary ['FCover']        =   [] 
    
    
    
    Dictionary2=Dictionary                          #we make a copy of the Dictionary because we need a variables to be in the 3600x3600 format. So some functions will put the stations map array of 10x6 into Dictionary, and the full picture of 3600x3600 into Dictionary2.
    return Dictionary2
    
def generateAquaTerraLSTdataset11_13(LST_Dataset,startimeWindow,EndTimeWindow):
     #Time window is 11:00 to 13:00. Also input is hour*10. For example 110 and 130, rather than 11 or 13.
    from pymasker import Masker
    import numpy as np
    #The LST_Dataset must containg the LST,ViewTime,Angle and quality flag band
    #Band positions in the LST dataset are:
        #Band 0 LST
        #Band 1 Quality flags
        #Band 2 Day view time 
        #Band 3 Day View Angle
    masker=Masker(np.uint8(LST_Dataset[:,:,1]))
    LST=LST_Dataset[:,:,0]
    ViewZenithAngle=LST_Dataset[:,:,3]
    GooDdata=masker.get_mask(0,2,'00') #00 = LST produced, good quality, not necessary to examine detailed QA
    #https://lpdaac.usgs.gov/dataset_discovery/modis/modis_products_table/mod11a1
    GooDdata_AverageLess1k=masker.get_mask(6,2,'00')    
        #Masked values are those with value 1
    Time=LST_Dataset[:,:,2]
    idx=(Time>=startimeWindow) & (Time<=EndTimeWindow)
    Mask=GooDdata+GooDdata_AverageLess1k+idx
    LST[Mask<2]=np.nan
    Time[Mask<2]=np.nan
    ViewZenithAngle[Mask<2]=np.nan
    SensorObs=np.zeros([Time.shape[0],Time.shape[1]])
    SensorObs[Mask >=2]=1
    return LST,Time,ViewZenithAngle,SensorObs
    
    
def FillDicTrAndTheta(Dictionary,Year,JulDay):
    #Fill Dictionary['T_R'] &Dictionary['theta1']=[]
    
    
    import glob
    import gdal
    import os
    import numpy as np
    Path_T_R ='E:\\HenrikTSEB\\Station_Data_Arrays\\Tr\\'
    Path_thetal  ='E:\\HenrikTSEB\\Station_Data_Arrays\\Theta\\'  
    
    
    
        
    temp_Map_array = np.load(Path_T_R +str(Year) + '-' + str(JulDay) + '.npy')
    
    
        
        
        
    temp2_Map_array = np.load(Path_thetal +str(Year) + '-' + str(JulDay) + '.npy')
        
    Dictionary ['T_R']       =  temp_Map_array
    Dictionary ['T_R'][:]
    Dictionary ['theta1']    =  temp2_Map_array
#    if  len(List_Aqua) is 0:
#        #SInce Aqua was launched later we need to account for the change of not existing data. IOn this cases only Terra is used
#        LST_Terra[LST_Terra==0]=np.nan
#        ViewZenithAngle_Terra[ViewZenithAngle_Terra==0]=np.nan
#        Dictionary ['T_R']       =   LST_Terra*0.02
#        Dictionary ['theta1']    =   np.abs(ViewZenithAngle_Terra-65)
        
    
    
    return Dictionary
    
    
def FillDicTairWaterVaporPressure(Dictionary,Year,JulDay):
    
    #Year = 2003
    #JulDay = 14
    
    import glob
    import gdal
    import os
    import numpy as np
    Path ='E:\\HenrikTSEB\\Station_Data_Arrays\\AirTemp\\'
    
    
    
    
    
    
    
    print('Reading the Air temperature file now. Please wait...!!!!')
    AirTemp=np.load(Path +str(Year) + '-' + str(JulDay) + '.npy')
    
        
        
    AirTemp_Map_array = AirTemp
        
    Dictionary ['T_A']       =  AirTemp_Map_array
    
    #ea=Calc_VaporPressure(AirTemp)
        
        #ea[np.isnan(Dictionary ['T_R'][:])]=np.nan
        
        #testttt=ea[2024,1067]
        
#-------------------------------------------------------------------------------------------------------------        
        
    Path ='E:\\HenrikTSEB\\Station_Data_Arrays\\ea\\'    
        
    ea_Map_array = np.load(Path +str(Year) + '-' + str(JulDay) + '.npy')        
        
        
    
    Dictionary['ea']=ea_Map_array
    
        
    return Dictionary
    
    
def FillDicPressure(Dictionary,Year,JulDay):
    
    
    
    
    
    
    
    import glob
    import gdal
    import os
    import numpy as np
    Path = 'E:\\HenrikTSEB\\Station_Data_Arrays\\pl\\'
    
        
    SPressure_Map_array = np.load(Path +str(Year) + '-' + str(JulDay) + '.npy')         
        
        
        
    Dictionary ['p1']       =   SPressure_Map_array
        
    
        
    return Dictionary    
#def FillFractionalCover(Dictionary,Year,JulDay):
#    import glob
#    import gdal
#    import os
#    import numpy as np
#    Path ='Y:\\Gorka\\EUROPA\\ERA_Europe_Radiation\\Outputs\\SP\\'
#    List= None  ;
#    List = glob.glob(Path + 'SP_' +str(Year) + str(JulDay) + 'Simuated12.tif')
#    if List:
#        fid = gdal.Open(List[0],gdal.GA_ReadOnly)
#
#        Dictionary ['FCover']       =   SPressure
#    else:
#        print('File not found. Check why!!!!')
#        
#    return Dictionary        
    
def FillDicLwIn(Dictionary,Year,JulDay):
    #Year= 2005
    #JulDay = 50
    
    import glob
    import gdal
    import os
    import numpy as np
    Path ='E:\\HenrikTSEB\\Station_Data_Arrays\\LWin\\'
    
        
    LwIn_Map_array = np.load(Path +str(Year) + '-' + str(JulDay) + '.npy')       
        

        #LwIn[np.isnan(Dictionary ['T_R'][:])]=np.nan
    Dictionary ['LWin']       =   LwIn_Map_array
    
    
        
    return Dictionary 
    
def returnGregorianDate(Year,JulDay):
    from jdcal import gcal2jd, jd2gcal
    Jul=gcal2jd(Year,1,1)
    Jul=list(Jul)
    #Jul[1]=Jul[1]+(JulDay-1)#Carefull here. THe Sunzenith has a bug in wiritting the filename and therefore the -1 is not necesary
    Jul[1]=Jul[1]+(JulDay)#Carefull here. THe Sunzenith has a bug in wiritting the filename and therefore the -1 is not necesary
    Gcal=jd2gcal(Jul[0],Jul[1])
    Year=str(Gcal[0])
    if Gcal[1] <10:
        Month='0'+str(Gcal[1])
    else:
        Month=str(Gcal[1])
        
    if Gcal[2] <10:
        Day='0'+str(Gcal[2])
    else:
        Day=str(Gcal[2])
    
    return Year,Month,Day
        
def FillSunZenithAngle(Dictionary,Year,JulDay):
    
    
    import glob
    import gdal
    import os
    import numpy as np
    
    
    Path ='E:\\HenrikTSEB\\Station_Data_Arrays\\sza\\'
       
    SunZenithAngle_Map_array = np.load(Path +str(Year) + '-' + str(JulDay) + '.npy')       
        
    
    Dictionary ['sza']       =   SunZenithAngle_Map_array
        
    return Dictionary      

def FillLAI(Dictionary,Year,JulDay):
    import glob
    import gdal
    import os
    import numpy as np
    
    Path ='Y:\\Gorka\\EUROPA\\Timesat\\Outputs\\LAI_Weight\\'
    List= None  ;
    List = glob.glob(Path + 'MCD15A2.005' +str(Year)+str(JulDay)+'.Weightsmosaic.asc')
    if List:
        print('Reading the LAI file now. Please wait...!!!!')
        array=np.fromfile(List[0],dtype='uint8')
        array=np.reshape(array, [3600,3600])
        LAI_array = array*1.0
        #LAI_array[np.isnan(Dictionary ['T_R'][:])]=np.nan
        
        LAI_array_stations=np.zeros([len(StationID),46])
        LAI_array_stations=extractStationsFromMap(LAI_array)
        
        LAI_Map_array = np.zeros([7,10])
        
        LAI_Map_array = MakeListIntoMap(LAI_array_stations)
        
    
        
        Dictionary ['LAI']       =   LAI_array_stations
    else:
        print('File not found. The date is not one 8 day composite. Starting interpolation between dates and filling the dictionary...')
        File1,File2=select8dayimage(Folder,Year,JulDay)
        
        
    return Dictionary           


    
def FillDicWindSpeedr(Dictionary,Year,JulDay):
    
    
    
    import glob
    import gdal
    import os
    import numpy as np
    Path = 'E:\\HenrikTSEB\\Station_Data_Arrays\\WindVel\\'
    
    
   
    WindVel= np.load(Path +str(Year) + '-' + str(JulDay) + '.npy')
    
    
    #------------------------------------------------------------------------------------------------------------
    
    
    WindVel_Map_array= WindVel        
    WindVel_Map_array[WindVel_Map_array <0.1]=0.1
    
    Dictionary['u']=WindVel_Map_array
    
    
    
    
    return Dictionary



    
def createDailyNetCDFfile4TSEB(year,doy,Dictionary,flag, Ts, Tc, T_AC,L_nS,L_nC, LE_C,H_C,LE_S,H_S,G,R_s,R_x,R_a,u_friction, L):
    import numpy   as np
    import netCDF4 as nc  
    import numpy.ma as ma
    import struct
    import time, os, sys                     # call current time for timestamp
    from writenetcdf import writenetcdf # from ufz
    from readnetcdf  import readnetcdf  # from ufz
    from fread       import fread       # from ufz
    from autostring  import astr        # from ufz
    from pyproj      import Proj
    import scipy.misc
    import scipy.ndimage
    from netCDF4 import Dataset
    import glob
    import re
    import gdal
    years              = str(year)+','+str(year)
    startyear    = int(years.split(',')[0])
    endyear      = int(years.split(',')[1])
    times      = np.arange(0,doy+1,doy)
    
    
    
    #
    # COORDINATE SYSTEM
    #   code 31463 defines GK (DHDN3 Zone 3)
    #   [website: http://www.spatialreference.org/ref/epsg/31463/]
    #   equal +proj=tmerc +ellps=bessel +lon_0=12 +x_0=3,500,000 +y_0=0
    # coord_sys = 'epsg:31467'
    coord_sys = 'ESRI:53008'  #https://epsg.io/2211-8653
    # HEADER FILE
    #   specifies the grid properties
    #   for example, use a copy of the header.txt
    #   and adapt cellsize, ncols, nrows to your hydrologic resolution
    
    
    headerfile = 'Y:\\Gorka\\EUROPA\\Header_Europe_stations.txt' #this is for the stations. Has a 10x6 grid instead of 3600x3600
    #headerfile = 'Y:\\Gorka\\EUROPA\\Header_Europe.txt'
    
    # OUTPUT FILE
    #   path to the output file, latlon.nc is hard-coded in mHM
    outfile = 'E:\\HenrikTSEB\\NetCDFfile4TSEB\\InputsTSEB_'+str(year)+'_'+str(doy)+'.nc'
    
    import optparse
    parser = optparse.OptionParser(usage='%prog [options]',
                                   description="Cretaes latitude and longitude grids for <coord_sys> with the domain defined in  <headerfile> into NetCDF file <outfile>.")
    
    # usage example with command line arguments
    # -------------------------------------------------------------------------
    #
    # py create_latlon.py -c 'epsg:31463' -f header.txt -o latlon.nc
    #
    # -------------------------------------------------------------------------
    
    parser.add_option('-c', '--coord_sys', action='store', dest='coord_sys', type='string',
                      default=coord_sys, metavar='Property',
                      help='Coordinate system specifier according to http://www.spatialreference.org. (default: epsg:31467)')
    parser.add_option('-f', '--header', action='store', dest='headerfile', type='string',
                      default=headerfile, metavar='Header file',
                      help='Header file containing information about e.g. number of rows and columns. (default: header.txt).')
    parser.add_option('-o', '--outfile', action='store', dest='outfile', type='string',
                      default=outfile, metavar='NetCDF file',
                      help='Name of NetCDF file. (default: PET_trimmed_1km.nc).')
    
    (opts, args) = parser.parse_args()
    
    headerfile         = opts.headerfile
    outfile            = opts.outfile   
    coord_sys          = opts.coord_sys 
    
    #############################################
    
    
    # check input files
    if not os.path.isfile(headerfile):
        sys.exit("No "+headerfile+" file found here, are you in the right directory?")

    # read header information
    header_info  = np.loadtxt( headerfile, dtype='|S20')
    ncols        = np.int(header_info[0,1])
    nrows        = np.int(header_info[1,1])
    xllcorner    = np.float(header_info[2,1])
    yllcorner    = np.float(header_info[3,1])
    cs           = np.float(header_info[4,1])
    NODATA_value = header_info[5,1]

    # create x and y grid
    xx          = np.arange( xllcorner + cs/2,            xllcorner + cs/2 + ncols*cs, cs)
    yy          = np.arange( yllcorner + cs/2 + nrows*cs, yllcorner + cs/2,-cs)
    xx, yy      = np.meshgrid(xx,yy)
    # Dimensions
    fhandle      = nc.Dataset(outfile, 'w', format='NETCDF4')
    #
    # determine latitude and longitude of the Aimgrid
    projAim    = Proj(init=coord_sys)
    lons, lats = projAim(xx, yy, inverse=True)
    FiAtt  = ([['description', 'Inputs and Outputs for TSEB'],
           ['history','Created ' + time.ctime(time.time()) ]])
    writenetcdf(fhandle, fileattributes=FiAtt)
    
    startTime    = ('days since '  + str(startyear-1) + '-' + str(12).zfill(2) + '-' + str(31).zfill(2) + ' ' 
                      + str(23) + ':' + str(0).zfill(2) + ':' + str(0).zfill(2))  
    
    
    nrows=nrows+1               #no idea why i have to do this.....
#    varName = 'time'
#    dims    = len(times)
#    
#    var     = np.zeros(1)+doy
#    varAtt  = ([['units'    , startTime],
#                ['calender day' , 'standard']])
#    writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, isdim=True, vartype='i4')
    # write netCDF
    #
    varAtt  = ([['axis'     , 'X']])
    varName = 'xc'
    dims    = ncols
    var     = xx[0,:] #np.arange(ncols)+1
    writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, isdim=True)
    
    
    #
    varAtt  = ([['axis'     , 'Y']])
    varName = 'yc'
    dims    = nrows
    var     = yy[:,0] #np.arange(nrows)+1
    
    
    
    writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, isdim=True)
    # Variables
    # lon
    varAtt  = ([['units'        , 'degrees_east'],
               ['long_name'     , 'longitude']])
    #
    varName = 'lon'
    var     = lons
    dims    = ['yc','xc']
    
      
    
    writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f8', comp=True)
    # lat
    varAtt  = ([['units'        , 'degrees_north'],
               ['long_name'     , 'latitude']])
    varName = 'lat'
    var     = lats
    dims    = ['yc','xc']
    
    
    
    writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f8', comp=True)
    #
    
    
    
    
    
    
    varAtt  = ([['long_name'     , 'LST'],
            ['units'         , 'K'],
            #['NoData' , 123.456], 
            ['coordinates'   , 'lon lat']])
    varName = 'LST'
    LST=Dictionary['T_R'][:]
    var     = Dictionary['T_R'][:]
    
    
    
    
    
    #var[np.isnan(var)]=123.456
    var=np.float64(var)
    dims    = ['yc','xc']
    
    
   
    
    writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f8', comp=True)
    
    
    varAtt  = ([['long_name'     , 'View Zenith Angle (degres)'],
            ['units'         , 'grades'],
            #['NoData' , 123.456], 
            ['coordinates'   , 'lon lat']])
    varName = 'VZA'
    var_theta     = Dictionary['theta1'][:]
    var_theta=np.float64(var_theta)
    var_theta[np.isnan(LST)]=np.nan
    dims    = ['yc','xc']
    writenetcdf(fhandle, name=varName, dims=dims, var=var_theta, attributes=varAtt, vartype='f8', comp=True)
    
    varAtt  = ([['long_name'     , 'Air Temperature'],
            ['units'         , 'K'],
            ['coordinates'   , 'lon lat']])
    
    varName = 'AirTemp'
    var_Airtemp     = Dictionary['T_A'][:]
    var_Airtemp=np.float64(var_Airtemp)
    var_Airtemp[np.isnan(LST)]=np.nan

    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var_Airtemp, attributes=varAtt, vartype='f4', comp=True)
    
    varAtt  = ([['long_name'     , 'Wind Speed'],
            ['units'         , 'm/s'],
            ['coordinates'   , 'lon lat']])
    varName = 'u'
    Dictionary['u'][np.isnan(LST)]=np.nan
    var     = Dictionary['u'][:]
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
    varAtt  = ([['long_name'     , 'Water vapour pressure above the canopy'],
            ['units'         , 'mb'],
            ['coordinates'   , 'lon lat']])
    varName = 'ea'
    var     = Dictionary['ea'][:]
    var=np.float64(var)
    var[np.isnan(LST)]=np.nan
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
    varAtt  = ([['long_name'     , 'Long Wave Radiation Downwards'],
            ['units'         , 'w/m2'],
            ['coordinates'   , 'lon lat']])
    varName = 'LWin'
    var     = Dictionary['LWin'][:]
    var=np.float64(var)
    var[np.isnan(LST)]=np.nan
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
    varAtt  = ([['long_name'     , 'Sun Zenith Angle from Aqua'],
            ['units'         , 'Degrees'],
            ['coordinates'   , 'lon lat']])
    varName = 'sza'
    var     = Dictionary['sza'][:]
    var=np.float64(var)
    var[np.isnan(LST)]=np.nan
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
    varAtt  = ([['long_name'     , 'Leaf Area Index interpolated betweeen dates. '],
            ['units'         , '[-]'],
            ['coordinates'   , 'lon lat']])
    varName = 'LAI'
    var     = Dictionary['LAI'][:]
    var=np.float64(var)
    #var[np.isnan(LST)]=np.nan
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
    varAtt  = ([['long_name'     , 'IGBP MODIS land Cover classification '],
            ['units'         , '[-]'],
            ['coordinates'   , 'lon lat']])
    varName = 'lc_IGBP'
    var     = Dictionary['lc_IGBP'][:]
    var=np.float64(var)
    var[np.isnan(LST)]=np.nan
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
    varAtt  = ([['long_name'     , 'Surface Pressure'],
            ['units'         , 'mb'],
            ['coordinates'   , 'lon lat']])
    varName = 'Sp'
    var     = Dictionary['p1'][:]
    var=np.float64(var)
    var[np.isnan(LST)]=np.nan
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
    varAtt  = ([['long_name'     , 'Height of canopy based on LAI'],
            ['units'         , 'm'],
            ['coordinates'   , 'lon lat']])
    varName = 'hc'
    #Dictionary['hc'][np.isnan(LST)]=np.nan
    var     = Dictionary['hc'][:]
    var=np.float64(var)
    var[np.isnan(LST)]=np.nan
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
    varAtt  = ([['long_name'     , 'Shortwave radiation downwards'],
            ['units'         , 'w/m2'],
            ['coordinates'   , 'lon lat']])
    varName = 'Sdn'
    #Dictionary['Sdn'][np.isnan(LST)]=np.nan
    var     = Dictionary['Sdn'][:]
    var=np.float64(var)
    var[np.isnan(LST)]=np.nan    
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
    varAtt  = ([['long_name'     , 'Net Radiation Soil'],
            ['units'         , 'w/m2'],
            ['coordinates'   , 'lon lat']])
    varName = 'Rn_sw_soil'
    #Dictionary['Rn_sw_soil'][np.isnan(LST)]=np.nan
    var     = Dictionary['Rn_sw_soil'][:]
    var=np.float64(var)
    var[np.isnan(LST)]=np.nan
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
    varAtt  = ([['long_name'     , 'Net Radiation vegetation'],
            ['units'         , 'w/m2'],
            ['coordinates'   , 'lon lat']])
    varName = 'Rn_sw_veg'
   # Dictionary['Rn_sw_veg'][np.isnan(LST)]=np.nan
    var     = Dictionary['Rn_sw_veg'][:]
    var=np.float64(var)
    var[np.isnan(LST)]=np.nan
    
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
    varAtt  = ([['long_name'     , 'ZoverM'],
            ['units'         , 'xxxx'],
            ['coordinates'   , 'lon lat']])
    varName = 'z_0M'
    var     = Dictionary['z_0M'][:]
    var=np.float64(var)
    var[np.isnan(LST)]=np.nan
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)    
    
        
    varAtt  = ([['long_name'     , 'doverM'],
            ['units'         , 'xxxx'],
            ['coordinates'   , 'lon lat']])
    varName = 'd_0M'
    var     = Dictionary['d_0M'][:]
    var=np.float64(var)
    var[np.isnan(LST)]=np.nan
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
    varAtt  = ([['long_name'     , 'zetau'],
            ['units'         , 'xxxx'],
            ['coordinates'   , 'lon lat']])
    varName = 'zu'
    var     = Dictionary['zu'][:]
    var=np.float64(var)
    var[np.isnan(LST)]=np.nan
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
    varAtt  = ([['long_name'     , 'zetat'],
            ['units'         , 'xxxx'],
            ['coordinates'   , 'lon lat']])
    varName = 'zt'
    var     = Dictionary['zt'][:]
    var=np.float64(var)
    var[np.isnan(LST)]=np.nan
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
    varAtt  = ([['long_name'     , 'TSEB Flag'],
            ['units'         , '[-]'],
            ['coordinates'   , 'lon lat']])
    varName = 'Flag'
    flag[np.isnan(LST)]=np.nan
    var     = flag
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
    varAtt  = ([['long_name'     , 'TSEB Soil Temperature'],
            ['units'         , 'K'],
            ['coordinates'   , 'lon lat']])
    varName = 'SoilTemperature'
    Ts[np.isnan(LST)]=np.nan
    var     = Ts
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
    varAtt  = ([['long_name'     , 'TSEB Canopy Temperature'],
            ['units'         , 'K'],
            ['coordinates'   , 'lon lat']])
    varName = 'CanopyTemperature'
    Tc[np.isnan(LST)]=np.nan
    var     = Tc
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
    varAtt  = ([['long_name'     , 'TSEB Air temperature at the canopy interface (Kelvin)'],
            ['units'         , 'K'],
            ['coordinates'   , 'lon lat']])
    varName = 'T_AC'
    T_AC[np.isnan(LST)]=np.nan
    var     = T_AC
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
    
    varAtt  = ([['long_name'     , 'TSEB Soil net longwave radiation'],
            ['units'         , 'W m-2'],
            ['coordinates'   , 'lon lat']])
    varName = 'L_nC'
    L_nC[np.isnan(LST)]=np.nan
    var     = L_nC
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
    
    varAtt  = ([['long_name'     , 'TSEB Canopy net longwave radiation '],
            ['units'         , 'W m-2'],
            ['coordinates'   , 'lon lat']])
    varName = 'L_nS'
    L_nS[np.isnan(LST)]=np.nan
    var     = L_nS
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
    
    varAtt  = ([['long_name'     , 'TSEB Canopy latent heat flux '],
            ['units'         , 'W m-2'],
            ['coordinates'   , 'lon lat']])
    varName = 'LE_C'
    LE_C[np.isnan(LST)]=np.nan
    var     = LE_C
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
    
    varAtt  = ([['long_name'     , 'TSEB Canopy sensible heat flux '],
            ['units'         , 'W m-2'],
            ['coordinates'   , 'lon lat']])
    varName = 'H_C'
    H_C[np.isnan(LST)]=np.nan
    var     = H_C
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
    
    varAtt  = ([['long_name'     , 'TSEB Soil latent heat flux  '],
            ['units'         , 'W m-2'],
            ['coordinates'   , 'lon lat']])
    varName = 'LE_S'
    LE_S[np.isnan(LST)]=np.nan
    var     = LE_S
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
    
    varAtt  = ([['long_name'     , 'TSEB Soil sensible heat flux'],
            ['units'         , 'W m-2'],
            ['coordinates'   , 'lon lat']])
    varName = 'H_S'
    H_S[np.isnan(LST)]=np.nan
    var     = H_S
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
    varAtt  = ([['long_name'     , 'TSEB Soil heat flux '],
            ['units'         , 'W m-2'],
            ['coordinates'   , 'lon lat']])
    varName = 'G'
    G[np.isnan(LST)]=np.nan
    var     = G
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
    varAtt  = ([['long_name'     , 'TSEB Soil aerodynamic resistance to heat transport '],
            ['units'         , 's m-1'],
            ['coordinates'   , 'lon lat']])
    varName = 'R_s'
    R_s[np.isnan(LST)]=np.nan
    var     = R_s
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
    varAtt  = ([['long_name'     , 'TSEB Bulk canopy aerodynamic resistance to heat transport'],
            ['units'         , 's m-1'],
            ['coordinates'   , 'lon lat']])
    varName = 'R_x'
    R_x[np.isnan(LST)]=np.nan
    var     = R_x
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
    varAtt  = ([['long_name'     , 'TSEB Friction velocity '],
            ['units'         , 's m-1'],
            ['coordinates'   , 'lon lat']])
    varName = 'u_friction'
    u_friction[np.isnan(LST)]=np.nan
    var     = u_friction
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
    varAtt  = ([['long_name'     , 'TSEB Aerodynamic resistance to heat transport '],
            ['units'         , 's m-1'],
            ['coordinates'   , 'lon lat']])
    varName = 'R_a'
    R_a[np.isnan(LST)]=np.nan
    var     = R_a
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
    varAtt  = ([['long_name'     , 'TSEB Monin-Obuhkov length '],
            ['units'         , 'm'],
            ['coordinates'   , 'lon lat']])
    varName = 'L'
    L[np.isnan(LST)]=np.nan
    var     = L
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
#    varAtt  = ([['long_name'     , 'L_sky'],
#            ['units'         , 'wm2'],
#            ['coordinates'   , 'lon lat']])
#    varName = 'Lsky'
#    Lsky[np.isnan(LST)]=np.nan
#    var     = Lsky
#    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
#    


   
#    varAtt  = ([['long_name'     , 'TSEB number of iterations until convergence of L'],
#            ['units'         , 'm'],
#            ['coordinates'   , 'lon lat']])
#    varName = 'Iterations'
#    n_iterations[np.isnan(LST)]=np.nan
#    var     = n_iterations
#    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
    
    varAtt  = ([['long_name'     , 'Dummy (Dummyt variable. Not meaningful)']])
    varName = 'ZZZZZZ'
    var   = np.zeros([7,10])
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
    
    
    
    return  

global Dictionary
Dictionary=MakeTSEBDictionary()
Dictionary2=MakeTSEBDictionary2()
global L_nC2
global L_nS2
L_nC2 = np.zeros([7,10])
L_nS2 = np.zeros([7,10])


JulDay = 60
Year = 2003

def GetRestOfData(Dictionary, Year, JulDay):
    
    Path_d_0M ='E:\\HenrikTSEB\\Station_Data_Arrays\\d_0M\\'
    Path_emisGrd ='E:\\HenrikTSEB\\Station_Data_Arrays\\emisGrd\\'
    Path_emisVeg ='E:\\HenrikTSEB\\Station_Data_Arrays\\emisVeg\\'
    Path_Sdn ='E:\\HenrikTSEB\\Station_Data_Arrays\\Sdn\\'
    Path_z_0M ='E:\\HenrikTSEB\\Station_Data_Arrays\\z_0M\\'
    Path_zt ='E:\\HenrikTSEB\\Station_Data_Arrays\\zt\\'
    Path_zu ='E:\\HenrikTSEB\\Station_Data_Arrays\\zu\\'
    Path_hc ='E:\\HenrikTSEB\\Station_Data_Arrays\\hc\\'
    
    Path_alpha_PT ='E:\\HenrikTSEB\\Station_Data_Arrays\\alpha_PT\\'
    Path_leaf_width ='E:\\HenrikTSEB\\Station_Data_Arrays\\leaf_width\\'
    Path_z0_soil ='E:\\HenrikTSEB\\Station_Data_Arrays\\z0_soil\\'
    Path_wc ='E:\\HenrikTSEB\\Station_Data_Arrays\\wc\\'
    Path_x_LAD ='E:\\HenrikTSEB\\Station_Data_Arrays\\x_LAD\\'
    
    Path_rho_leaf_vis ='E:\\HenrikTSEB\\Station_Data_Arrays\\rho_leaf_vis\\'
    Path_tau_leaf_vis ='E:\\HenrikTSEB\\Station_Data_Arrays\\tau_leaf_vis\\'
    Path_rho_leaf_nir ='E:\\HenrikTSEB\\Station_Data_Arrays\\rho_leaf_nir\\'
    Path_tau_leaf_nir ='E:\\HenrikTSEB\\Station_Data_Arrays\\tau_leaf_nir\\'
    
    
    
    
    
    d_0M = np.load(Path_d_0M +str(Year) + '-' + str(JulDay) + '.npy')
    emisGrd = np.load(Path_emisGrd +str(Year) + '-' + str(JulDay) + '.npy')
    emisVeg = np.load(Path_emisVeg +str(Year) + '-' + str(JulDay) + '.npy')
    Sdn = np.load(Path_Sdn +str(Year) + '-' + str(JulDay) + '.npy')
    z_0M = np.load(Path_z_0M +str(Year) + '-' + str(JulDay) + '.npy')
    zt = np.load(Path_zt +str(Year) + '-' + str(JulDay) + '.npy')
    zu = np.load(Path_zu +str(Year) + '-' + str(JulDay) + '.npy')
    hc = np.load(Path_hc +str(Year) + '-' + str(JulDay) + '.npy')
    
    alpha_PT = np.load(Path_alpha_PT +str(Year) + '-' + str(JulDay) + '.npy')
    leaf_width = np.load(Path_leaf_width +str(Year) + '-' + str(JulDay) + '.npy')
    z0_soil = np.load(Path_z0_soil +str(Year) + '-' + str(JulDay) + '.npy')
    wc = np.load(Path_wc +str(Year) + '-' + str(JulDay) + '.npy')    
    x_LAD = np.load(Path_x_LAD +str(Year) + '-' + str(JulDay) + '.npy')
    
    rho_leaf_vis = np.load(Path_rho_leaf_vis +str(Year) + '-' + str(JulDay) + '.npy')
    tau_leaf_vis = np.load(Path_tau_leaf_vis +str(Year) + '-' + str(JulDay) + '.npy')
    rho_leaf_nir = np.load(Path_rho_leaf_nir +str(Year) + '-' + str(JulDay) + '.npy')
    tau_leaf_nir = np.load(Path_tau_leaf_nir +str(Year) + '-' + str(JulDay) + '.npy')
    
    
    
    
    
    Dictionary['d_0M'] = d_0M
    Dictionary['emisGrd'] = emisGrd
    Dictionary['emisVeg'] = emisVeg
    Dictionary['Sdn'] = Sdn
    Dictionary['z_0M'] = z_0M
    Dictionary['zt'] = zt
    Dictionary['zu'] = zu
    Dictionary['hc'] = hc
    
    Dictionary['alpha_PT'] = alpha_PT
    Dictionary['leaf_width'] = leaf_width
    Dictionary['z0_soil'] = z0_soil
    Dictionary['wc'] = wc
    Dictionary['x_LAD'] = x_LAD
    
    Dictionary['rho_leaf_vis'] = rho_leaf_vis
    Dictionary['tau_leaf_vis'] = tau_leaf_vis
    Dictionary['rho_leaf_nir'] = rho_leaf_nir
    Dictionary['tau_leaf_nir '] = tau_leaf_nir 
    
        
    return Dictionary    


#np.save("E:\\HenrikTSEB\\StationNames\\StaionID.npy" , StationID)
    test222 = []
def runMultiTSEB(Year1, JulDay1):           #this method runs TSEB for a single day. MAIN METHOD
    
    #Year1 = 2007
    #JulDay1 = 175
    
    global Dictionary
    global JulDay
    global Year
    
    JulDay = JulDay1                    #This is probably not needed.
    Year = Year1
    
    
    
    
    #Now we fill in the dictionary with all the data from the arrays...
    print('starting..'+str(Year)+'_'+str(JulDay))
    try:
        Dictionary=MakeTSEBDictionary()
        Dictionary2=MakeTSEBDictionary2()
        Dictionary=FillDicTrAndTheta(Dictionary,Year,JulDay) # DONE
        Dictionary=FillDicTairWaterVaporPressure(Dictionary,Year,JulDay) # DONE
        Dictionary=FillDicWindSpeedr(Dictionary,Year,JulDay) # DONE
        Dictionary=FillDicLwIn(Dictionary,Year,JulDay) # DONE
        Dictionary=FillSunZenithAngle(Dictionary,Year,JulDay) # DONE 
        Dictionary=GetInterpolatedLAI('Y:\\Gorka\\EUROPA\\Timesat\\Outputs\\LAI_Weight\\',Year,JulDay,Dictionary) #DONE
        Dictionary=GetLancover_IGBP_Gorka(Dictionary,Year, JulDay) # DONE
        Dictionary=FillDicPressure(Dictionary,Year,JulDay) # DONE
        Dictionary=GetShortWaveRatiation(Dictionary, Year, JulDay) #DONE
        Dictionary=GetFGDataFromTiff(Dictionary, JulDay)
        Dictionary=GetRestOfData(Dictionary, Year, JulDay)
    
        #test222 = Dictionary["LWin"]
        #Then we run TSEB using the dictionary we just filled.
        print('runningTSEB')
        flag, Ts, Tc, T_AC,L_nS,L_nC, LE_C,H_C,LE_S,H_S,G,R_s,R_x,R_a,u_friction, L,n_iterations = runTSEB(Dictionary)  #runs the TSEB code with the dictionary data.
        #test2 =  (LE_C + LE_S)      
        
        
        #landcoverTestMine = Dictionary['lc_IGBP']
        print("here")
    except: 
        #if this happens it means that we're probably missing an array for this day and we can't fill the dictionary.
        print('couldnt run TSEB, data is probably missing')
         
        #we set all the TSEB return values to -999 arrays so when we put it in excel we know that for this day data is missing/wrong.
        flag = np.full((7,10),-999)         
        Ts= np.full((7,10),-999)
        Tc= np.full((7,10),-999)
        T_AC= np.full((7,10),-999)
        L_nS= np.full((7,10),-999)
        L_nC= np.full((7,10),-999)
        LE_C= np.full((7,10),-999)
        H_C= np.full((7,10),-999)
        LE_S= np.full((7,10),-999)
        H_S= np.full((7,10),-999)
        G= np.full((7,10),-999)
        R_s= np.full((7,10),-999)
        R_x= np.full((7,10),-999)
        R_a= np.full((7,10),-999)
        u_friction= np.full((7,10),-999)
        L= np.full((7,10),-999)
        n_iterations= np.full((7,10),-999)
        
        Dictionary ['T_R']       =   np.full((7,10),-999)
        Dictionary ['theta1']    =   np.full((7,10),-999)
        Dictionary ['T_A']      =   np.full((7,10),-999)
        Dictionary ['u']        =   np.full((7,10),-999)
        Dictionary ['ea1']        =   np.full((7,10),-999)
        Dictionary ['p1']        =   np.full((7,10),-999)
        Dictionary ['Sdn'] =         np.full((7,10),-999)
        Dictionary ['Rn_sw_veg'] =   np.full((7,10),-999)
        Dictionary ['Rn_sw_soil']=   np.full((7,10),-999)
        Dictionary ['LWin']      =   np.full((7,10),-999)
        Dictionary ['Fg']        =   np.full((7,10),-999)
        Dictionary ['hc']        =   np.full((7,10),-999)
        Dictionary ['emisVeg']   =   np.full((7,10),-999)
        Dictionary ['emisGrd']   =   np.full((7,10),-999)
        Dictionary ['z_0M']      =   np.full((7,10),-999)
        Dictionary ['d_0M']      =   np.full((7,10),-999)
        Dictionary ['zu']        =   np.full((7,10),-999)
        Dictionary ['zt']        =   np.full((7,10),-999)
        Dictionary ['sza']        =   np.full((7,10),-999)
        Dictionary ['lc_IGBP']        =   np.full((7,10),-999)
        Dictionary ['LAI']        =   np.full((7,10),-999)  
        Dictionary ['FCover']        =   np.full((7,10),-999) 
        
    #Results_Array = Dictionary['T_R']
    #Results_Array[Results_Array==0]=-600
    
      
    
    #remove data where the HC is above 700
    Ts[H_C > 700] = -999
    Tc[H_C > 700] = -999
    T_AC[H_C > 700] = -999
    L_nS[H_C > 700] = -999
    L_nC[H_C > 700] = -999
    LE_C[H_C > 700] = -999
    LE_S[H_C > 700] = -999
    G[H_C > 700] = -999
    R_s[H_C > 700] = -999
    R_x[H_C > 700] = -999
    R_a[H_C > 700] = -999
    u_friction[H_C > 700] = -999
    L[H_C > 700] = -999
    H_S[H_C > 700] = -999
    H_C[H_C > 700] = -999
    
    
        
    global L_nC2
    global L_nS2
    global LE_C2
    global LE_S2
    global LC_S2
    global LE_S2
    global H_S2
    global H_C2
    global L_nS2
    global L_nC2
    global R_a2
    global R_s2
    global R_x2
    global G2
    global Ts2
    global Tc2
    
    
    
    R_a2 = R_a
    R_s2 = R_s
    R_x2 = R_x
    Ts2 = Ts
    Tc2 = Tc
    G2 = G
    LC_S2 = LE_C                        #cant remember why i did this...
    LE_S2 = LE_S
    H_S2 = H_S
    H_C2 = H_C
    L_nS2 = L_nS
    L_nC2 = L_nC
    
    
    
    
    #Now we store the data we got from running TSEB.
    print('writing to excel or array...')
    writeToExcel(JulDay, Year, False)          #writes the selected outputs into the excel file
    
    
    LandCoverVector=np.load('Y:\\Gorka\\EUROPA\\FluxnetSitesCoordinates\\OldFiles\\output_LandCover_stations.npy')
    LandCoverVectorMine=np.load('Y:\\Gorka\\EUROPA\\FluxnetSitesCoordinates\\output_LandCover_stations.npy')
    test=np.load('Y:\\Gorka\\EUROPA\\FluxnetSitesCoordinates\\OldFiles\\output_lc_IGBP.npy')
    
def writeToExcel(JulDay, Year, excel): #writes to either excel if "excel" = true, or to arrays if "excel" = false. We write 3 different variables to either 3 sheets in an excel file, or to 3 different folders holding .npy arrays.
    #JulDay = 14 
    #Year = 2003

    excel = False
    
    
    """
    * JulDay = The day to write
    * Year= the year to write
    * Excel = true/false write to excel or array. If true then we write data to excel. If false then we write data to an array for that year.
    
    """    
    
    if(excel==True):
        from openpyxl.reader.excel import load_workbook
        if(JulDay < 180):                       #doing this since we have 2 excel sheets for each year to make the program run faster
            workbookr = load_workbook(filename="E:\\HenrikTSEB\\TSEB_Station_Output_UsingSmallArrays"+str(Year)+"-1.xlsx")
        if(JulDay > 180):
            workbookr = load_workbook(filename="E:\\HenrikTSEB\\TSEB_Station_Output_UsingSmallArrays"+str(Year)+"-2.xlsx")    
              
        global LC_S2
        global LE_S2
        global H_S2
        global H_C2
        global L_nS2
        global L_nC2
        global R_a2
        global R_s2
        global R_x2
        global G2
        global Ts2
        global Tc2
        
        
        
        ColumnYearFirst = Year - 2003
        ColumnYearSecond = ColumnYearFirst * 365
        ColumnYearSecond = 0                        #doing this since we dont need this anymore since we have an excel for each year now.
        
        #---------------wrinting ET into excel in sheet 1.--------------------------
        
        sheet = workbookr.get_sheet_by_name('ET')
        
        ET_Array = np.zeros([7,10])
        ET_Array = LE_S2 + LC_S2
        ET_Array[np.isnan(ET_Array)]=0                                #get error if we write NaN values, so overwrite with 0s?
        
        
        sheet.cell(row=1, column=JulDay + ColumnYearSecond).value = str(Year) + '-' + str(JulDay)
        
            
        wolo=-1
        count=-0
        for column in range(0,10):
                    
            wolo=wolo+1
            for row in range(0,7):
                count=count+1
                sheet.cell(row=count+1, column=JulDay + ColumnYearSecond).value = ET_Array[row, column]  
                    
        #-------------------writing LST into sheet 2---------------------
        sheet = workbookr.get_sheet_by_name('LST')
        
        sheet.cell(row=1, column=JulDay + ColumnYearSecond).value = str(Year) + '-' + str(JulDay)
        
        
        LST_Array = np.zeros([7,10])
        LST_Array = Dictionary['T_R']
        LST_Array[np.isnan(LST_Array)]=0
        
        wolo=-1
        count=-0
        for column in range(0,10):
                    
            wolo=wolo+1
            for row in range(0,7):
                count=count+1
                sheet.cell(row=count+1, column=JulDay + ColumnYearSecond).value = LST_Array[row, column]
        
        
        #-----------------------------------writing air temp into sheet 3---------------------
        sheet = workbookr.get_sheet_by_name('AirTemp')
        
        sheet.cell(row=1, column=JulDay + ColumnYearSecond).value = str(Year) + '-' + str(JulDay)
        
        AirTemp_Array = np.zeros([7,10])
        AirTemp_Array = Dictionary['T_A']
        AirTemp_Array[np.isnan(AirTemp_Array)]=0
        
        wolo=-1
        count=-0
        for column in range(0,10):
                    
            wolo=wolo+1
            for row in range(0,7):
                count=count+1
                sheet.cell(row=count+1, column=JulDay + ColumnYearSecond).value = AirTemp_Array[row, column]
        
        
        
        if(JulDay < 180): 
            workbookr.save("E:\\HenrikTSEB\\TSEB_Station_Output_UsingSmallArrays"+str(Year)+"-1.xlsx")
        if(JulDay > 180):
          workbookr.save("E:\\HenrikTSEB\\TSEB_Station_Output_UsingSmallArrays"+str(Year)+"-2.xlsx") 
#______________________________________________________________________________________________________________________________________________________ 
#................................................................................................................................................................
#...............................................................................................................................................................          
#_______________________________________________________________________________________________________________________________________________________          
          
    if(excel==False):  #if we want to write to arrays instead of excel
        print("excel false")
        Path_ET = 'E:\\HenrikTSEB\\Station_Result_Arrays_ET\\'
        
        Results_Array = np.load(Path_ET +str(Year) + '.npy')
        
        #initializing parameters. This is not the correct way to do it... FIX
        global LC_S2
        global LE_S2
        global H_S2
        global H_C2
        global L_nS2
        global L_nC2
        global R_a2
        global R_s2
        global R_x2
        global G2
        global Ts2
        global Tc2
        
        ColumnYearFirst = Year - 2003
        ColumnYearSecond = ColumnYearFirst * 365
        ColumnYearSecond = 0                        #doing this since we dont need this anymore since we have an excel for each year now.
        
        ET_Array = np.zeros([7,10])
        ET_Array = LE_S2 + LC_S2
        ET_Array[np.isnan(ET_Array)]=-999                                #get error if we write NaN values, so overwrite with 0s?
        #ET_Array[ET_Array == 0] = -999
          
        wolo=-1
        count=0 
        for column in range(0,10):                                                  #now we iterate through the array we just got by running TSEB and save it in a new 7*10 array.
                    
            wolo=wolo+1
            for row in range(0,7):
                count=count+1                                                       #count keeps track of the station number here.
                Results_Array[count,JulDay-1] = ET_Array[row, column]               #-1 from JulDay to fix problem in excel of only getting 364 days.
                 
        np.save(Path_ET +str(Year) , Results_Array)
        
        #----------------------------------------------------------------------------------------------------
        
        Path_LST = 'E:\\HenrikTSEB\\Station_Result_Arrays_LST\\'
        
        Results_Array = np.load(Path_LST +str(Year) + '.npy')
        
        LST_Array = np.zeros([7,10])
        LST_Array = Dictionary['T_R']
        #LST_Array[LST_Array==0] =-999
        #LST_Array[LST_Array==0.000] =-999
        LST_Array[np.isnan(LST_Array)]=-999                                #get error if we write NaN values, so overwrite with -999s?
        
        
        wolo=-1
        count=-0
        for column in range(0,10):
                    
            wolo=wolo+1
            for row in range(0,7):
                count=count+1
                Results_Array[count,JulDay-1] = LST_Array[row, column]
                     
        np.save(Path_LST +str(Year) , Results_Array)
        
        #-------------------------------------------------------------------------------------------------------
        
        Path_AirTemp = 'E:\\HenrikTSEB\\Station_Result_Arrays_AirTemp\\'
        
        Results_Array = np.load(Path_AirTemp +str(Year) + '.npy')
        
        AirTemp_Array = np.zeros([7,10])
        AirTemp_Array = Dictionary['T_A']
        AirTemp_Array[np.isnan(AirTemp_Array)]=-999                               #get error if we write NaN values, so overwrite with 0s?
        #AirTemp_Array[AirTemp_Array == 0] = -999
        
        wolo=-1
        count=-0
        for column in range(0,10):
                    
            wolo=wolo+1
            for row in range(0,7):
                count=count+1
                Results_Array[count,JulDay-1] = AirTemp_Array[row, column]
                    
        np.save(Path_AirTemp +str(Year) , Results_Array)
        
        #-----------------------------------------------------------------------------------------------
        
        Path_HC = 'E:\\HenrikTSEB\\Station_Result_Arrays_HC\\'
    
        
        Results_Array = np.load(Path_HC +str(Year) + '.npy')
        
                      
                            
        HC_Array = np.zeros([7,10])
        HC_Array = H_S2 + H_C2
        HC_Array[np.isnan(HC_Array)]=-999                                #get error if we write NaN values, so overwrite with 0s?
        HC_Array[HC_Array > 700]=-999
        #HC_Array[HC_Array == 0] = -999
        
        
            
        wolo=-1
        count=-0
        for column in range(0,10):
                    
            wolo=wolo+1
            for row in range(0,7):
                count=count+1
                Results_Array[count,JulDay-1] = HC_Array[row, column]
                
       
        np.save(Path_HC +str(Year) , Results_Array)
        
        #------------------------------------------------------------------------------------------------
        Path_Sn = 'E:\\HenrikTSEB\\Station_Result_Arrays_Sn\\'
    
        
        Results_Array = np.load(Path_Sn +str(Year) + '.npy') 
        
        Sn_Array = np.zeros([7,10])
        Sn_Array = Dictionary['Rn_sw_soil'] + Dictionary['Rn_sw_veg']
        Sn_Array[np.isnan(Sn_Array)]=-999                                #get error if we write NaN values, so overwrite with 0s?
        #Sn_Array[Sn_Array == 0] = -999
        
            
        wolo=-1
        count=-0
        for column in range(0,10):
                    
            wolo=wolo+1
            for row in range(0,7):
                count=count+1
                Results_Array[count,JulDay-1] = Sn_Array[row, column]
                
       
        np.save(Path_Sn +str(Year) , Results_Array)
        
        #-----------------------------------------------------------------------------------------------
        Path_Ln = 'E:\\HenrikTSEB\\Station_Result_Arrays_Ln\\'
    
        
        Results_Array = np.load(Path_Ln +str(Year) + '.npy')
        
        Ln_Array = np.zeros([7,10])
        Ln_Array = L_nS2 + L_nC2
        Ln_Array[np.isnan(Ln_Array)]=-999                                #get error if we write NaN values, so overwrite with 0s?
        #Ln_Array[Ln_Array == 0] = -999
        
             
        wolo=-1
        count=-0
        for column in range(0,10):
                    
            wolo=wolo+1
            for row in range(0,7):
                count=count+1
                Results_Array[count,JulDay-1] = Ln_Array[row, column]
                
    
        np.save(Path_Ln +str(Year) , Results_Array)
                
       #---------------------------------------------------------------------------------------------         
        Path_R_a = 'E:\\HenrikTSEB\\Station_Result_Arrays_R_a\\'
    
        
        Results_Array = np.load(Path_R_a +str(Year) + '.npy')
        
                              

        R_a_Array = np.zeros([7,10])
        R_a_Array = R_a2
        R_a_Array[np.isnan(R_a_Array)]=-999                                #get error if we write NaN values, so overwrite with 0s?
        #R_a_Array[R_a_Array == 0] = -999
        
        
            
        wolo=-1
        count=-0
        for column in range(0,10):
                    
            wolo=wolo+1
            for row in range(0,7):
                count=count+1
                Results_Array[count,JulDay-1] = R_a_Array[row, column]         
                
        np.save(Path_R_a +str(Year) , Results_Array)
        
        #------------------------------------------------------------------------
        Path_R_s = 'E:\\HenrikTSEB\\Station_Result_Arrays_R_s\\'
    
        
        Results_Array = np.load(Path_R_s +str(Year) + '.npy')
        
                              

        R_s_Array = np.zeros([7,10])
        R_s_Array = R_s2
        R_s_Array[np.isnan(R_s_Array)]=-999                                #get error if we write NaN values, so overwrite with 0s?
        #R_s_Array[R_s_Array == 0] = -999
        
        
            
        wolo=-1
        count=-0
        for column in range(0,10):
                    
            wolo=wolo+1
            for row in range(0,7):
                count=count+1
                Results_Array[count,JulDay-1] = R_s_Array[row, column]         
                
        np.save(Path_R_s +str(Year) , Results_Array)
        
        #-----------------------------------------------------------------------------
        Path_R_x = 'E:\\HenrikTSEB\\Station_Result_Arrays_R_x\\'
    
        
        Results_Array = np.load(Path_R_x +str(Year) + '.npy')
        
                              

        R_x_Array = np.zeros([7,10])
        R_x_Array = R_x2
        R_x_Array[np.isnan(R_x_Array)]=-999                                #get error if we write NaN values, so overwrite with 0s?
        #R_x_Array[R_x_Array == 0] = -999
        
        
            
        wolo=-1
        count=-0
        for column in range(0,10):
                    
            wolo=wolo+1
            for row in range(0,7):
                count=count+1
                Results_Array[count,JulDay-1] = R_x_Array[row, column]         
                
        np.save(Path_R_x +str(Year) , Results_Array)
        
        #---------------------------------------------------------------------
        Path_G = 'E:\\HenrikTSEB\\Station_Result_Arrays_G\\'
    
        
        Results_Array = np.load(Path_G +str(Year) + '.npy')
        
        R_G_Array = np.zeros([7,10])
        R_G_Array = G2
        R_G_Array[np.isnan(R_G_Array)]=-999                                #get error if we write NaN values, so overwrite with 0s?
        #R_G_Array[R_G_Array == 0] = -999
           
        wolo=-1
        count=-0
        for column in range(0,10):
                    
            wolo=wolo+1
            for row in range(0,7):
                count=count+1
                Results_Array[count,JulDay-1] = R_G_Array[row, column]         
                
        np.save(Path_G +str(Year) , Results_Array)
        
        
        #---------------------------------------------------------------------------------------------
        Path_Ts = 'E:\\HenrikTSEB\\Station_Result_Arrays_Ts\\'
    
        
        Results_Array = np.load(Path_Ts +str(Year) + '.npy')
        
        Ts_Array = np.zeros([7,10])
        Ts_Array = Ts2
        Ts_Array[np.isnan(Ts_Array)]=-999                                #get error if we write NaN values, so overwrite with 0s?
        #R_G_Array[R_G_Array == 0] = -999
        
        
            
        wolo=-1
        count=-0
        for column in range(0,10):
                    
            wolo=wolo+1
            for row in range(0,7):
                count=count+1
                Results_Array[count,JulDay-1] = Ts_Array[row, column]         
                
        np.save(Path_Ts +str(Year) , Results_Array)
        
        #-------------------------------------------------------------------------------------------------
        
        Path_Tc = 'E:\\HenrikTSEB\\Station_Result_Arrays_Tc\\'
    
        
        Results_Array = np.load(Path_Tc +str(Year) + '.npy')
        
        Tc_Array = np.zeros([7,10])
        Tc_Array = Tc2
        Tc_Array[np.isnan(Tc_Array)]=-999                                #get error if we write NaN values, so overwrite with 0s?
        #R_G_Array[R_G_Array == 0] = -999
           
        wolo=-1
        count=-0
        for column in range(0,10):
                    
            wolo=wolo+1
            for row in range(0,7):
                count=count+1
                Results_Array[count,JulDay-1] = Tc_Array[row, column]         
                
        np.save(Path_Tc +str(Year) , Results_Array)
    
startimeWindow=125
EndTimeWindow=155


reference_image='Y:\\Gorka\\EUROPA\\MYD11A1.006\\MYD11A1.0062002185.mosaic.tif'
#reference_image='F:\\LW_in_Subset\\LWin_2002001_12.tif'
metadata=getSceneMetadata(reference_image)
proj=metadata.get('proj')
geotransform=metadata.get('gt')


global Year
global JulDay

Year = 1
JulDay = 1


 


#Year=2010
#JulDay= 195
#print('starting..'+str(Year)+'_'+str(JulDay))
##    for JulDay in JulDays:
#Dictionary=MakeTSEBDictionary()
#Dictionary2=MakeTSEBDictionary2()
#Dictionary=FillDicTrAndTheta(Dictionary,Year,JulDay) #
#Dictionary=FillDicTairWaterVaporPressure(Dictionary,Year,JulDay) #
#Dictionary=FillDicWindSpeedr(Dictionary,Year,JulDay) #
#Dictionary=FillDicLwIn(Dictionary,Year,JulDay) #
#Dictionary=FillSunZenithAngle(Dictionary,Year,JulDay) #
#Dictionary=GetInterpolatedLAI('Y:\\Gorka\\EUROPA\\Timesat\\Outputs\\LAI_Weight\\',Year,JulDay,Dictionary) #
#Dictionary=GetLancover_IGBP_Gorka(Dictionary,Year) #
#Dictionary=FillDicPressure(Dictionary,Year,JulDay) #
#Dictionary=GetShortWaveRatiation(Dictionary)
#Dictionary=GetFGDataFromTiff(Dictionary, JulDay)
#
#
##Folder='Y:\\Gorka\\EUROPA\\Timesat\\Outputs\\LAI_Weight\\'
## #
##
##PruebaTA=Dictionary['T_A'][:]
##PruebaTR=Dictionary['T_R'][:]
##PruebaAngle=Dictionary['theta1'][:]
##Pruebau=Dictionary['u'][:]
##Pruebaea=Dictionary ['ea'][:]
##PruebaLWin=Dictionary ['LWin'][:]
##PruebaSunZenithAngle=Dictionary ['sza'][:]
##PruebaLAI=Dictionary ['LAI'][:]
##Pruebahc=Dictionary ['hc'][:]
##Pruebazu=Dictionary ['zu'][:]
##Pruebazt=Dictionary ['zt'][:]
##Pruebaz_0M=Dictionary ['z_0M'][:]
##Pruebad_0M=Dictionary ['d_0M'][:]
##Prueba_SP=Dictionary ['p1'][:]
##Prueba_Sdn=Dictionary ['Sdn'] 
##Prueba_Rn_sw_veg=Dictionary['Rn_sw_veg']
##Prueba_Rn_sw_soilg=Dictionary['Rn_sw_soil']
#
#print('runningTSEB')
#flag, Ts, Tc, T_AC,L_nS,L_nC, LE_C,H_C,LE_S,H_S,G,R_s,R_x,R_a,u_friction, L,n_iterations = runTSEB(Dictionary)
#
##print(Ts)
#
#print('writing to excel...')
#writeToExcel(JulDay)

#print('createdailynet')
#createDailyNetCDFfile4TSEB(Year,JulDay,Dictionary,flag, Ts, Tc, T_AC,L_nS,L_nC, LE_C,H_C,LE_S,H_S,G,R_s,R_x,R_a,u_friction, L)

#scriptfolder=os.path.dirname(__file__) 
#scriptfolder=scriptfolder+'/'
#scriptfolder=scriptfolder.replace('/','\\')        
#createoutputfolders(scriptfolder)        
#folderout=os.path.dirname(__file__)     
#folderout=folderout+'/outputs/'
#folderout=folderout.replace('/','\\')
#savefilesTSEB(folderout,Year,JulDay,geotransform,proj,flag, Ts, Tc, T_AC,L_nS,L_nC, LE_C,H_C,LE_S,H_S,G,R_s,R_x,R_a,u_friction, L,n_iterations)

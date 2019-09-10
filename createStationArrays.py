# -*- coding: utf-8 -*-
"""
Created on Fri May 19 12:50:38 2017

@author: gmg

T_R1
theta1
T_A1
u1
ea1
p1
Rn_sw_veg
Rn_sw_soil
LWin
Fg
hc
emisVeg
emisGrd
z_0M
d_0M
zu
zt
leaf_width=leaf_width
z0_soil=z0_soil
alpha_PT=alpha_PT
x_LAD=x_L

----------------------------------------------------------------------------------
This script creates the 7*10 arrays for all the data, used for running TSEB for the stations
----------------------------------------------------------------------------------

"""
from __future__ import print_function
import numpy as np
import numpy.ma as ma
from osgeo import gdal
import pyTSEB
import os
import multiprocessing
import pandas
from joblib import Parallel, delayed
import xlsxwriter
from openpyxl import workbook
from PIL import Image
from skimage import io

fileName='Y:\\Gorka\\EUROPA\\FluxnetSitesCoordinates\\FluxnetSitesCoordinates.xlsx'

ExcelFile=pandas.read_excel(fileName,sheetname=0)
Rows=ExcelFile.RowPos
Cols=ExcelFile.ColPos
IGBP=ExcelFile.IGBP
StationID=ExcelFile.SITE_ID
StationName=ExcelFile.SITE_NAME





def GetFGDataFromTiff(Dictionary, JulDay):
    
   im = io.imread('E:\\HenrikTSEB\\Fg_stations_only_output\\Fg_Image.mosaic.tif')           #reading the tiff file created in the FG_stations script.
   FgDoyArray = np.zeros([7, 10])                                                       
   TimeStep = JulDay/8    #this needs to round up!!                                                                   #Finding the correct timestep to read.
   FgDoyArray = im[:, :, TimeStep]                                                          #reading one of the 46 time steps into the array for this specific DoY.
   
   Dictionary ['Fg'] = FgDoyArray
   
   return Dictionary
   
    
    


def getSceneMetadata(reference_image):
    import gdal
    metadata = {'xsize':0 ,'ysize': 0, 'bands':0, 'gt':[], 'proj':''}
    # Get reprojection parameters from the first dataset
#    scene = '"'+scene+'":Grid:band1'
    inputImage = gdal.Open(reference_image,gdal.GA_ReadOnly)
    if inputImage:
        metadata['xsize'] = inputImage.RasterXSize
        metadata['ysize'] = inputImage.RasterYSize
        metadata['bands'] = inputImage.RasterCount
        metadata['gt'] = inputImage.GetGeoTransform()
        metadata['proj'] = inputImage.GetProjection()
        inputImage = None
    return metadata
    
def readValueArray(array,row,col,ID):           #reads data from an array by row and column. Used for reading stations.
    
    import netCDF4
    import numpy as np
    
    
    
    value = array[row,col]   
    
    
    return value
    
    
def extractStationsFromMap(Map):                        #gets the data for just the stations out of a 3600x3600 array
    station_list_data=np.zeros([len(StationID),46])
    globalMaxLaiArray_Out_put=np.zeros([len(StationID)])
    LandCoverVector=np.zeros([len(StationID)])          #creates array of 60
    ID_all= [''  for x in xrange(len(Rows))]
    IGBP_all= [''  for x in xrange(len(Rows))]
    i=0
    


    for Station in range(0,len(Rows)):
           
        
        #testRead = readValueArray(ea,2311,2914,'DE-Geb')
        #testtt=ea[2336,2567]
        
        #testttii=ea[2336,2567]
        
        
        Row=Rows[Station]
        Col=Cols[Station]
        ID=StationID[Station]
        IGBP_Class=IGBP[Station]
        TestRead=readValueArray(Map,Row,Col,ID)
        #TestRead[TestRead==255]=np.nan
        station_list_data[i,:]=TestRead
        #globalMaxLaiArray_Out_put[i]=globalMaxLaiArray[Row,Col]
        #print(Row,Col,LAI,ID)
        ID_all[i]=ID
        IGBP_all[i]=IGBP_Class
        
        #LandCoverVector[i]=landcover[Row,Col]
        
    #    output_LAI[:,i]=LAI
        
        i=i+1
    
    return station_list_data
      

def MakeListIntoMap(mlist):          #takes a list of stations and makes it into a 10x60 array like a map
    Map_array = np.zeros([7,10])
    countRow = 0
    
    for n in range(0,10):
        
        for o in range(0,7):
            
            if(countRow<61):
                Map_array[o,n]=mlist[countRow, 0]
                countRow=countRow+1

    return Map_array    
    
    

 
def savefilesTSEB(folder,year,doy,geotransform,proj,flag, Ts, Tc, T_AC,L_nS,L_nC, LE_C,H_C,LE_S,H_S,G,R_s,R_x,R_a,u_friction, L,n_iterations):
    import numpy as np
    
    stryear=str(year)
    if doy < 100:
        if doy<10:
            strdoy='00'+str(doy)
        else:
            strdoy='0'+str(doy)
    else:
         strdoy=str(doy)   
    yeardoy= stryear+ strdoy  
    outputfile=folder+'flag\\Flag_'+yeardoy+'.tif'
    Save_array_tiff(flag, outputfile,geotransform, proj)
    outputfile=folder+'Ts\\Ts_'+yeardoy+'.tif'
    Ts[flag==255]=np.nan
    #Ts[Ts==0]=np.nan
    Save_array_tiff(Ts, outputfile,geotransform, proj)
    outputfile=folder+'Tc\\Tc_'+yeardoy+'.tif'
    Tc[flag==255]=np.nan
    #Tc[Tc==0]=np.nan
    Save_array_tiff(Tc, outputfile,geotransform, proj)
    outputfile=folder+'T_AC\\T_AC_'+yeardoy+'.tif'
    T_AC[flag==255]=np.nan
    #T_AC[T_AC==0]=np.nan
    Save_array_tiff(T_AC, outputfile,geotransform, proj)
    outputfile=folder+'L_nS\\L_nS_'+yeardoy+'.tif'
    L_nS[flag==255]=np.nan
    #L_nS[L_nS==0]=np.nan
    Save_array_tiff(L_nS, outputfile,geotransform, proj)
    outputfile=folder+'L_nC\\L_nC_'+yeardoy+'.tif'
    L_nC[flag==255]=np.nan
    #L_nC[L_nC==0]=np.nan
    Save_array_tiff(L_nC, outputfile,geotransform, proj)
    outputfile=folder+'LE_C\\LE_C_'+yeardoy+'.tif'
    LE_C[flag==255]=np.nan
    LE_C[LE_C<=0]=np.nan
    Save_array_tiff(LE_C, outputfile,geotransform, proj)
    outputfile=folder+'H_C\\H_C_'+yeardoy+'.tif'
    H_C[flag==255]=np.nan
    H_C[H_C<=0]=np.nan
    Save_array_tiff(H_C, outputfile,geotransform, proj)
    outputfile=folder+'LE_S\\LE_S_'+yeardoy+'.tif'
    LE_S[flag==255]=np.nan
    LE_S[LE_S<=0]=np.nan
    Save_array_tiff(LE_S, outputfile,geotransform, proj)
    outputfile=folder+'H_S\\H_S_'+yeardoy+'.tif'
    H_S[flag==255]=np.nan
    H_S[H_S<=0]=np.nan
    Save_array_tiff(H_S, outputfile,geotransform, proj)
    outputfile=folder+'G\\G_'+yeardoy+'.tif'
    G[flag==255]=np.nan
    G[G<=0]=np.nan
    Save_array_tiff(G, outputfile,geotransform, proj)
    outputfile=folder+'R_s\\R_s_'+yeardoy+'.tif'
    R_s[flag==255]=np.nan
    #R_s[R_s==0]=np.nan
    Save_array_tiff(R_s, outputfile,geotransform, proj)
    outputfile=folder+'R_x\\R_x_'+yeardoy+'.tif'
    R_x[flag==255]=np.nan
    #R_x[R_x==0]=np.nan
    Save_array_tiff(R_x, outputfile,geotransform, proj)
    outputfile=folder+'R_a\\R_a_'+yeardoy+'.tif'
    R_a[flag==255]=np.nan
    #R_a[R_a==0]=np.nan
    Save_array_tiff(R_a, outputfile,geotransform, proj)
    outputfile=folder+'u_friction\\u_frictions_'+yeardoy+'.tif'
    u_friction[flag==255]=np.nan
    u_friction[u_friction==0]=np.nan
    Save_array_tiff(u_friction, outputfile,geotransform, proj)
    outputfile=folder+'L\\L_'+yeardoy+'.tif'
    L[flag==255]=np.nan
    #L[L<=0]=np.nan
    Save_array_tiff(L, outputfile,geotransform, proj)
#    outputfile=folder+'n_iterations\\n_iterations_'+yeardoy+'.tif'
#    Save_array_tiff(n_iterations, outputfile,geotransform, proj)
    LE_Total=LE_C+LE_S
    LE_Total[LE_Total<=0]=np.nan
    outputfile=folder+'LE_Total\\'+yeardoy+'_LE_EUROPE.tif'
    Save_array_tiff(LE_Total, outputfile,geotransform, proj)
 
 
#    outputfile=folder+'Landcover\\Landcover_Map.tif'
#    Save_array_tiff(landcover, outputfile,geotransform, proj)
#    landcover
def Save_array_tiff(array, outputfile,geotransform, proj):
 
 
    from osgeo import gdal, osr
 
    # Start the gdal driver for GeoTIFF
    driver = gdal.GetDriverByName("GTiff")
 
    # Write array
    shape=array.shape
    print(len(shape))
 
    if len(shape) > 2:
        ds = driver.Create(outputfile, shape[1], shape[0], shape[2], gdal.GDT_Float32)
        ds.SetProjection(proj)
        ds.SetGeoTransform(geotransform)
 
        for i in range(shape[2]):
            print(i)
            ds.GetRasterBand(i+1).WriteArray(array[:,:,i])
 
    else:
        ds = driver.Create(outputfile, shape[1], shape[0], 1, gdal.GDT_Float32)
        ds.SetProjection(proj)
        ds.SetGeoTransform(geotransform)
        ds.GetRasterBand(1).WriteArray(array)
 
    ds=None
 
    print('Saved ' +outputfile )
 
    return


def createoutputfolders(folder):
    if not os.path.exists(folder+'outputs\\flag\\'):
        newpath=folder+'outputs\\flag'
        os.makedirs(newpath)
    if not os.path.exists(folder+'outputs\\G'):
        newpath=folder+'outputs\\G'
        os.makedirs(newpath)
    if not os.path.exists(folder+'outputs\\H_C'):
        newpath=folder+'outputs\\H_C'
        os.makedirs(newpath)
    if not os.path.exists(folder+'outputs\\H_S'):
        newpath=folder+'outputs\\H_S'
        os.makedirs(newpath)
    if not os.path.exists(folder+'outputs\\L'):
        newpath=folder+'outputs\\L'
        os.makedirs(newpath)
    if not os.path.exists(folder+'outputs\\L_nC'):
        newpath=folder+'outputs\\L_nC'
        os.makedirs(newpath)
    if not os.path.exists(folder+'outputs\\L_nS'):
        newpath=folder+'outputs\\L_nS'
        os.makedirs(newpath)
    if not os.path.exists(folder+'outputs\\Landcover'):
        newpath=folder+'outputs\\Landcover'
        os.makedirs(newpath)
    if not os.path.exists(folder+'outputs\\LE_C'):
        newpath=folder+'outputs\\LE_C'
        os.makedirs(newpath)
    if not os.path.exists(folder+'outputs\\LE_S'):
        newpath=folder+'outputs\\LE_S'
        os.makedirs(newpath)
    if not os.path.exists(folder+'outputs\\LE_Total'):
        newpath=folder+'outputs\\LE_Total'
        os.makedirs(newpath)
    if not os.path.exists(folder+'outputs\\n_iterations'):
        newpath=folder+'outputs\\n_iterations'
        os.makedirs(newpath)
    if not os.path.exists(folder+'outputs\\R_a'):
        newpath=folder+'outputs\\R_a'
        os.makedirs(newpath)
    if not os.path.exists(folder+'outputs\\R_s'):
        newpath=folder+'outputs\\R_s'
        os.makedirs(newpath)
    if not os.path.exists(folder+'outputs\\R_x'):
        newpath=folder+'outputs\\R_x'
        os.makedirs(newpath)
    if not os.path.exists(folder+'outputs\\T_AC'):
        newpath=folder+'outputs\\T_AC'
        os.makedirs(newpath)
    if not os.path.exists(folder+'outputs\\Tc'):
        newpath=folder+'outputs\\Tc'
        os.makedirs(newpath)
    if not os.path.exists(folder+'outputs\\Ts'):
        newpath=folder+'outputs\\Ts'
        os.makedirs(newpath)
    if not os.path.exists(folder+'outputs\\u_friction'):
        newpath=folder+'outputs\\u_friction'
        os.makedirs(newpath)
        
def runTSEB(Dictionary):
    from TSEB import TSEB_PT
    Tr_K=Dictionary['T_R'][:]
    vza=Dictionary['theta1'][:] 
    Ta_K=Dictionary['T_A'][:] 
    vza=Dictionary['theta1'][:]
    u=Dictionary['u'][:]
    ea=Dictionary ['ea'][:]
    p=Dictionary ['p1'][:]
    Sn_C=Dictionary['Rn_sw_veg'][:] 
    Sn_S=Dictionary['Rn_sw_soil'][:] 
    Lsky=Dictionary ['LWin'][:]
    LAI=(Dictionary ['LAI'][:])
    #LAI[LAI==25.5]=0.0
    hc=Dictionary ['hc'][:]
    emisVeg=(Dictionary['emisVeg'][:])
    emisGrd=(Dictionary['emisGrd'][:])
    z_0M=Dictionary ['z_0M'][:]
    d_0=Dictionary ['d_0M'][:]
    zu=Dictionary ['zu'][:]
    zt=Dictionary ['zt'][:]
    alpha_PT=(Dictionary['alpha_PT'][:])
    leaf_width=(Dictionary['leaf_width'][:])
    z0_soil=(Dictionary['z0_soil'][:])
    wc=Dictionary['wc'][:]
    x_LAD=Dictionary['x_LAD'][:]
    fg=Dictionary['Fg'][:]
    
    #FG set to 1:
    #flag, Ts, Tc, T_AC,L_nS,L_nC, LE_C,H_C,LE_S,H_S,G,R_s,R_x,R_a,u_friction, L,n_iterations=TSEB_PT(Tr_K,vza,Ta_K,u,ea,p,Sn_C,Sn_S,Lsky,LAI,hc,emisVeg,emisGrd,z_0M,d_0,zu,zt,leaf_width=leaf_width,z0_soil=z0_soil,alpha_PT=alpha_PT,x_LAD=x_LAD,f_c=0.999990,f_g=0.999990,wc=wc,Resistance_flag=0,calcG_params=[[1],0.35], UseL=False)
    
    #using real FG:
    flag, Ts, Tc, T_AC,L_nS,L_nC, LE_C,H_C,LE_S,H_S,G,R_s,R_x,R_a,u_friction, L,n_iterations=TSEB_PT(Tr_K,vza,Ta_K,u,ea,p,Sn_C,Sn_S,Lsky,LAI,hc,emisVeg,emisGrd,z_0M,d_0,zu,zt,leaf_width=leaf_width,z0_soil=z0_soil,alpha_PT=alpha_PT,x_LAD=x_LAD,f_c=0.999990,f_g=fg,wc=wc,Resistance_flag=0,calcG_params=[[1],0.35], UseL=False)
    return flag, Ts, Tc, T_AC,L_nS,L_nC, LE_C,H_C,LE_S,H_S,G,R_s,R_x,R_a,u_friction, L,n_iterations
    
    
def GetShortWaveRatiation(Dictionary):
    from netRadiation import CalcDifuseRatio,CalcSnCampbell
    import glob
    import gdal
    import os
    import numpy as np
    sza=Dictionary2['sza']                      #we use dictionary 2 here because this function calls outside scripts where the arrays need to be 3600x3600
    press=Dictionary2['p1']
    LAI=Dictionary2['LAI']
    Path ='Y:\\Gorka\\EUROPA\\ERA_Europe_Radiation\\Outputs\\SSRD\\'
    List= None  ;
    if int(JulDay) <100:
        if int(JulDay) <10:
            strJulDay='00'+str(JulDay)
        else:
            strJulDay='0'+str(JulDay)
    else:
        strJulDay=str(JulDay)
    List = glob.glob(Path + 'SSRD_' +str(Year) + strJulDay + 'Simuated13.tif')
    if List:
        fid = gdal.Open(List[0],gdal.GA_ReadOnly)
        print('Reading the ShortWave incoming radiation file now. Please wait...!!!!')
        Sdn=fid.GetRasterBand(1).ReadAsArray()
        
        
        Sdn_stations=np.zeros([len(StationID),46])
        Sdn_stations=extractStationsFromMap(Sdn)
        
        Sdn_Map_array = np.zeros([7,10])
        
        Sdn_Map_array = MakeListIntoMap(Sdn_stations) 
        
        CreateStationArray(Year, JulDay, Sdn_Map_array, 'Sdn')
        
        
        
        Sdn_Map_array[np.isnan(Dictionary ['T_R'][:])]=np.nan
        Dictionary ['Sdn']=Sdn_Map_array #removo this line after solar is ok
        
        #Dictionary ['LWin']       =   Swin
    
    
    difvis,difnir,fvis,fnir=CalcDifuseRatio(Sdn,sza,press=press,SOLAR_CONSTANT=1360)
    Sdn_dif=((difvis*Sdn)+(difnir*Sdn))/2
    Sdn_dir=Sdn-Sdn_dif
    rho_leaf_vis=Dictionary2['rho_leaf_vis']
    tau_leaf_vis=Dictionary2['tau_leaf_vis']
    rho_leaf_nir=Dictionary2['rho_leaf_nir']
    tau_leaf_nir=Dictionary2['tau_leaf_nir']
    rsoilv=0.15
    rsoiln=0.4
    
    
    Rn_sw_veg,Rn_sw_soil=CalcSnCampbell (LAI, sza, Sdn_dir, Sdn_dif, fvis,fnir, rho_leaf_vis,
                                         tau_leaf_vis,rho_leaf_nir, tau_leaf_nir, rsoilv, rsoiln,                                         x_LAD=1, LAI_eff=None)
    
    
    ####################################################################################################
    Rn_sw_veg_stations=np.zeros([len(StationID),46])
    Rn_sw_veg_stations=extractStationsFromMap(Rn_sw_veg)
        
    Rn_sw_veg_Map_array = np.zeros([7,10])
        
    Rn_sw_veg_Map_array = MakeListIntoMap(Rn_sw_veg_stations) 
        
    Rn_sw_soil_stations=np.zeros([len(StationID),46])
    Rn_sw_soil_stations=extractStationsFromMap(Rn_sw_soil)
        
    Rn_sw_soil_Map_array = np.zeros([7,10])
        
    Rn_sw_soil_Map_array = MakeListIntoMap(Rn_sw_soil_stations)
        
    Dictionary['Rn_sw_veg']=  Rn_sw_veg_Map_array  #return DiffuseSwRad,DirectSwRad,Sdn
    Dictionary ['Rn_sw_soil']=   Rn_sw_soil_Map_array
    
    CreateStationArray(Year, JulDay, Rn_sw_veg_Map_array, 'Rn_sw_veg')
    CreateStationArray(Year, JulDay, Rn_sw_soil_Map_array, 'Rn_sw_soil')
    
    return Dictionary
    
    #fid = gdal.Open(List[0],gdal.GA_ReadOnly)
    #thy = fid.GetRasterBand(1).
    #fid.



def GetLancover_IGBP_Gorka(Dictionary,Year):
    import netRadiation
    import resistances
    import numpy as np
    import numpy.ma as ma
    from osgeo import gdal
    landcover_file='Y:\\Gorka\\EUROPA\\MCD12Q1\\MCD12Q1'+str(Year)+'-01-01'
    landcover=np.fromfile(landcover_file,dtype='uint8')
    landcover=np.reshape(landcover, [3600,3600])
    landcover[landcover>=17]=0
    landcover2=landcover
    
    landcover_stations=np.zeros([len(StationID),46])
    landcover_stations=extractStationsFromMap(landcover)
        
    landcover_Map_array = np.zeros([7,10])
        
    landcover_Map_array = MakeListIntoMap(landcover_stations)
    
    landcover=np.zeros([7,10])
    landcover=landcover_Map_array 

    CreateStationArray(Year, JulDay, landcover_Map_array, 'lc_IGBP')           
            
       
    
    LAI=Dictionary['LAI'][:]
    #LAI[LAI==255]=np.nan
    #LAI[np.isnan(Dictionary ['T_R'][:])]=np.nan
    
    
       # land cover specific data
#    
#
#    # IGBP classification
#    #                       Everg   Everg.  Decid.  Decid.                                                                          Crop
#    #                       Needle  Broad   Needle  Broad   Mixed   Closed  Open    Woody           Grass   Wet     Crop            Veg     Snow
#    #                 Water Forest  Forest  Forest  Forest  Forest  shrubs  shrubs  Savanna Savanna land    lands   lands  Urban    Mosaic  Ice     Baren   
#    landCoverClasses= [0,   1,      2,      3,      4,      5,      6,      7,      8,      9,      10,     11,     12,    13,      14,     15,     16 ];  # land cover class, from MOD12Q1 
#    C_h_C =           [0.0, 22.0,   17.0,   22.0,   22.0,   20.0,   5.5,    3.5,    10.0,   2.0,    1.5,    2.5,    2.0,    4.0,    2.0,   1.0,    1.0];  # max height
#    C_defaultLAI =    [0.01,7.0,    7.0,    6.0,    6.0,    5.0,    4.0,    3.0,    3.0,    4.0,    3.0,    3.0,    6.0,   2.0,     3.0,   0.01,   0.05];  # fully developed LAI
#    C_f_C =           [1.0, 0.5,    0.9,    0.5,    0.9,    0.7,    0.9,    0.9,    0.9,    1.0,    1.0,    1.0,    0.9,   1.0,     0.9,    1.0,    1.0];  # clumping factor, based on Houbourg 2009 paper
#    C_D =             [0.0, 3.5,    1.5,    3.5,    1.5,    2.5,    1.5,    1.5,    1.5,    1.0,    1.0,    1.0,    1.0,   0.0,     1.5,    0.0,    1.0];  # ration of canopy width to height, Hector's value slightly different to Houbourg 2009 
#    C_s =             [0.0, 0.05,   0.10,   0.05,   0.10,   0.07,   0.10,   0.1,    0.02,   0.02,   0.02,   0.02,   0.2,   0.0,     0.1,    0.0,    0.02]; # leaf size, based on Houbourg 2009 paper
#    C_z_T =           [2.0, 50.0,   50.0,   50.0,   50.0,   50.0,   50.0,   50.0,   50.0,   50.0,   50.0,   50.0,   50.0,  50.0,    50.0,   50.0,   50.0]; # modelled temp 50m above ground/canopy
#    C_z_U =           [10.0,30.0,   25.0,   30.0,   25.0,   30.0,   10.0,   10.0,   20.0,   13.0,   10.0,   10.0,   10.0,  15.0,    20.0,   10.0,   10.0]; # modelled wind 10m above ground/canopy
#    NDVIFracGreen=    [0.01, 1.0,      1.0,    1.0,   1.0,     1.0,   1.0,    1.0,    1.0,   1.0,     1.2,    1.0,     1.2,  1.0,      1.0,     1,     1.0]; # modelled wind 10m above ground/canopy
#    
    # IGBP classification
    #                       Everg   Everg.  Decid.  Decid.                                                                          Crop
    #                       Needle  Broad   Needle  Broad   Mixed   Closed  Open    Woody           Grass   Wet     Crop            Veg     Snow
    #                 Water Forest  Forest  Forest  Forest  Forest  shrubs  shrubs  Savanna Savanna land    lands   lands  Urban    Mosaic  Ice     Baren   
    landCoverClasses= [0,   1,      2,      3,      4,      5,      6,      7,      8,      9,      10,     11,     12,    13,      14,     15,     16 ];  # land cover class, from MOD12Q1 
    C_h_C =           [0.0, 20.0,   20.0,   20.0,   20.0,   20.0,   2.0,    2.0,    10.0,   3.0,    1.5,    1.5,    2,     2.0,     2.0,   0.5,    1.0];  # max height
    C_defaultLAI =    [0.01, 5.1,    6.1,   6.8,   7.3,    6.6,    6.6,    4.3,    5.0,   5.5,      5.3,    5.1,    3.5,   4.7,     5.1,   0.01,   4.5];  # fully developed LAI
    C_f_C =           [1.0, 0.5,    0.9,    0.5,    0.9,    0.7,    0.9,    0.9,    0.9,    1.0,    1.0,    1.0,    0.9,   1.0,     0.9,    1.0,    1.0];  # clumping factor, based on Houbourg 2009 paper
    C_D =             [0.0, 3.5,    1.5,    3.5,    1.5,    2.5,    1.5,    1.5,    1.5,    1.0,    1.0,    1.0,    1.0,   0.0,     1.5,    0.0,    1.0];  # ration of canopy width to height, Hector's value slightly different to Houbourg 2009 
    C_s =             [0.0, 0.05,   0.10,   0.05,   0.10,   0.07,   0.10,   0.1,    0.02,   0.02,   0.02,   0.02,   0.2,   0.0,     0.1,    0.0,    0.02]; # leaf size, based on Houbourg 2009 paper
    C_z_T =           [2.0, 50.0,   50.0,   50.0,   50.0,   50.0,   50.0,   50.0,   50.0,   50.0,   50.0,   50.0,   50.0,  50.0,    50.0,   50.0,   50.0]; # modelled temp 50m above ground/canopy
    C_z_U =           [10.0,30.0,   25.0,   30.0,   25.0,   30.0,   10.0,   10.0,   20.0,   13.0,   10.0,   10.0,   10.0,  15.0,    20.0,   10.0,   10.0]; # modelled wind 10m above ground/canopy
    NDVIFracGreen=    [0.01, 1.0,      1.0,    1.0,   1.0,     1.0,   1.0,    1.0,    1.0,   1.0,     1.2,    1.0,     1.0,  1.0,      1.0,     1.0,     1.0]; # modelled wind 10m above ground/canopy
    
    
    
    Dictionary['hc']=np.zeros(landcover.shape) 
    Dictionary['zu']=np.zeros(landcover.shape) 
    Dictionary['zt']=np.zeros(landcover.shape) 
    Dictionary['z_0M']=np.zeros(landcover.shape) 
    Dictionary['d_0M']=np.zeros(landcover.shape)
    
    Dictionary2['rho_leaf_vis']=np.zeros(landcover2.shape) +0.2
    Dictionary2['tau_leaf_vis']=np.zeros(landcover2.shape) +0.2
    Dictionary2['rho_leaf_nir']=np.zeros(landcover2.shape) +0.4
    Dictionary2['tau_leaf_nir']=np.zeros(landcover2.shape) +0.5
    
    Dictionary['rho_leaf_vis']=np.zeros(landcover.shape) +0.2
    Dictionary['tau_leaf_vis']=np.zeros(landcover.shape) +0.2
    Dictionary['rho_leaf_nir']=np.zeros(landcover.shape) +0.4
    Dictionary['tau_leaf_nir']=np.zeros(landcover.shape) +0.5
    Dictionary['emisVeg']=np.zeros(landcover.shape) +0.999
    Dictionary['emisGrd']=np.zeros(landcover.shape) +0.96
    Dictionary['alpha_PT']=np.zeros(landcover.shape) +1.26 #original 1.26
    Dictionary['leaf_width']=np.zeros(landcover.shape) +0.1
    Dictionary['z0_soil']=np.zeros(landcover.shape) +0.01
    Dictionary['wc']=np.zeros(landcover.shape) +0.1
    Dictionary['x_LAD']=np.zeros(landcover.shape) +0.1
    
    # Set the parameters for each present land cover class 
    for lcClass in np.unique(landcover[~np.isnan(landcover)]):
        lcPixels = np.where(landcover == lcClass)
        lcIndex = landCoverClasses.index(lcClass)
        
        # scale height with LAI for grasslands, savanas, croplands, etc.
        Dictionary['hc'][lcPixels] = 0.1*C_h_C[lcIndex] + 0.9*C_h_C[lcIndex]*np.minimum(LAI[lcPixels]/C_defaultLAI[lcIndex], 1.0)
        #LAI[np.isnan(Dictionary ['T_R'][:])]=np.nan
        Dictionary['zu'][lcPixels] =Dictionary['hc'][lcPixels]+12
        Dictionary['zt'][lcPixels] =Dictionary['hc'][lcPixels]+12
        Dictionary['z_0M'][lcPixels]=resistances.CalcZ_0M(Dictionary['hc'][lcPixels])
        Dictionary['d_0M'][lcPixels]=resistances.CalcZ_0H (Dictionary['z_0M'][lcPixels],kB=0)
    
   # data['lc_IGBP'] =   landcoverData   
    
    
    
    Dictionary['lc_IGBP'] =np.float64( landcover)
    
    CreateStationArray(Year, JulDay, Dictionary['d_0M'], 'd_0M')
    CreateStationArray(Year, JulDay, Dictionary['emisVeg'], 'emisVeg')
    CreateStationArray(Year, JulDay, Dictionary['emisGrd'], 'emisGrd')
    CreateStationArray(Year, JulDay, Dictionary['hc'], 'hc')
    CreateStationArray(Year, JulDay, Dictionary['z_0M'], 'z_0M')
    CreateStationArray(Year, JulDay, Dictionary['zt'], 'zt')
    CreateStationArray(Year, JulDay, Dictionary['zu'], 'zu')
    CreateStationArray(Year, JulDay, Dictionary['alpha_PT'], 'alpha_PT')
    CreateStationArray(Year, JulDay, Dictionary['leaf_width'], 'leaf_width')
    CreateStationArray(Year, JulDay, Dictionary['z0_soil'], 'z0_soil')
    CreateStationArray(Year, JulDay, Dictionary['wc'], 'wc')
    CreateStationArray(Year, JulDay, Dictionary['x_LAD'], 'x_LAD')
    
    CreateStationArray(Year, JulDay, Dictionary['rho_leaf_vis'], 'rho_leaf_vis')
    CreateStationArray(Year, JulDay, Dictionary['tau_leaf_vis'], 'tau_leaf_vis')
    CreateStationArray(Year, JulDay, Dictionary['rho_leaf_nir'], 'rho_leaf_nir')
    CreateStationArray(Year, JulDay, Dictionary['tau_leaf_nir'], 'tau_leaf_nir')
    
    
        
    return Dictionary



def GetInterpolatedLAI(Folder,Year,JulDay,Dictionary):
    
    
    
#    Dictionary              =   {}
#    Dictionary2 = {}
#    Dictionary ['T_R']       =   [] #Done
#    Dictionary ['theta1']    =   [] #Done
#    Dictionary ['T_A']      =   [] #Done
#    Dictionary ['u']        =   [] #Done
#    Dictionary ['ea1']        =   []
#    Dictionary ['p1']        =   []
#    Dictionary ['Sdn'] =         []
#    Dictionary ['Rn_sw_veg'] =   []
#    Dictionary ['Rn_sw_soil']=   []
#    Dictionary ['LWin']      =   []
#    Dictionary ['Fg']        =   []
#    Dictionary ['hc']        =   []
#    Dictionary ['emisVeg']   =   []
#    Dictionary ['emisGrd']   =   []
#    Dictionary ['z_0M']      =   []
#    Dictionary ['d_0M']      =   []
#    Dictionary ['zu']        =   []
#    Dictionary ['zt']        =   []
#    Dictionary ['sza']        =   []
#    Dictionary ['lc_IGBP']        =   []
#    Dictionary ['LAI']        =   []  
#    Dictionary ['FCover']        =   [] 
#    Folder = 'Y:\\Gorka\\EUROPA\\Timesat\\Outputs\\LAI_Weight\\'
#    Year = 2005
#    JulDay = 365
    
    print('Getting the LAI data....!!!!')
    import glob
    import numpy as np
    File1=None
    File2=None
    JulDayIni=None
    JulDayEnd=None
    Interpolated=None
    JulDay8=(  '001','009','017','025','033','041','049','057','065','073','081',
               '089','097','105','113','121','129','137','145','153','161','169',
               '177','185','193','201','209','217','225','233','241','249','257',
               '265','273','281','289','297','305','313','321','329','337','345',
               '353','361')
    JulDayIni=('001','009','017','025','033','041','049','057','065','073','081',
               '089','097','105','113','121','129','137','145','153','161','169',
               '177','185','193','201','209','217','225','233','241','249','257',
               '265','273','281','289','297','305','313','321','329','337','345',
               '353')
    JulDayEnd=('009','017','025','033','041','049','057','065','073','081','089',
               '097','105','113','121','129','137','145','153','161','169','177',
               '185','193','201','209','217','225','233','241','249','257','265',
               '273','281','289','297','305','313','321','329','337','345','353',
               '361')
    for i in range(0,len(JulDay8)):
        
        JulDayini=None
        JulDayend=None
        
        if JulDay == int(JulDay8[i]):
            JulDayFixed=JulDay8[i]
            print('No interpolation between dates needed. DOY is within the 8 MODIS step products!!!!')
            File1=glob.glob(Folder + 'MCD15A2.005' +str(Year)+JulDayFixed+'.Weightsmosaic.asc')
            array1=np.fromfile(File1[0],dtype='uint8')
            array1=np.reshape(array1, [3600,3600])/10.0
            #array1 = array1.astype(float)/10
            #array1[np.isnan(Dictionary ['T_R'][:])]=np.nan
            #array1[array1>=250]=np.nan
            array1[array1>=12.0]=0
            
            ###########################################
            
            array1_stations=np.zeros([len(StationID),46])
            array1_stations=extractStationsFromMap(array1)
        
            array1_Map_array = np.zeros([7,10])
        
            array1_Map_array = MakeListIntoMap(array1_stations)             
            
            Dictionary['LAI']=array1_Map_array
            
            Dictionary2['LAI']=array1
            
            CreateStationArray(Year, JulDay, array1_Map_array, 'LAI')
            
            return Dictionary
            
        if (JulDay > int(JulDayIni[i])) and (JulDay <int(JulDayEnd[i])):
            print('Interpolating between dates!!!!')
            JulDayini=JulDayIni[i]
            JulDayend=JulDayEnd[i]
            File1=glob.glob(Folder + 'MCD15A2.005' +str(Year)+JulDayini+'.Weightsmosaic.asc')
            File2=glob.glob(Folder + 'MCD15A2.005' +str(Year)+JulDayend+'.Weightsmosaic.asc')
        #Reads the files and calculates the linear interpolated value    
            array1=np.fromfile(File1[0],dtype='uint8')
            array1=np.reshape(array1, [3600,3600])/10.0
            #array1[array1>=250]=np.nan
            #array1 = array1.astype(float) /10 
            
            array2=np.fromfile(File2[0],dtype='uint8')
            array2=np.reshape(array2, [3600,3600])/10.0
            #array2[array2>=250]=np.nan
            #array2 = array2.astype(float)  /10
            
            Step=(array2-array1)/8.
            StepPos=JulDay-int(JulDayini)
            Interpolated=array1+(Step*StepPos)
            #Interpolated[np.isnan(Dictionary ['T_R'][:])]=np.nan
            Interpolated[Interpolated>=12.0]=0
            
            
########################################################            
            
            Interpolated_stations=np.zeros([len(StationID),46])
            Interpolated_stations=extractStationsFromMap(Interpolated)
        
            Interpolated_Map_array = np.zeros([7,10])
        
            Interpolated_Map_array = MakeListIntoMap(Interpolated_stations)            
            
            Dictionary['LAI']=Interpolated_Map_array
            Dictionary2['LAI']=Interpolated
            
            CreateStationArray(Year, JulDay, Interpolated_Map_array, 'LAI')
            
            return Dictionary
            
        if JulDay >=int('361'):
            print('Last winter days. Interpolating between dates!!!!')
            JulDayini=JulDayEnd[i]
            JulDayend='001'
            File1=glob.glob(Folder + 'MCD15A2.005' +str(Year)+JulDayini+'.Weightsmosaic.asc')
            File2=glob.glob(Folder + 'MCD15A2.005' +str(Year+1)+JulDayend+'.Weightsmosaic.asc')
        #Reads the files and calculates the linear interpolated value    
            array1=np.fromfile(File1[0],dtype='uint8')
            array1=np.reshape(array1, [3600,3600])/10.0
            #array1[array1>=250]=np.nan
           # array1 = array1.astype(float) /10 
            
            array2=np.fromfile(File2[0],dtype='uint8')
            array2=np.reshape(array2, [3600,3600])/10.0
            #array2[array2>=250]=np.nan
           # array2 = array2.astype(float)  /10
            
            Step=(array2-array1)/8.
            StepPos=JulDay-int(JulDayini)
            Interpolated=array1+(Step*StepPos)
            #Interpolated[np.isnan(Dictionary ['T_R'][:])]=np.nan
            Interpolated[Interpolated>=12.0]=0
            
            
            Interpolated_stations=np.zeros([len(StationID),46])
            Interpolated_stations=extractStationsFromMap(Interpolated)
        
            Interpolated_Map_array = np.zeros([7,10])
        
            Interpolated_Map_array = MakeListIntoMap(Interpolated_stations)            
            
            Dictionary['LAI']=Interpolated_Map_array
            Dictionary2['LAI']=Interpolated
            
            CreateStationArray(Year, JulDay, Interpolated_Map_array, 'LAI')
            
            return Dictionary
        
    
        
def Calc_VaporPressure(T_K):
    
    # Calculates the vapor pressure in hPa, required temperature in Kelvin
    from numpy import exp,zeros,shape
    ea=zeros(shape(T_K))-1
    T_C=T_K-273.15
    ea = 6.112 * exp((17.67*T_C)/(T_C + 243.5))
    return ea
    
def openEuropeanLSTfile(Filename):
    import os
    import gdal
    import numpy as np
    filename=os.path.split(Filename)
    print ('The file '+filename[1]+' was found. Extracting the information from the file...')
    fid = gdal.Open(Filename,gdal.GA_ReadOnly)
    LST_Dataset=np.zeros([fid.RasterXSize,fid.RasterYSize,fid.RasterCount])
    LST_Dataset[LST_Dataset==0]=np.nan
    for i in range(fid.RasterCount):
        print(filename[1]+' succesfully opened. Reading band '+str(i))
        LST_Dataset[:,:,i]=fid.GetRasterBand(i+1).ReadAsArray()
    
    return LST_Dataset
def MakeTSEBDictionary():
    Dictionary              =   {}
    Dictionary2 = {}
    Dictionary ['T_R']       =   [] #Done
    Dictionary ['theta1']    =   [] #Done
    Dictionary ['T_A']      =   [] #Done
    Dictionary ['u']        =   [] #Done
    Dictionary ['ea1']        =   []
    Dictionary ['p1']        =   []
    Dictionary ['Sdn'] =         []
    Dictionary ['Rn_sw_veg'] =   []
    Dictionary ['Rn_sw_soil']=   []
    Dictionary ['LWin']      =   []
    Dictionary ['Fg']        =   []
    Dictionary ['hc']        =   []
    Dictionary ['emisVeg']   =   []
    Dictionary ['emisGrd']   =   []
    Dictionary ['z_0M']      =   []
    Dictionary ['d_0M']      =   []
    Dictionary ['zu']        =   []
    Dictionary ['zt']        =   []
    Dictionary ['sza']        =   []
    Dictionary ['lc_IGBP']        =   []
    Dictionary ['LAI']        =   []  
    Dictionary ['FCover']        =   [] 
    
    
    
    Dictionary2=Dictionary                          #we make a copy of the Dictionary because we need a variables to be in the 3600x3600 format. So some functions will put the stations map array of 10x6 into Dictionary, and the full picture of 3600x3600 into Dictionary2.
    return Dictionary

def MakeTSEBDictionary2():
    Dictionary              =   {}
    Dictionary2 = {}
    Dictionary ['T_R']       =   [] #Done
    Dictionary ['theta1']    =   [] #Done
    Dictionary ['T_A']      =   [] #Done
    Dictionary ['u']        =   [] #Done
    Dictionary ['ea1']        =   []
    Dictionary ['p1']        =   []
    Dictionary ['Sdn'] =         []
    Dictionary ['Rn_sw_veg'] =   []
    Dictionary ['Rn_sw_soil']=   []
    Dictionary ['LWin']      =   []
    Dictionary ['Fg']        =   []
    Dictionary ['hc']        =   []
    Dictionary ['emisVeg']   =   []
    Dictionary ['emisGrd']   =   []
    Dictionary ['z_0M']      =   []
    Dictionary ['d_0M']      =   []
    Dictionary ['zu']        =   []
    Dictionary ['zt']        =   []
    Dictionary ['sza']        =   []
    Dictionary ['lc_IGBP']        =   []
    Dictionary ['LAI']        =   []  
    Dictionary ['FCover']        =   [] 
    
    
    
    Dictionary2=Dictionary                          #we make a copy of the Dictionary because we need a variables to be in the 3600x3600 format. So some functions will put the stations map array of 10x6 into Dictionary, and the full picture of 3600x3600 into Dictionary2.
    return Dictionary2
    
def generateAquaTerraLSTdataset11_13(LST_Dataset,startimeWindow,EndTimeWindow):
     #Time window is 11:00 to 13:00. Also input is hour*10. For example 110 and 130, rather than 11 or 13.
    from pymasker import Masker
    import numpy as np
    #The LST_Dataset must containg the LST,ViewTime,Angle and quality flag band
    #Band positions in the LST dataset are:
        #Band 0 LST
        #Band 1 Quality flags
        #Band 2 Day view time 
        #Band 3 Day View Angle
    masker=Masker(np.uint8(LST_Dataset[:,:,1]))
    LST=LST_Dataset[:,:,0]
    ViewZenithAngle=LST_Dataset[:,:,3]
    GooDdata=masker.get_mask(0,2,'00') #00 = LST produced, good quality, not necessary to examine detailed QA
    #https://lpdaac.usgs.gov/dataset_discovery/modis/modis_products_table/mod11a1
    GooDdata_AverageLess1k=masker.get_mask(6,2,'00')    
        #Masked values are those with value 1
    Time=LST_Dataset[:,:,2]
    idx=(Time>=startimeWindow) & (Time<=EndTimeWindow)
    Mask=GooDdata+GooDdata_AverageLess1k+idx
    LST[Mask<2]=np.nan
    Time[Mask<2]=np.nan
    ViewZenithAngle[Mask<2]=np.nan
    SensorObs=np.zeros([Time.shape[0],Time.shape[1]])
    SensorObs[Mask >=2]=1
    return LST,Time,ViewZenithAngle,SensorObs
    
    
def FillDicTrAndTheta(Dictionary,Year,JulDay):
    #Fill Dictionary['T_R'] &Dictionary['theta1']=[]
    import glob
    import gdal
    import os
    import numpy as np
    Path_Terra ='Y:\\Gorka\\EUROPA\\MOD11A1.006\\'
    Path_Aqua  ='Y:\\Gorka\\EUROPA\\MYD11A1.006\\'  
    
    #Find Terra Image
    List_Terra= None  ;     List_Aqua=None
    
    if int(JulDay) <100:
        if int(JulDay) <10:
            strJulDay='00'+str(JulDay)
        else:
            strJulDay='0'+str(JulDay)
    else:
        strJulDay=str(JulDay)
    List_Terra = glob.glob(Path_Terra + '*' +str(Year) + strJulDay + '.mosaic.tif')
    List_Aqua  = glob.glob(Path_Aqua  +'*'  +str(Year) + strJulDay + '.mosaic.tif')
    
#    if len(List_Terra) is not 0:
#        TerraLST_Dataset=openEuropeanLSTfile(List_Terra[0])
#        LST_Terra,Time_Terra,ViewZenithAngle_Terra,SensorObs_Terras=generateAquaTerraLSTdataset11_13(TerraLST_Dataset,startimeWindow,EndTimeWindow)
    if  len(List_Aqua) is not 0:
        AquaLST_Dataset=openEuropeanLSTfile(List_Aqua[0])
        LST_Aqua,Time_Aqua,ViewZenithAngle_Aqua,SensorObs_Aqua=generateAquaTerraLSTdataset11_13(AquaLST_Dataset,startimeWindow,EndTimeWindow)
        Obstotal=np.zeros([Time_Aqua.shape[0],Time_Aqua.shape[1]])  
#        Obstotal[SensorObs_Terras==1]=1
#        Obstotal[SensorObs_Aqua==1]=2
#        Obstotal[(SensorObs_Aqua+SensorObs_Terras)==2]=3
    #Creates the composed Aqua and terra data for LST, angles and time of observation. Terra
    #Is used asbasic and aqua overlays over it.
    
#        LST_Composed = np.where(LST_Terra != 0, LST_Terra, LST_Aqua)
#        TIME_Composed = np.where(Time_Terra != 0, Time_Terra, Time_Aqua)
#        ViewZenithAngle_Composed = np.where(ViewZenithAngle_Terra != 0, ViewZenithAngle_Terra, ViewZenithAngle_Aqua)
#        LST_Composed[LST_Composed==0]=np.nan
#        ViewZenithAngle_Composed[ViewZenithAngle_Composed==0]=np.nan
        temp=LST_Aqua*0.02 
        temp[temp==0]=-1.0
        temp2=np.abs(ViewZenithAngle_Aqua-65)
        temp2[temp==0]=np.nan
        temp2[temp2==65]=np.nan
        
        
        temp_stations=np.zeros([len(StationID),46])
        temp_stations=extractStationsFromMap(temp)
        
        temp_Map_array = np.zeros([7,10])
        
        temp_Map_array = MakeListIntoMap(temp_stations) 
        
        temp2_stations=np.zeros([len(StationID),46])
        temp2_stations=extractStationsFromMap(temp2)
        
        temp2_Map_array = np.zeros([7,10])
        
        temp2_Map_array = MakeListIntoMap(temp2_stations) 
        
        
        CreateStationArray(Year, JulDay, temp_Map_array, 'Tr')
        CreateStationArray(Year, JulDay, temp2_Map_array, 'Theta')
        
        Dictionary ['T_R']       =  temp_Map_array
        Dictionary ['T_R'][:]
        Dictionary ['theta1']    =  temp2_Map_array
#    if  len(List_Aqua) is 0:
#        #SInce Aqua was launched later we need to account for the change of not existing data. IOn this cases only Terra is used
#        LST_Terra[LST_Terra==0]=np.nan
#        ViewZenithAngle_Terra[ViewZenithAngle_Terra==0]=np.nan
#        Dictionary ['T_R']       =   LST_Terra*0.02
#        Dictionary ['theta1']    =   np.abs(ViewZenithAngle_Terra-65)
        
    
    
    return Dictionary
    



    
def FillDicTairWaterVaporPressure(Dictionary,Year,JulDay):
    
    import glob
    import gdal
    import os
    import numpy as np
    
    #Year = 2003
    #JulDay = 10
    
    Path ='Y:\\Gorka\\EUROPA\\ERA_Europe_Radiation\\Outputs\\2T\\'
    List= None  ;
    if int(JulDay) <100:
        if int(JulDay) <10:
            strJulDay='00'+str(JulDay)
        else:
            strJulDay='0'+str(JulDay)
    else:
        strJulDay=str(JulDay)
    List = glob.glob(Path + '2T_' +str(Year) + strJulDay + 'Simuated12.tif')
    if List:
        fid = gdal.Open(List[0],gdal.GA_ReadOnly)
        print('Reading the Air temperature file now. Please wait...!!!!')
        AirTemp=fid.GetRasterBand(1).ReadAsArray()
        AirTemp[np.isnan(Dictionary ['theta1'][:])]=-1
        
        AirTemp_stations=np.zeros([len(StationID),46])
        AirTemp_stations=extractStationsFromMap(AirTemp)
        
        AirTemp_Map_array = np.zeros([7,10])
        
        AirTemp_Map_array = MakeListIntoMap(AirTemp_stations)
        
        Dictionary ['T_A']       =  np.float64(AirTemp_Map_array)
        
        AirTemp_Map_array = np.float64(AirTemp_Map_array)
        
        CreateStationArray(Year, JulDay, AirTemp_Map_array, 'AirTemp')
        ea=Calc_VaporPressure(AirTemp)
        
        
        #ea[np.isnan(Dictionary ['T_R'][:])]=np.nan
        
        #testttt=ea[2024,1067]
        
################################################################################        
        
        ea_stations=np.zeros([len(StationID),46])
        ea_stations=extractStationsFromMap(ea)
        
        ea_Map_array = np.zeros([7,10])
        
        ea_Map_array = MakeListIntoMap(ea_stations)        
        
        
        ea_Map_array[np.isnan(Dictionary ['T_R'][:])]=np.nan
        Dictionary ['ea']=ea_Map_array
        
        
        CreateStationArray(Year, JulDay, ea_Map_array, 'ea')
        
        
    else:
        print('File not found. Check why!!!!')
        
    return Dictionary
    
    
def FillDicPressure(Dictionary,Year,JulDay):
    
    
    
    import glob
    import gdal
    import os
    import numpy as np
    Path ='Y:\\Gorka\\EUROPA\\ERA_Europe_Radiation\\Outputs\\SP\\'
    List= None  ;
    if int(JulDay) <100:
        if int(JulDay) <10:
            strJulDay='00'+str(JulDay)
        else:
            strJulDay='0'+str(JulDay)
    else:
        strJulDay=str(JulDay)
    List = glob.glob(Path + 'SP_' +str(Year) + strJulDay + 'Simuated13.tif')
    if List:
        fid = gdal.Open(List[0],gdal.GA_ReadOnly)
        print('Reading the Surface Pressure file now. Please wait...!!!!')
        SPressure=fid.GetRasterBand(1).ReadAsArray()
        SPressure=np.float64(SPressure*0.01)
        ###############################################################################

        SPressure_stations=np.zeros([len(StationID),46])
        SPressure_stations=extractStationsFromMap(SPressure)
        
        SPressure_Map_array = np.zeros([7,10])
        
        SPressure_Map_array = MakeListIntoMap(SPressure_stations)        
        
        SPressure_Map_array[np.isnan(Dictionary ['T_R'][:])]=np.nan        
        
        
        SPressure_Map_array[np.isnan(Dictionary ['T_R'][:])]=np.nan
        Dictionary ['p1']       =   SPressure_Map_array
        Dictionary2['p1'] = SPressure
        
        CreateStationArray(Year, JulDay, SPressure_Map_array, 'pl')
        
        
    else:
        print('File not found. Check why!!!!')
        
    return Dictionary    
#def FillFractionalCover(Dictionary,Year,JulDay):
#    import glob
#    import gdal
#    import os
#    import numpy as np
#    Path ='Y:\\Gorka\\EUROPA\\ERA_Europe_Radiation\\Outputs\\SP\\'
#    List= None  ;
#    List = glob.glob(Path + 'SP_' +str(Year) + str(JulDay) + 'Simuated12.tif')
#    if List:
#        fid = gdal.Open(List[0],gdal.GA_ReadOnly)
#
#        Dictionary ['FCover']       =   SPressure
#    else:
#        print('File not found. Check why!!!!')
#        
#    return Dictionary        
    
def FillDicLwIn(Dictionary,Year,JulDay):
    import glob
    import gdal
    import os
    import numpy as np
    Path ='Y:\\Gorka\\EUROPA\\ERA_Europe_Radiation\\Outputs\\STRD\\'
    List= None  ;
    if int(JulDay) <100:
        if int(JulDay) <10:
            strJulDay='00'+str(JulDay)
        else:
            strJulDay='0'+str(JulDay)
    else:
        strJulDay=str(JulDay)
    List = glob.glob(Path + 'STRD_' +str(Year) + strJulDay + 'Simuated13.tif')
    if List:
        fid = gdal.Open(List[0],gdal.GA_ReadOnly)
        print('Reading the Long Wave incoming radiation file now. Please wait...!!!!')
        LwIn=fid.GetRasterBand(1).ReadAsArray()
        #####################################################################################
        LwIn_stations=np.zeros([len(StationID),46])
        LwIn_stations=extractStationsFromMap(LwIn)
        
        LwIn_Map_array = np.zeros([7,10])
        
        LwIn_Map_array = MakeListIntoMap(LwIn_stations)        
        
        LwIn_Map_array[np.isnan(Dictionary['T_R'][:])]=np.nan
        
        
        #LwIn[np.isnan(Dictionary ['T_R'][:])]=np.nan
        Dictionary ['LWin']       =   np.float64(LwIn_Map_array)
        
        LwIn_Map_array = np.float64(LwIn_Map_array)
        
        
        CreateStationArray(Year, JulDay, LwIn_Map_array, 'LWin')
        
    else:
        print('File not found. Check why!!!!')
        
    return Dictionary 
    
def returnGregorianDate(Year,JulDay):
    from jdcal import gcal2jd, jd2gcal
    Jul=gcal2jd(Year,1,1)
    Jul=list(Jul)
    #Jul[1]=Jul[1]+(JulDay-1)#Carefull here. THe Sunzenith has a bug in wiritting the filename and therefore the -1 is not necesary
    Jul[1]=Jul[1]+(JulDay)#Carefull here. THe Sunzenith has a bug in wiritting the filename and therefore the -1 is not necesary
    Gcal=jd2gcal(Jul[0],Jul[1])
    Year=str(Gcal[0])
    if Gcal[1] <10:
        Month='0'+str(Gcal[1])
    else:
        Month=str(Gcal[1])
        
    if Gcal[2] <10:
        Day='0'+str(Gcal[2])
    else:
        Day=str(Gcal[2])
    
    return Year,Month,Day
        
def FillSunZenithAngle(Dictionary,Year,JulDay):
    import glob
    import gdal
    import os
    import numpy as np
    
    Year = 2003
    JulDay = 10
    
    
    Path ='Y:\\Gorka\\EUROPA\\AquaSunZenithAngle\\'
    Y,M,D=returnGregorianDate(Year,JulDay)
    List= None  ;
    List = glob.glob(Path + 'SunZenithAngle*' +Y +'-'+M+'-'+ D)
    if List:
        print('Reading the SunZenithAngle file now. Please wait...!!!!')
        array=np.fromfile(List[0],dtype='uint8')
        array=np.reshape(array, [3600,3600])
        SunZenithAngle = array.astype(float)   *1
        
        ###############################################################################
        SunZenithAngle_stations=np.zeros([len(StationID),46])
        SunZenithAngle_stations=extractStationsFromMap(SunZenithAngle)
        
        SunZenithAngle_Map_array = np.zeros([7,10])
        
        SunZenithAngle_Map_array = MakeListIntoMap(SunZenithAngle_stations)        
        
        SunZenithAngle_Map_array[np.isnan(Dictionary ['T_R'][:])]=np.nan
        Dictionary ['sza']       =   SunZenithAngle_Map_array
        
        Dictionary2['sza']= SunZenithAngle
        
        
        CreateStationArray(Year, JulDay, SunZenithAngle_Map_array, 'sza')
        
    else:
        print('File not found. Check why!!!!')
        
    return Dictionary      

def FillLAI(Dictionary,Year,JulDay):
    import glob
    import gdal
    import os
    import numpy as np
    
    Year = 2004
    JulDay = 365
    
    Path ='Y:\\Gorka\\EUROPA\\Timesat\\Outputs\\LAI_Weight\\'
    List= None  ;
    List = glob.glob(Path + 'MCD15A2.005' +str(Year)+str(JulDay)+'.Weightsmosaic.asc')
    
    
    if List:
        print('Reading the LAI file now. Please wait...!!!!')
        array=np.fromfile(List[0],dtype='uint8')
        array=np.reshape(array, [3600,3600])
        LAI_array = array*1.0
        #LAI_array[np.isnan(Dictionary ['T_R'][:])]=np.nan
        
        LAI_array_stations=np.zeros([len(StationID),46])
        LAI_array_stations=extractStationsFromMap(LAI_array)
        
        LAI_Map_array = np.zeros([7,10])
        
        LAI_Map_array = MakeListIntoMap(LAI_array_stations)
        
    
        
        Dictionary ['LAI']       =   LAI_array_stations
    else:
        print('File not found. The date is not one 8 day composite. Starting interpolation between dates and filling the dictionary...')
        File1,File2=select8dayimage(Folder,Year,JulDay)
        
        
        
        
    return Dictionary           


    
def FillDicWindSpeedr(Dictionary,Year,JulDay):
    
    import glob
    import gdal
    import os
    import numpy as np
    Path_Vertical ='Y:\\Gorka\\EUROPA\\ERA_Europe_Radiation\\Outputs\\10V\\'
    Path_Horizontal ='Y:\\Gorka\\EUROPA\\ERA_Europe_Radiation\\Outputs\\10U\\'
    
    if int(JulDay) <100:
        if int(JulDay) <10:
            strJulDay='00'+str(JulDay)
        else:
            strJulDay='0'+str(JulDay)
    else:
        strJulDay=str(JulDay)
    List= None  ;
    List_Vertical = glob.glob(Path_Vertical + '10V_' +str(Year) + strJulDay + 'Simuated12.tif')
    List_Horizontal = glob.glob(Path_Horizontal + '10U_' +str(Year) + strJulDay + 'Simuated12.tif')
    if List_Vertical:
        fid = gdal.Open(List_Vertical[0],gdal.GA_ReadOnly)
        print('Reading the Wind Velocity Vertical component now. Please wait...!!!!')
        WindVel_V=fid.GetRasterBand(1).ReadAsArray()
        fid=None
        print('Done...!!!!')
    if List_Horizontal:
        fid = gdal.Open(List_Horizontal[0],gdal.GA_ReadOnly)
        print('Reading the Wind Velocity horizontal component now. Please wait...!!!!')
        WindVel_H=fid.GetRasterBand(1).ReadAsArray()
        fid=None
        print('Done...!!!!')
    print('Calculating now the wind velocity using Pithagoras theorem...!!!!')
    
    if List_Vertical:
        WindVel= (WindVel_V**2+WindVel_H**2)**0.5
    else:
        print("file not found windvel")
        return
    if List_Horizontal:
        WindVel= (WindVel_V**2+WindVel_H**2)**0.5
    else:
        print("file not found windvel")
        return
    
    
    #########################################################################################
    WindVel_stations=np.zeros([len(StationID),46])
    WindVel_stations=extractStationsFromMap(WindVel)            #gets the Wind velocity for the stations
    
    WindVel_Map_array = np.zeros([7,10])
    
    WindVel_Map_array=MakeListIntoMap(WindVel_stations)         #makes the wind velocity of the stations into a 6*10 array
    
    WindVel_Map_array[np.isnan(Dictionary['T_R'][:])]=np.nan
    Dictionary['u']=WindVel_Map_array
    
    CreateStationArray(Year, JulDay, WindVel_Map_array, 'WindVel')
    
    return Dictionary

def CreateStationArray(Year, JulDay, Map_Array, ParameterName):         #creates and saves a 7*10 array for the selected data for the stations. Used to make quicker runs so we dont have to load 3600x3600 arrays later when running stations.
    Path = 'E:\\HenrikTSEB\\Station_Data_Arrays\\'+ParameterName+'\\'
    
    np.save(Path +str(Year) + '-' +str(JulDay) , Map_Array)
    

    
def createDailyNetCDFfile4TSEB(year,doy,Dictionary,flag, Ts, Tc, T_AC,L_nS,L_nC, LE_C,H_C,LE_S,H_S,G,R_s,R_x,R_a,u_friction, L):
    import numpy   as np
    import netCDF4 as nc  
    import numpy.ma as ma
    import struct
    import time, os, sys                     # call current time for timestamp
    from writenetcdf import writenetcdf # from ufz
    from readnetcdf  import readnetcdf  # from ufz
    from fread       import fread       # from ufz
    from autostring  import astr        # from ufz
    from pyproj      import Proj
    import scipy.misc
    import scipy.ndimage
    from netCDF4 import Dataset
    import glob
    import re
    import gdal
    years              = str(year)+','+str(year)
    startyear    = int(years.split(',')[0])
    endyear      = int(years.split(',')[1])
    times      = np.arange(0,doy+1,doy)
    
    
    
    #
    # COORDINATE SYSTEM
    #   code 31463 defines GK (DHDN3 Zone 3)
    #   [website: http://www.spatialreference.org/ref/epsg/31463/]
    #   equal +proj=tmerc +ellps=bessel +lon_0=12 +x_0=3,500,000 +y_0=0
    # coord_sys = 'epsg:31467'
    coord_sys = 'ESRI:53008'  #https://epsg.io/2211-8653
    # HEADER FILE
    #   specifies the grid properties
    #   for example, use a copy of the header.txt
    #   and adapt cellsize, ncols, nrows to your hydrologic resolution
    
    
    headerfile = 'Y:\\Gorka\\EUROPA\\Header_Europe_stations.txt' #this is for the stations. Has a 10x6 grid instead of 3600x3600
    #headerfile = 'Y:\\Gorka\\EUROPA\\Header_Europe.txt'
    
    # OUTPUT FILE
    #   path to the output file, latlon.nc is hard-coded in mHM
    outfile = 'E:\\HenrikTSEB\\NetCDFfile4TSEB\\InputsTSEB_'+str(year)+'_'+str(doy)+'.nc'
    
    import optparse
    parser = optparse.OptionParser(usage='%prog [options]',
                                   description="Cretaes latitude and longitude grids for <coord_sys> with the domain defined in  <headerfile> into NetCDF file <outfile>.")
    
    # usage example with command line arguments
    # -------------------------------------------------------------------------
    #
    # py create_latlon.py -c 'epsg:31463' -f header.txt -o latlon.nc
    #
    # -------------------------------------------------------------------------
    
    parser.add_option('-c', '--coord_sys', action='store', dest='coord_sys', type='string',
                      default=coord_sys, metavar='Property',
                      help='Coordinate system specifier according to http://www.spatialreference.org. (default: epsg:31467)')
    parser.add_option('-f', '--header', action='store', dest='headerfile', type='string',
                      default=headerfile, metavar='Header file',
                      help='Header file containing information about e.g. number of rows and columns. (default: header.txt).')
    parser.add_option('-o', '--outfile', action='store', dest='outfile', type='string',
                      default=outfile, metavar='NetCDF file',
                      help='Name of NetCDF file. (default: PET_trimmed_1km.nc).')
    
    (opts, args) = parser.parse_args()
    
    headerfile         = opts.headerfile
    outfile            = opts.outfile   
    coord_sys          = opts.coord_sys 
    
    #############################################
    
    
    # check input files
    if not os.path.isfile(headerfile):
        sys.exit("No "+headerfile+" file found here, are you in the right directory?")

    # read header information
    header_info  = np.loadtxt( headerfile, dtype='|S20')
    ncols        = np.int(header_info[0,1])
    nrows        = np.int(header_info[1,1])
    xllcorner    = np.float(header_info[2,1])
    yllcorner    = np.float(header_info[3,1])
    cs           = np.float(header_info[4,1])
    NODATA_value = header_info[5,1]

    # create x and y grid
    xx          = np.arange( xllcorner + cs/2,            xllcorner + cs/2 + ncols*cs, cs)
    yy          = np.arange( yllcorner + cs/2 + nrows*cs, yllcorner + cs/2,-cs)
    xx, yy      = np.meshgrid(xx,yy)
    # Dimensions
    fhandle      = nc.Dataset(outfile, 'w', format='NETCDF4')
    #
    # determine latitude and longitude of the Aimgrid
    projAim    = Proj(init=coord_sys)
    lons, lats = projAim(xx, yy, inverse=True)
    FiAtt  = ([['description', 'Inputs and Outputs for TSEB'],
           ['history','Created ' + time.ctime(time.time()) ]])
    writenetcdf(fhandle, fileattributes=FiAtt)
    
    startTime    = ('days since '  + str(startyear-1) + '-' + str(12).zfill(2) + '-' + str(31).zfill(2) + ' ' 
                      + str(23) + ':' + str(0).zfill(2) + ':' + str(0).zfill(2))  
    
    
    nrows=nrows+1               #no idea why i have to do this.....
#    varName = 'time'
#    dims    = len(times)
#    
#    var     = np.zeros(1)+doy
#    varAtt  = ([['units'    , startTime],
#                ['calender day' , 'standard']])
#    writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, isdim=True, vartype='i4')
    # write netCDF
    #
    varAtt  = ([['axis'     , 'X']])
    varName = 'xc'
    dims    = ncols
    var     = xx[0,:] #np.arange(ncols)+1
    writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, isdim=True)
    
    
    #
    varAtt  = ([['axis'     , 'Y']])
    varName = 'yc'
    dims    = nrows
    var     = yy[:,0] #np.arange(nrows)+1
    
    
    
    writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, isdim=True)
    # Variables
    # lon
    varAtt  = ([['units'        , 'degrees_east'],
               ['long_name'     , 'longitude']])
    #
    varName = 'lon'
    var     = lons
    dims    = ['yc','xc']
    
      
    
    writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f8', comp=True)
    # lat
    varAtt  = ([['units'        , 'degrees_north'],
               ['long_name'     , 'latitude']])
    varName = 'lat'
    var     = lats
    dims    = ['yc','xc']
    
    
    
    writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f8', comp=True)
    #
    
    
    
    
    
    
    varAtt  = ([['long_name'     , 'LST'],
            ['units'         , 'K'],
            #['NoData' , 123.456], 
            ['coordinates'   , 'lon lat']])
    varName = 'LST'
    LST=Dictionary['T_R'][:]
    var     = Dictionary['T_R'][:]
    
    
    
    
    
    #var[np.isnan(var)]=123.456
    var=np.float64(var)
    dims    = ['yc','xc']
    
    
   
    
    writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f8', comp=True)
    
    
    varAtt  = ([['long_name'     , 'View Zenith Angle (degres)'],
            ['units'         , 'grades'],
            #['NoData' , 123.456], 
            ['coordinates'   , 'lon lat']])
    varName = 'VZA'
    var_theta     = Dictionary['theta1'][:]
    var_theta=np.float64(var_theta)
    var_theta[np.isnan(LST)]=np.nan
    dims    = ['yc','xc']
    writenetcdf(fhandle, name=varName, dims=dims, var=var_theta, attributes=varAtt, vartype='f8', comp=True)
    
    varAtt  = ([['long_name'     , 'Air Temperature'],
            ['units'         , 'K'],
            ['coordinates'   , 'lon lat']])
    
    varName = 'AirTemp'
    var_Airtemp     = Dictionary['T_A'][:]
    var_Airtemp=np.float64(var_Airtemp)
    var_Airtemp[np.isnan(LST)]=np.nan

    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var_Airtemp, attributes=varAtt, vartype='f4', comp=True)
    
    varAtt  = ([['long_name'     , 'Wind Speed'],
            ['units'         , 'm/s'],
            ['coordinates'   , 'lon lat']])
    varName = 'u'
    Dictionary['u'][np.isnan(LST)]=np.nan
    var     = Dictionary['u'][:]
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
    varAtt  = ([['long_name'     , 'Water vapour pressure above the canopy'],
            ['units'         , 'mb'],
            ['coordinates'   , 'lon lat']])
    varName = 'ea'
    var     = Dictionary['ea'][:]
    var=np.float64(var)
    var[np.isnan(LST)]=np.nan
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
    varAtt  = ([['long_name'     , 'Long Wave Radiation Downwards'],
            ['units'         , 'w/m2'],
            ['coordinates'   , 'lon lat']])
    varName = 'LWin'
    var     = Dictionary['LWin'][:]
    var=np.float64(var)
    var[np.isnan(LST)]=np.nan
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
    varAtt  = ([['long_name'     , 'Sun Zenith Angle from Aqua'],
            ['units'         , 'Degrees'],
            ['coordinates'   , 'lon lat']])
    varName = 'sza'
    var     = Dictionary['sza'][:]
    var=np.float64(var)
    var[np.isnan(LST)]=np.nan
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
    varAtt  = ([['long_name'     , 'Leaf Area Index interpolated betweeen dates. '],
            ['units'         , '[-]'],
            ['coordinates'   , 'lon lat']])
    varName = 'LAI'
    var     = Dictionary['LAI'][:]
    var=np.float64(var)
    #var[np.isnan(LST)]=np.nan
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
    varAtt  = ([['long_name'     , 'IGBP MODIS land Cover classification '],
            ['units'         , '[-]'],
            ['coordinates'   , 'lon lat']])
    varName = 'lc_IGBP'
    var     = Dictionary['lc_IGBP'][:]
    var=np.float64(var)
    var[np.isnan(LST)]=np.nan
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
    varAtt  = ([['long_name'     , 'Surface Pressure'],
            ['units'         , 'mb'],
            ['coordinates'   , 'lon lat']])
    varName = 'Sp'
    var     = Dictionary['p1'][:]
    var=np.float64(var)
    var[np.isnan(LST)]=np.nan
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
    varAtt  = ([['long_name'     , 'Height of canopy based on LAI'],
            ['units'         , 'm'],
            ['coordinates'   , 'lon lat']])
    varName = 'hc'
    #Dictionary['hc'][np.isnan(LST)]=np.nan
    var     = Dictionary['hc'][:]
    var=np.float64(var)
    var[np.isnan(LST)]=np.nan
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
    varAtt  = ([['long_name'     , 'Shortwave radiation downwards'],
            ['units'         , 'w/m2'],
            ['coordinates'   , 'lon lat']])
    varName = 'Sdn'
    #Dictionary['Sdn'][np.isnan(LST)]=np.nan
    var     = Dictionary['Sdn'][:]
    var=np.float64(var)
    var[np.isnan(LST)]=np.nan    
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
    varAtt  = ([['long_name'     , 'Net Radiation Soil'],
            ['units'         , 'w/m2'],
            ['coordinates'   , 'lon lat']])
    varName = 'Rn_sw_soil'
    #Dictionary['Rn_sw_soil'][np.isnan(LST)]=np.nan
    var     = Dictionary['Rn_sw_soil'][:]
    var=np.float64(var)
    var[np.isnan(LST)]=np.nan
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
    varAtt  = ([['long_name'     , 'Net Radiation vegetation'],
            ['units'         , 'w/m2'],
            ['coordinates'   , 'lon lat']])
    varName = 'Rn_sw_veg'
   # Dictionary['Rn_sw_veg'][np.isnan(LST)]=np.nan
    var     = Dictionary['Rn_sw_veg'][:]
    var=np.float64(var)
    var[np.isnan(LST)]=np.nan
    
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
    varAtt  = ([['long_name'     , 'ZoverM'],
            ['units'         , 'xxxx'],
            ['coordinates'   , 'lon lat']])
    varName = 'z_0M'
    var     = Dictionary['z_0M'][:]
    var=np.float64(var)
    var[np.isnan(LST)]=np.nan
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)    
    
        
    varAtt  = ([['long_name'     , 'doverM'],
            ['units'         , 'xxxx'],
            ['coordinates'   , 'lon lat']])
    varName = 'd_0M'
    var     = Dictionary['d_0M'][:]
    var=np.float64(var)
    var[np.isnan(LST)]=np.nan
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
    varAtt  = ([['long_name'     , 'zetau'],
            ['units'         , 'xxxx'],
            ['coordinates'   , 'lon lat']])
    varName = 'zu'
    var     = Dictionary['zu'][:]
    var=np.float64(var)
    var[np.isnan(LST)]=np.nan
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
    varAtt  = ([['long_name'     , 'zetat'],
            ['units'         , 'xxxx'],
            ['coordinates'   , 'lon lat']])
    varName = 'zt'
    var     = Dictionary['zt'][:]
    var=np.float64(var)
    var[np.isnan(LST)]=np.nan
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
    varAtt  = ([['long_name'     , 'TSEB Flag'],
            ['units'         , '[-]'],
            ['coordinates'   , 'lon lat']])
    varName = 'Flag'
    flag[np.isnan(LST)]=np.nan
    var     = flag
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
    varAtt  = ([['long_name'     , 'TSEB Soil Temperature'],
            ['units'         , 'K'],
            ['coordinates'   , 'lon lat']])
    varName = 'SoilTemperature'
    Ts[np.isnan(LST)]=np.nan
    var     = Ts
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
    varAtt  = ([['long_name'     , 'TSEB Canopy Temperature'],
            ['units'         , 'K'],
            ['coordinates'   , 'lon lat']])
    varName = 'CanopyTemperature'
    Tc[np.isnan(LST)]=np.nan
    var     = Tc
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
    varAtt  = ([['long_name'     , 'TSEB Air temperature at the canopy interface (Kelvin)'],
            ['units'         , 'K'],
            ['coordinates'   , 'lon lat']])
    varName = 'T_AC'
    T_AC[np.isnan(LST)]=np.nan
    var     = T_AC
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
    
    varAtt  = ([['long_name'     , 'TSEB Soil net longwave radiation'],
            ['units'         , 'W m-2'],
            ['coordinates'   , 'lon lat']])
    varName = 'L_nC'
    L_nC[np.isnan(LST)]=np.nan
    var     = L_nC
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
    
    varAtt  = ([['long_name'     , 'TSEB Canopy net longwave radiation '],
            ['units'         , 'W m-2'],
            ['coordinates'   , 'lon lat']])
    varName = 'L_nS'
    L_nS[np.isnan(LST)]=np.nan
    var     = L_nS
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
    
    varAtt  = ([['long_name'     , 'TSEB Canopy latent heat flux '],
            ['units'         , 'W m-2'],
            ['coordinates'   , 'lon lat']])
    varName = 'LE_C'
    LE_C[np.isnan(LST)]=np.nan
    var     = LE_C
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
    
    varAtt  = ([['long_name'     , 'TSEB Canopy sensible heat flux '],
            ['units'         , 'W m-2'],
            ['coordinates'   , 'lon lat']])
    varName = 'H_C'
    H_C[np.isnan(LST)]=np.nan
    var     = H_C
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
    
    varAtt  = ([['long_name'     , 'TSEB Soil latent heat flux  '],
            ['units'         , 'W m-2'],
            ['coordinates'   , 'lon lat']])
    varName = 'LE_S'
    LE_S[np.isnan(LST)]=np.nan
    var     = LE_S
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
    
    varAtt  = ([['long_name'     , 'TSEB Soil sensible heat flux'],
            ['units'         , 'W m-2'],
            ['coordinates'   , 'lon lat']])
    varName = 'H_S'
    H_S[np.isnan(LST)]=np.nan
    var     = H_S
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
    varAtt  = ([['long_name'     , 'TSEB Soil heat flux '],
            ['units'         , 'W m-2'],
            ['coordinates'   , 'lon lat']])
    varName = 'G'
    G[np.isnan(LST)]=np.nan
    var     = G
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
    varAtt  = ([['long_name'     , 'TSEB Soil aerodynamic resistance to heat transport '],
            ['units'         , 's m-1'],
            ['coordinates'   , 'lon lat']])
    varName = 'R_s'
    R_s[np.isnan(LST)]=np.nan
    var     = R_s
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
    varAtt  = ([['long_name'     , 'TSEB Bulk canopy aerodynamic resistance to heat transport'],
            ['units'         , 's m-1'],
            ['coordinates'   , 'lon lat']])
    varName = 'R_x'
    R_x[np.isnan(LST)]=np.nan
    var     = R_x
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
    varAtt  = ([['long_name'     , 'TSEB Friction velocity '],
            ['units'         , 's m-1'],
            ['coordinates'   , 'lon lat']])
    varName = 'u_friction'
    u_friction[np.isnan(LST)]=np.nan
    var     = u_friction
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
    varAtt  = ([['long_name'     , 'TSEB Aerodynamic resistance to heat transport '],
            ['units'         , 's m-1'],
            ['coordinates'   , 'lon lat']])
    varName = 'R_a'
    R_a[np.isnan(LST)]=np.nan
    var     = R_a
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
    varAtt  = ([['long_name'     , 'TSEB Monin-Obuhkov length '],
            ['units'         , 'm'],
            ['coordinates'   , 'lon lat']])
    varName = 'L'
    L[np.isnan(LST)]=np.nan
    var     = L
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
#    varAtt  = ([['long_name'     , 'L_sky'],
#            ['units'         , 'wm2'],
#            ['coordinates'   , 'lon lat']])
#    varName = 'Lsky'
#    Lsky[np.isnan(LST)]=np.nan
#    var     = Lsky
#    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
#    


   
#    varAtt  = ([['long_name'     , 'TSEB number of iterations until convergence of L'],
#            ['units'         , 'm'],
#            ['coordinates'   , 'lon lat']])
#    varName = 'Iterations'
#    n_iterations[np.isnan(LST)]=np.nan
#    var     = n_iterations
#    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
    
    varAtt  = ([['long_name'     , 'Dummy (Dummyt variable. Not meaningful)']])
    varName = 'ZZZZZZ'
    var   = np.zeros([7,10])
    handle=writenetcdf(fhandle, name=varName, dims=dims, var=var, attributes=varAtt, vartype='f4', comp=True)
    
    
    
    
    return  

global Dictionary
Dictionary=MakeTSEBDictionary()
Dictionary2=MakeTSEBDictionary2()
global L_nC2
global L_nS2
L_nC2 = np.zeros([7,10])
L_nS2 = np.zeros([7,10])




    
def runMultiTSEB_stations(Year1, JulDay1):           #this method runs TSEB for a single day.
    

    #ulDay1 = 362
   #Year1 = 2006
    
    global Dictionary
    global JulDay
    global Year
    
    JulDay = JulDay1
    Year = Year1
    
    
    
    print('JULDAY AND YEAR:')
    print(JulDay1)
    print(Year1)
    print('starting..'+str(Year)+'_'+str(JulDay))
    #    for JulDay in JulDays:
    Dictionary=MakeTSEBDictionary()
    Dictionary2=MakeTSEBDictionary2()
    #Dictionary=FillDicTrAndTheta(Dictionary,Year,JulDay) #
    #Dictionary=FillDicTairWaterVaporPressure(Dictionary,Year,JulDay) #
    #Dictionary=FillDicWindSpeedr(Dictionary,Year,JulDay) #
    #Dictionary=FillDicLwIn(Dictionary,Year,JulDay) #
    #Dictionary=FillSunZenithAngle(Dictionary,Year,JulDay) #
    Dictionary=GetInterpolatedLAI('Y:\\Gorka\\EUROPA\\Timesat\\Outputs\\LAI_Weight\\',Year,JulDay,Dictionary) #
    Dictionary=GetLancover_IGBP_Gorka(Dictionary,Year) #
    #Dictionary=FillDicPressure(Dictionary,Year,JulDay) #
    #Dictionary=GetShortWaveRatiation(Dictionary)
    #Dictionary=GetFGDataFromTiff(Dictionary, JulDay)
    
#    print('runningTSEB')
#    flag, Ts, Tc, T_AC,L_nS,L_nC, LE_C,H_C,LE_S,H_S,G,R_s,R_x,R_a,u_friction, L,n_iterations = runTSEB(Dictionary)  #runs the TSEB code with the dictionary data.
#    
#    global L_nC2
#    global L_nS2
#
#    L_nC2 = L_nC                        #cant remember why i did this...
#    L_nS2 = L_nS
#    #print(Ts)
#    
#    print(Dictionary['sza'])
#    print('writing to excel...')
#    writeToExcel(JulDay, Year)          #writes the selected outputs into the excel file
    
    

def writeToExcel(JulDay, Year):
    from openpyxl.reader.excel import load_workbook
    workbookr = load_workbook(filename="E:\\HenrikTSEB\\TSEB_Station_Output.xlsx")
    
    global L_nC2
    global L_nS2
    
    ColumnYearFirst = Year - 2003
    ColumnYearSecond = ColumnYearFirst * 365
    
    #wrinting ET into excel in sheet 1.
    
    sheet = workbookr.get_sheet_by_name('ET')
    
    ET_Array = np.zeros([7,10])
    ET_Array = L_nC2 + L_nS2
    ET_Array[np.isnan(ET_Array)]=0                                #get error if we write NaN values, so overwrite with 0s?
    
    
    sheet.cell(row=1, column=JulDay + ColumnYearSecond).value = JulDay
    
        
    wolo=-1
    count=-0
    for column in range(0,10):
                
        wolo=wolo+1
        for row in range(0,7):
            count=count+1
            sheet.cell(row=count+1, column=JulDay + ColumnYearSecond).value = ET_Array[row, column]
            
                
                
    #writing LST into sheet 2
    sheet = workbookr.get_sheet_by_name('LST')
    
    LST_Array = np.zeros([7,10])
    LST_Array = Dictionary['T_R']
    LST_Array[np.isnan(LST_Array)]=0
    
    wolo=-1
    count=-0
    for column in range(0,10):
                
        wolo=wolo+1
        for row in range(0,7):
            count=count+1
            sheet.cell(row=count+1, column=JulDay + ColumnYearSecond).value = LST_Array[row, column]
    
    
    #writing air temp into sheet 3
    sheet = workbookr.get_sheet_by_name('AirTemp')
    
    AirTemp_Array = np.zeros([7,10])
    AirTemp_Array = Dictionary['T_A']
    AirTemp_Array[np.isnan(AirTemp_Array)]=0
    
    wolo=-1
    count=-0
    for column in range(0,10):
                
        wolo=wolo+1
        for row in range(0,7):
            count=count+1
            sheet.cell(row=count+1, column=JulDay + ColumnYearSecond).value = AirTemp_Array[row, column]
    
    
    
    
    workbookr.save("E:\\HenrikTSEB\\TSEB_Station_Output.xlsx")
    
    

    
    
startimeWindow=125
EndTimeWindow=155

 



reference_image='Y:\\Gorka\\EUROPA\\MYD11A1.006\\MYD11A1.0062002185.mosaic.tif'
#reference_image='F:\\LW_in_Subset\\LWin_2002001_12.tif'
metadata=getSceneMetadata(reference_image)
proj=metadata.get('proj')
geotransform=metadata.get('gt')


global Year
global JulDay

Year = 1
JulDay = 1


 


#Year=2010
#JulDay= 195
#print('starting..'+str(Year)+'_'+str(JulDay))
##    for JulDay in JulDays:
#Dictionary=MakeTSEBDictionary()
#Dictionary2=MakeTSEBDictionary2()
#Dictionary=FillDicTrAndTheta(Dictionary,Year,JulDay) #
#Dictionary=FillDicTairWaterVaporPressure(Dictionary,Year,JulDay) #
#Dictionary=FillDicWindSpeedr(Dictionary,Year,JulDay) #
#Dictionary=FillDicLwIn(Dictionary,Year,JulDay) #
#Dictionary=FillSunZenithAngle(Dictionary,Year,JulDay) #
#Dictionary=GetInterpolatedLAI('Y:\\Gorka\\EUROPA\\Timesat\\Outputs\\LAI_Weight\\',Year,JulDay,Dictionary) #
#Dictionary=GetLancover_IGBP_Gorka(Dictionary,Year) #
#Dictionary=FillDicPressure(Dictionary,Year,JulDay) #
#Dictionary=GetShortWaveRatiation(Dictionary)
#Dictionary=GetFGDataFromTiff(Dictionary, JulDay)
#
#
##Folder='Y:\\Gorka\\EUROPA\\Timesat\\Outputs\\LAI_Weight\\'
## #
##
##PruebaTA=Dictionary['T_A'][:]
##PruebaTR=Dictionary['T_R'][:]
##PruebaAngle=Dictionary['theta1'][:]
##Pruebau=Dictionary['u'][:]
##Pruebaea=Dictionary ['ea'][:]
##PruebaLWin=Dictionary ['LWin'][:]
##PruebaSunZenithAngle=Dictionary ['sza'][:]
##PruebaLAI=Dictionary ['LAI'][:]
##Pruebahc=Dictionary ['hc'][:]
##Pruebazu=Dictionary ['zu'][:]
##Pruebazt=Dictionary ['zt'][:]
##Pruebaz_0M=Dictionary ['z_0M'][:]
##Pruebad_0M=Dictionary ['d_0M'][:]
##Prueba_SP=Dictionary ['p1'][:]
##Prueba_Sdn=Dictionary ['Sdn'] 
##Prueba_Rn_sw_veg=Dictionary['Rn_sw_veg']
##Prueba_Rn_sw_soilg=Dictionary['Rn_sw_soil']
#
#print('runningTSEB')
#flag, Ts, Tc, T_AC,L_nS,L_nC, LE_C,H_C,LE_S,H_S,G,R_s,R_x,R_a,u_friction, L,n_iterations = runTSEB(Dictionary)
#
##print(Ts)
#
#print('writing to excel...')
#writeToExcel(JulDay)

#print('createdailynet')
#createDailyNetCDFfile4TSEB(Year,JulDay,Dictionary,flag, Ts, Tc, T_AC,L_nS,L_nC, LE_C,H_C,LE_S,H_S,G,R_s,R_x,R_a,u_friction, L)

#scriptfolder=os.path.dirname(__file__) 
#scriptfolder=scriptfolder+'/'
#scriptfolder=scriptfolder.replace('/','\\')        
#createoutputfolders(scriptfolder)        
#folderout=os.path.dirname(__file__)     
#folderout=folderout+'/outputs/'
#folderout=folderout.replace('/','\\')
#savefilesTSEB(folderout,Year,JulDay,geotransform,proj,flag, Ts, Tc, T_AC,L_nS,L_nC, LE_C,H_C,LE_S,H_S,G,R_s,R_x,R_a,u_friction, L,n_iterations)

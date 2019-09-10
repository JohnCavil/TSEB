# -*- coding: utf-8 -*-
"""
Created on Tue Jan 08 09:39:49 2019

@author: hlan
"""

import xlsxwriter
from openpyxl import workbook
import numpy as np
from datetime import date
from dateutil.rrule import rrule, DAILY

def CalculateMonthlyAndYearlyAverages():
    #this script takes the monthly averages for various parameters and writes them in the "averages" tab of the TSEB_and_Measured.xlsx file.
    from openpyxl.reader.excel import load_workbook
    
    
    StationIndex = np.zeros([60,1]) #for enumeration of all stations.
        
    StationIndex=[
    "BE-Lon",
    "CH-Oe2",
    "DE-Geb",
    "DE-Kli",
    "DE-RuS",
    "DE-Seh",
    "DK-Fou",
    "FR-Gri",
    "IT-BCi",
    "IT-CA2",
    "IT-Noe",
    "DE-Hai",
    "DE-Lnf",
    "DK-Sor",
    "FR-Fon",
    "IT-CA1",
    "IT-CA3",
    "IT-Col",
    "IT-Isp",
    "IT-PT1",
    "IT-Ro1",
    "IT-Ro2",
    "FR-Pue",
    "IT-Cp2",
    "IT-Cpz",
    "CH-Dav",
    "CZ-BK1",
    "DE-Lkb",
    "DE-Obe",
    "DE-Tha",
    "FR-LBr",
    "IT-La2",
    "IT-Lav",
    "IT-Ren",
    "IT-SR2",
    "IT-SRo",
    "NL-Loo",
    "RU-Fyo",
    "AT-Neu",
    "CH-Cha",
    "CH-Fru",
    "CH-Oe1",
    "CZ-BK2",
    "DE-Gri",
    "DE-RuR",
    "DK-Eng",
    "IT-MBo",
    "IT-Tor",
    "NL-Hor",
    "BE-Bra",
    "BE-Vie",
    "CH-Lae",
    "ES-Amo",
    "ES-LgS",
    "ES-LJu",
    "ES-Ln2",
    "CZ-wet",
    "DE-Akm",
    "DE-SfN",
    "DE-Spw",
    "DE-Zrk",
    ]
    
    
    workbookTSEBandMeasured = load_workbook(filename="E:\\HenrikTSEB\\TSEB_and_measured.xlsx")
        
    sheet_TSEB_ET = workbookTSEBandMeasured.get_sheet_by_name('TSEB_ET')
    sheet_LE_CorrM = workbookTSEBandMeasured.get_sheet_by_name('LE_Corr')
    sheet_Averages = workbookTSEBandMeasured.get_sheet_by_name('Averages')
    
    sheet_LE_Corr25M = workbookTSEBandMeasured.get_sheet_by_name('LE_Corr25')
    sheet_LE_Corr75M = workbookTSEBandMeasured.get_sheet_by_name('LE_Corr75')
    
    #date = sheet_TSEB_ET.cell(row=1, column=5000).value

    
    
    from datetime import date
    start = date(1996, 1, 1)
    end = date(2016, 1, 1)
    
    #Dictionary              =   {}
    #Dictionary ['jan']       =   [float(3),float(2),float(2),float(2)]
    
    #jan = []
    
    #jan.append(float(2))
    #test = Dictionary['jan']
    #test.append(float(32))
    
    
    #print float((float(sum(test))/float(len(test))))
    #Dictionary['jan'] = test
    
    
    
    
    
    
    
           
        
    #Now we run it for each parameter....    
    
    #For TSEB_ET--------------------------------------------------------------------------------------------    
    for stationName in StationIndex:
        dayNumber = 0
        monthValues = {}
        monthValues['1'] = [] 
        monthValues['2'] = [] 
        monthValues['3'] = [] 
        monthValues['4'] = [] 
        monthValues['5'] = [] 
        monthValues['6'] = [] 
        monthValues['7'] = [] 
        monthValues['8'] = [] 
        monthValues['9'] = [] 
        monthValues['10'] = [] 
        monthValues['11'] = [] 
        monthValues['12'] = []
        
        
        for dt in rrule(DAILY, dtstart=start, until=end):
            day =  int(dt.strftime("%d"))
            year =  int(dt.strftime("%Y"))
            month =  int(dt.strftime("%m"))
            dayNumber = dayNumber + 1
        
        #for day in range(1,7307):
            #dayNumber = dayNumber + 1
            print dt         
        
        
        
         
            
            
    
            date = sheet_TSEB_ET.cell(row=1, column=dayNumber).value
            print dayNumber
            print date
            try:
                whichMonth = date[date.find(":")+len(":"):date.rfind(":")]
                
            except:
                print "oh well"
            dayValue = sheet_TSEB_ET.cell(row=StationIndex.index(stationName)+2, column=dayNumber).value
            dayValue = str(dayValue)
            
            print StationIndex.index(stationName)+2
            #print dayValue
                
            if(dayValue != "" and dayValue != "None"):
                #print(whichMonth)
                
                temp = monthValues[whichMonth]
                temp.append(float(dayValue))
                monthValues[whichMonth] = temp
            
            else:
                temp = monthValues[whichMonth]
                #temp.append(float(dayValue))
                monthValues[whichMonth] = temp
            
        for nmonth in range(1,13): 
            averageNumber = "blank"
            arrayForMonth = monthValues[str(nmonth)]
            print(nmonth)
            print(stationName)
                            
            if(len(arrayForMonth) != 0):
                averageNumber = sum(arrayForMonth)/len(arrayForMonth)
            else:
                print("huh")
                averageNumber = "what?"
            sheet_Averages.cell(row=StationIndex.index(stationName)+3, column = nmonth+1).value = averageNumber 
                        
                    #print(averageNumber)
                    #print "the daily value is: " + dayValue
                    #print("row number: " + str(StationIndex.index(stationName)+3) + " for station: " + stationName)
                    #print averageNumber
                    #print nmonth
                    #print monthValues
                        
    #Now we do it for LE_Corr------------------------------------------------------------------------------------------------------                   
    for stationName in StationIndex:
        dayNumber = 0
        monthValues = {}
        monthValues['1'] = [] 
        monthValues['2'] = [] 
        monthValues['3'] = [] 
        monthValues['4'] = [] 
        monthValues['5'] = [] 
        monthValues['6'] = [] 
        monthValues['7'] = [] 
        monthValues['8'] = [] 
        monthValues['9'] = [] 
        monthValues['10'] = [] 
        monthValues['11'] = [] 
        monthValues['12'] = []
        
        
        for dt in rrule(DAILY, dtstart=start, until=end):
            day =  int(dt.strftime("%d"))
            year =  int(dt.strftime("%Y"))
            month =  int(dt.strftime("%m"))
            dayNumber = dayNumber + 1
        
        #for day in range(1,7307):
            #dayNumber = dayNumber + 1
            print dt         
        
        
        
         
            
            
    
            date = sheet_LE_CorrM.cell(row=1, column=dayNumber).value
            print dayNumber
            print date
            try:
                whichMonth = date[date.find(":")+len(":"):date.rfind(":")]
                
            except:
                print "oh well"
            dayValue = sheet_LE_CorrM.cell(row=StationIndex.index(stationName)+2, column=dayNumber).value
            dayValue = str(dayValue)
            
            print StationIndex.index(stationName)+2
            #print dayValue
                
            if(dayValue != "" and dayValue != "None"):
                #print(whichMonth)
                
                temp = monthValues[whichMonth]
                temp.append(float(dayValue))
                monthValues[whichMonth] = temp
            
            else:
                temp = monthValues[whichMonth]
                #temp.append(float(dayValue))
                monthValues[whichMonth] = temp
            
        for nmonth in range(1,13): 
            averageNumber = "blank"
            arrayForMonth = monthValues[str(nmonth)]
            print(nmonth)
            print(stationName)
                            
            if(len(arrayForMonth) != 0):
                averageNumber = sum(arrayForMonth)/len(arrayForMonth)
            else:
                print("huh")
                averageNumber = "what?"
            sheet_Averages.cell(row=StationIndex.index(stationName)+3, column = nmonth+16).value = averageNumber 
                        
                    #print(averageNumber)
                    #print "the daily value is: " + dayValue
                    #print("row number: " + str(StationIndex.index(stationName)+3) + " for station: " + stationName)
                    #print averageNumber
                    #print nmonth
                    #print monthValues
      #---------------------------------------------------------------------------------------------------------------------------                  
    workbookTSEBandMeasured.save("E:\\HenrikTSEB\\TSEB_and_measured.xlsx") 
    
     
def combineTSEBandMeasured(): #this function combines the TSEB and measured excel sheets into one, only keeping values when both sheets have good values for that day.
    
    print("starting to combine")
    StationIndex = np.zeros([1,1]) #for enumeration of all stations.
        
    StationIndex=[
    "BE-Lon",
    "CH-Oe2",
    "DE-Geb",
    "DE-Kli",
    "DE-RuS",
    "DE-Seh",
    "DK-Fou",
    "FR-Gri",
    "IT-BCi",
    "IT-CA2",
    "IT-Noe",
    "DE-Hai",
    "DE-Lnf",
    "DK-Sor",
    "FR-Fon",
    "IT-CA1",
    "IT-CA3",
    "IT-Col",
    "IT-Isp",
    "IT-PT1",
    "IT-Ro1",
    "IT-Ro2",
    "FR-Pue",
    "IT-Cp2",
    "IT-Cpz",
    "CH-Dav",
    "CZ-BK1",
    "DE-Lkb",
    "DE-Obe",
    "DE-Tha",
    "FR-LBr",
    "IT-La2",
    "IT-Lav",
    "IT-Ren",
    "IT-SR2",
    "IT-SRo",
    "NL-Loo",
    "RU-Fyo",
    "AT-Neu",
    "CH-Cha",
    "CH-Fru",
    "CH-Oe1",
    "CZ-BK2",
    "DE-Gri",
    "DE-RuR",
    "DK-Eng",
    "IT-MBo",
    "IT-Tor",
    "NL-Hor",
    "BE-Bra",
    "BE-Vie",
    "CH-Lae",
    "ES-Amo",
    "ES-LgS",
    "ES-LJu",
    "ES-Ln2",
    "CZ-wet",
    "DE-Akm",
    "DE-SfN",
    "DE-Spw",
    "DE-Zrk",
    ]
    
    from openpyxl.reader.excel import load_workbook
    
    
    
    workbookTSEBandMeasured = load_workbook(filename="E:\\HenrikTSEB\\Day_and_hour_data\\TSEB_and_fluxnet_daily_sunny_alpha.xlsx")
    workbookTSEB = load_workbook(filename="E:\\HenrikTSEB\\RESULTS\\RevisedRuns\\calculatedTSEB_correct_alpha126.xlsx") #the one where i have lined up the data correctly
    workbookMeasured = load_workbook(filename="E:\\HenrikTSEB\\Day_and_hour_data\\Fluxnet_combined_Daily.xlsx")
    
    
    
    
    #The sheets for the measured excel.
    sheet_LE_Corr = workbookMeasured.get_sheet_by_name('LE_Corr')
    sheet_LE_Corr25 = workbookMeasured.get_sheet_by_name('LE_Corr25')
    sheet_LE_Corr75 = workbookMeasured.get_sheet_by_name('LE_Corr75')
    sheet_H_Corr = workbookMeasured.get_sheet_by_name('H_Corr')
    sheet_H_Corr25 = workbookMeasured.get_sheet_by_name('H_Corr25')
    sheet_H_Corr75 = workbookMeasured.get_sheet_by_name('H_Corr75')
    sheet_LE_CorrDay = workbookMeasured.get_sheet_by_name('LE_CorrDay')
    sheet_LE_Corr25Day = workbookMeasured.get_sheet_by_name('LE_Corr25Day')
    sheet_LE_Corr75Day = workbookMeasured.get_sheet_by_name('LE_Corr75Day')
    sheet_H_CorrDay = workbookMeasured.get_sheet_by_name('H_CorrDay')
    sheet_H_Corr25Day = workbookMeasured.get_sheet_by_name('H_Corr25Day')
    sheet_H_Corr75Day = workbookMeasured.get_sheet_by_name('H_Corr75Day')
    
    #the sheet for the TSEB ET.
    sheet_TSEB = workbookTSEB.get_sheet_by_name('Ark1')
    
    #The sheets for the combined excel.
    sheet_TSEB_ET = workbookTSEBandMeasured.get_sheet_by_name('TSEB_ET')
    sheet_LE_CorrM = workbookTSEBandMeasured.get_sheet_by_name('LE_Corr')
    sheet_LE_Corr25M = workbookTSEBandMeasured.get_sheet_by_name('LE_Corr25')
    sheet_LE_Corr75M = workbookTSEBandMeasured.get_sheet_by_name('LE_Corr75')
    sheet_H_CorrM = workbookTSEBandMeasured.get_sheet_by_name('H_Corr')
    sheet_H_Corr25M = workbookTSEBandMeasured.get_sheet_by_name('H_Corr25')
    sheet_H_Corr75M = workbookTSEBandMeasured.get_sheet_by_name('H_Corr75')
    sheet_LE_CorrDayM = workbookTSEBandMeasured.get_sheet_by_name('LE_CorrDay')
    sheet_LE_Corr25DayM = workbookTSEBandMeasured.get_sheet_by_name('LE_Corr25Day')
    sheet_LE_Corr75DayM = workbookTSEBandMeasured.get_sheet_by_name('LE_Corr75Day')
    sheet_H_CorrDayM = workbookTSEBandMeasured.get_sheet_by_name('H_CorrDay')
    sheet_H_Corr25DayM = workbookTSEBandMeasured.get_sheet_by_name('H_Corr25Day')
    sheet_H_Corr75DayM = workbookTSEBandMeasured.get_sheet_by_name('H_Corr75Day')
    
    cellToKeep = False
    
    
    dayNumber = 0
    
    import datetime
    start = date(1996, 1, 1)
    end = date(2016, 1, 1)
    
    
    for dt in rrule(DAILY, dtstart=start, until=end):
        day =  int(dt.strftime("%d"))
        year =  int(dt.strftime("%Y"))
        month =  int(dt.strftime("%m"))
        dayNumber = dayNumber + 1 
        

        
        
        
        #now we go through the cells in the measure and TSEB excel and find out which ones to keep. (the ones where we have good values in both.)
        
        for stationName in StationIndex:
            
              DateToWrite = str(day) +  ":" + str(month) + ":" + str(year)
              print("Starting: " + stationName + " day: " + DateToWrite)
              print dayNumber
              

          
              
              sheet_TSEB_ET.cell(row=1, column = dayNumber+1).value = DateToWrite
              sheet_LE_CorrM.cell(row=1, column = dayNumber+1).value = DateToWrite
              
              sheet_LE_Corr25M.cell(row=1, column = dayNumber+1).value = DateToWrite
              sheet_LE_Corr75M.cell(row=1, column = dayNumber+1).value = DateToWrite
              
              
               
              measuredValue = sheet_LE_Corr.cell(row=StationIndex.index(stationName)+2, column = dayNumber+1).value
              measuredValue = str(measuredValue)
              
              measuredValue25 = sheet_LE_Corr25.cell(row=StationIndex.index(stationName)+2, column = dayNumber+1).value
              measuredValue25 = str(measuredValue25)
              
              measuredValue75 = sheet_LE_Corr75.cell(row=StationIndex.index(stationName)+2, column = dayNumber+1).value
              measuredValue75 = str(measuredValue75)
              
              
              cellToKeep = False
               
              if measuredValue != "" and measuredValue != "-999" and measuredValue != "-9999" and measuredValue != "None":
                  
                  #print("good value measured")
                  measuredValueTSEB = sheet_TSEB.cell(row=StationIndex.index(stationName)+2, column = dayNumber+1).value
                  measuredValueTSEB = str(measuredValueTSEB)
                   
                  if measuredValueTSEB != "" and measuredValueTSEB != "-999" and measuredValueTSEB != "-9999" and measuredValueTSEB != "-1998" and measuredValueTSEB!= "None":
                      #print("good value TSEB") 
                      
                      print (measuredValueTSEB + ", " + measuredValue)
                      cellToKeep = True
                      sheet_TSEB_ET.cell(row=StationIndex.index(stationName)+2, column = dayNumber+1).value = float(measuredValueTSEB)
                      sheet_LE_CorrM.cell(row=StationIndex.index(stationName)+2, column = dayNumber+1).value = float(measuredValue)
                      
                      sheet_LE_Corr25M.cell(row=StationIndex.index(stationName)+2, column = dayNumber+1).value = float(measuredValue25)
                      sheet_LE_Corr75M.cell(row=StationIndex.index(stationName)+2, column = dayNumber+1).value = float(measuredValue75)
                       
              if cellToKeep == False:
                   sheet_TSEB_ET.cell(row=StationIndex.index(stationName)+2, column = dayNumber+1).value = ""
                   sheet_LE_CorrM.cell(row=StationIndex.index(stationName)+2, column = dayNumber+1).value = ""
                   
                   sheet_LE_Corr25M.cell(row=StationIndex.index(stationName)+2, column = dayNumber+1).value = ""
                   sheet_LE_Corr75M.cell(row=StationIndex.index(stationName)+2, column = dayNumber+1).value = ""
                       
                   
                              
    workbookTSEBandMeasured.save("E:\\HenrikTSEB\\RESULTS\\RevisedRuns\\TSEB_and_fluxnet_alpha126.xlsx")                   
                   
               
        

def runItForAll(): #takes the data from all the station excel files and puts what we need into FluxnetCombined.xlsx
    
    print("STARTING")
    StationIndex = np.zeros([60,1]) #for enumeration of all stations.
        
    StationIndex=[
    "BE-Lon",
    "CH-Oe2",
    "DE-Geb",
    "DE-Kli",
    "DE-RuS",
    "DE-Seh",
    "DK-Fou",
    "FR-Gri",
    "IT-BCi",
    "IT-CA2",
    "IT-Noe",
    "DE-Hai",
    "DE-Lnf",
    "DK-Sor",
    "FR-Fon",
    "IT-CA1",
    "IT-CA3",
    "IT-Col",
    "IT-Isp",
    "IT-PT1",
    "IT-Ro1",
    "IT-Ro2",
    "FR-Pue",
    "IT-Cp2",
    "IT-Cpz",
    "CH-Dav",
    "CZ-BK1",
    "DE-Lkb",
    "DE-Obe",
    "DE-Tha",
    "FR-LBr",
    "IT-La2",
    "IT-Lav",
    "IT-Ren",
    "IT-SR2",
    "IT-SRo",
    "NL-Loo",
    "RU-Fyo",
    "AT-Neu",
    "CH-Cha",
    "CH-Fru",
    "CH-Oe1",
    "CZ-BK2",
    "DE-Gri",
    "DE-RuR",
    "DK-Eng",
    "IT-MBo",
    "IT-Tor",
    "NL-Hor",
    "BE-Bra",
    "BE-Vie",
    "CH-Lae",
    "ES-Amo",
    "ES-LgS",
    "ES-LJu",
    "ES-Ln2",
    "CZ-wet",
    "DE-Akm",
    "DE-SfN",
    "DE-Spw",
    "DE-Zrk",
    ]
        
    for station in StationIndex:
            
        try:
            print ("getting data for station: " + station)
            getDataFromFluxnet(station)
            
        except:
            print("couldn't get data for station: " + station)
            continue

def getDataFromFluxnet(StationName):  #creates the excel sheets based on the arrays.
    
        print("init")
        #StationName = "AT-Neu"
        StationIndex = np.zeros([60,1])         #this is used to write the stations names in the excel sheet.
        
        StationIndex=[
        "BE-Lon",
        "CH-Oe2",
        "DE-Geb",
        "DE-Kli",
        "DE-RuS",
        "DE-Seh",
        "DK-Fou",
        "FR-Gri",
        "IT-BCi",
        "IT-CA2",
        "IT-Noe",
        "DE-Hai",
        "DE-Lnf",
        "DK-Sor",
        "FR-Fon",
        "IT-CA1",
        "IT-CA3",
        "IT-Col",
        "IT-Isp",
        "IT-PT1",
        "IT-Ro1",
        "IT-Ro2",
        "FR-Pue",
        "IT-Cp2",
        "IT-Cpz",
        "CH-Dav",
        "CZ-BK1",
        "DE-Lkb",
        "DE-Obe",
        "DE-Tha",
        "FR-LBr",
        "IT-La2",
        "IT-Lav",
        "IT-Ren",
        "IT-SR2",
        "IT-SRo",
        "NL-Loo",
        "RU-Fyo",
        "AT-Neu",
        "CH-Cha",
        "CH-Fru",
        "CH-Oe1",
        "CZ-BK2",
        "DE-Gri",
        "DE-RuR",
        "DK-Eng",
        "IT-MBo",
        "IT-Tor",
        "NL-Hor",
        "BE-Bra",
        "BE-Vie",
        "CH-Lae",
        "ES-Amo",
        "ES-LgS",
        "ES-LJu",
        "ES-Ln2",
        "CZ-wet",
        "DE-Akm",
        "DE-SfN",
        "DE-Spw",
        "DE-Zrk",
        ]
    
        StationIndex.index("ES-Amo")
    
    
    
        from openpyxl.reader.excel import load_workbook
        workbookr = load_workbook(filename="E:\\HenrikTSEB\\FluxnetDataset\\FluxnetDataset\\Fluxnet_xlxs\\FLX_"+StationName+".xlsx") #these have been created manually.
        
        workbookCombined = load_workbook(filename="E:\\HenrikTSEB\\FluxnetDataset\\FluxnetDataset\\Fluxnet_xlxs\\Fluxnet_combined.xlsx")
        
        
        sheet = workbookr.get_sheet_by_name('Ark1')
        sheet_LE_Corr = workbookCombined.get_sheet_by_name('LE_Corr')
        sheet_LE_Corr25 = workbookCombined.get_sheet_by_name('LE_Corr25')
        sheet_LE_Corr75 = workbookCombined.get_sheet_by_name('LE_Corr75')
        sheet_H_Corr = workbookCombined.get_sheet_by_name('H_Corr')
        sheet_H_Corr25 = workbookCombined.get_sheet_by_name('H_Corr25')
        sheet_H_Corr75 = workbookCombined.get_sheet_by_name('H_Corr75')
        sheet_LE_CorrDay = workbookCombined.get_sheet_by_name('LE_CorrDay')
        sheet_LE_Corr25Day = workbookCombined.get_sheet_by_name('LE_Corr25Day')
        sheet_LE_Corr75Day = workbookCombined.get_sheet_by_name('LE_Corr75Day')
        sheet_H_CorrDay = workbookCombined.get_sheet_by_name('H_CorrDay')
        sheet_H_Corr25Day = workbookCombined.get_sheet_by_name('H_Corr25Day')
        sheet_H_Corr75Day = workbookCombined.get_sheet_by_name('H_Corr75Day')
       
        a = date(1996, 1, 1)
        b = date(2015, 1, 1)
        dayNumber=1         #this is the day number for the station itself. So if it starts in 2002 then it will be day one 2002/01/01
        realDayNumber=1       #this is the day number between 1996 and 2015. Not using this?
        
        yearValue = sheet.cell(row=2, column=1).value
        
        
        yearValue = int(str(yearValue)[:4])   #gets the first 4 digits of the first row in the first column, aka the year that we start at for that station.
        
        
        #now we get the number of days we need to start writing at. So if the station starts in 2005 we need 5*365 (+ leap days) before we write.
        import datetime
        startDate = datetime.date(1996, 01, 01)
        stationStartDate = datetime.date(yearValue, 01, 01)
        diff = stationStartDate - startDate
        daysToStart = diff.days           #this is used to know in what cell we should start writing at for that station. So how many days after 01/01/1996 the station started working at.                                         

        
        
        for dt in rrule(DAILY, dtstart=a, until=b):
            day =  int(dt.strftime("%d"))
            year =  int(dt.strftime("%Y"))/int(a.strftime("%Y"))
            month =  int(dt.strftime("%m"))
                   
    
            #first we now write the averages for the 12:00 to 14:00 period for the 6 variables. (This is also what TSEB measures?)
                   
            try:
                
                if(day==1):
                    value1200M = sheet.cell(row=dayNumber * 48-22, column=4).value
                    value1230M = sheet.cell(row=dayNumber * 48-21, column=4).value
                    value1300M = sheet.cell(row=dayNumber * 48-20, column=4).value
                    value1330M = sheet.cell(row=dayNumber * 48-19, column=4).value
                else:
                    value1200M = sheet.cell(row=dayNumber * 48+26, column=4).value
                    value1230M = sheet.cell(row=dayNumber * 48+27, column=4).value
                    value1300M = sheet.cell(row=dayNumber * 48+28, column=4).value
                    value1330M = sheet.cell(row=dayNumber * 48+29, column=4).value
                    
                sheet_LE_Corr25.cell(row=1, column=dayNumber+1).value = dayNumber     
                sheet_LE_Corr25.cell(row=StationIndex.index(StationName)+2, column=daysToStart+1).value = np.mean([value1200M, value1230M, value1300M, value1330M])
                
                if(day==1):
                    if(value1200M == -9999 or value1230M == -9999 or value1300M == -9999 or value1330M == -9999): #this is done so that if we have one bad value then all the values become bad.
                        value1200 = -9999
                        value1230 = -9999
                        value1300 = -9999
                        value1330 = -9999
                    else:
                        value1200 = sheet.cell(row=dayNumber * 48-22, column=3).value
                        value1230 = sheet.cell(row=dayNumber * 48-21, column=3).value
                        value1300 = sheet.cell(row=dayNumber * 48-20, column=3).value
                        value1330 = sheet.cell(row=dayNumber * 48-19, column=3).value
                else:
                    if(value1200M == -9999 or value1230M == -9999 or value1300M == -9999 or value1330M == -9999):
                        value1200 = -9999
                        value1230 = -9999
                        value1300 = -9999
                        value1330 = -9999
                    else:
                        value1200 = sheet.cell(row=dayNumber * 48+26, column=3).value
                        value1230 = sheet.cell(row=dayNumber * 48+27, column=3).value
                        value1300 = sheet.cell(row=dayNumber * 48+28, column=3).value
                        value1330 = sheet.cell(row=dayNumber * 48+29, column=3).value
                 
                sheet_LE_Corr.cell(row=1, column=dayNumber+1).value = dayNumber  
                sheet_LE_Corr.cell(row=StationIndex.index(StationName)+2, column=daysToStart+1).value = np.mean([value1200, value1230, value1300, value1330])
                
                  
                
                if(day==1):
                    value1200 = sheet.cell(row=dayNumber * 48-22, column=5).value
                    value1230 = sheet.cell(row=dayNumber * 48-21, column=5).value
                    value1300 = sheet.cell(row=dayNumber * 48-20, column=5).value
                    value1330 = sheet.cell(row=dayNumber * 48-19, column=5).value
                else:
                    value1200 = sheet.cell(row=dayNumber * 48+26, column=5).value
                    value1230 = sheet.cell(row=dayNumber * 48+27, column=5).value
                    value1300 = sheet.cell(row=dayNumber * 48+28, column=5).value
                    value1330 = sheet.cell(row=dayNumber * 48+29, column=5).value
                 
                sheet_LE_Corr75.cell(row=1, column=dayNumber+1).value = dayNumber    
                sheet_LE_Corr75.cell(row=StationIndex.index(StationName)+2, column=daysToStart+1).value = np.mean([value1200, value1230, value1300, value1330])
                
                if(day==1):
                    if(value1200M == -9999 or value1230M == -9999 or value1300M == -9999 or value1330M == -9999):
                        value1200 = -9999
                        value1230 = -9999
                        value1300 = -9999
                        value1330 = -9999
                    else:
                        value1200 = sheet.cell(row=dayNumber * 48-22, column=3).value
                        value1230 = sheet.cell(row=dayNumber * 48-21, column=3).value
                        value1300 = sheet.cell(row=dayNumber * 48-20, column=3).value
                        value1330 = sheet.cell(row=dayNumber * 48-19, column=3).value
                else:
                    if(value1200M == -9999 or value1230M == -9999 or value1300M == -9999 or value1330M == -9999):
                        value1200 = -9999
                        value1230 = -9999
                        value1300 = -9999
                        value1330 = -9999
                    else:
                        value1200 = sheet.cell(row=dayNumber * 48+26, column=3).value
                        value1230 = sheet.cell(row=dayNumber * 48+27, column=3).value
                        value1300 = sheet.cell(row=dayNumber * 48+28, column=3).value
                        value1330 = sheet.cell(row=dayNumber * 48+29, column=3).value
                 
                sheet_H_Corr.cell(row=1, column=dayNumber+1).value = dayNumber    
                sheet_H_Corr.cell(row=StationIndex.index(StationName)+2, column=daysToStart+1).value = np.mean([value1200, value1230, value1300, value1330])
                
                if(day==1):
                    value1200 = sheet.cell(row=dayNumber * 48-22, column=7).value
                    value1230 = sheet.cell(row=dayNumber * 48-21, column=7).value
                    value1300 = sheet.cell(row=dayNumber * 48-20, column=7).value
                    value1330 = sheet.cell(row=dayNumber * 48-19, column=7).value
                else:
                    value1200 = sheet.cell(row=dayNumber * 48+26, column=7).value
                    value1230 = sheet.cell(row=dayNumber * 48+27, column=7).value
                    value1300 = sheet.cell(row=dayNumber * 48+28, column=7).value
                    value1330 = sheet.cell(row=dayNumber * 48+29, column=7).value
                 
                sheet_H_Corr25.cell(row=1, column=dayNumber+1).value = dayNumber    
                sheet_H_Corr25.cell(row=StationIndex.index(StationName)+2, column=daysToStart+1).value = np.mean([value1200, value1230, value1300, value1330])
                
                if(day==1):
                    value1200 = sheet.cell(row=dayNumber * 48-22, column=8).value
                    value1230 = sheet.cell(row=dayNumber * 48-21, column=8).value
                    value1300 = sheet.cell(row=dayNumber * 48-20, column=8).value
                    value1330 = sheet.cell(row=dayNumber * 48-19, column=8).value
                else:
                    value1200 = sheet.cell(row=dayNumber * 48+26, column=8).value
                    value1230 = sheet.cell(row=dayNumber * 48+27, column=8).value
                    value1300 = sheet.cell(row=dayNumber * 48+28, column=8).value
                    value1330 = sheet.cell(row=dayNumber * 48+29, column=8).value
                 
                sheet_H_Corr75.cell(row=1, column=dayNumber+1).value = dayNumber    
                sheet_H_Corr75.cell(row=StationIndex.index(StationName)+2, column=daysToStart+1).value = np.mean([value1200, value1230, value1300, value1330])
                    
                
            except:     #if something is wrong we just write -9999 (maybe we should leave it blank instead?)
                #print("data missing")
                sheet_LE_Corr.cell(row=StationIndex.index(StationName)+2, column=daysToStart+1).value = -9999
                sheet_LE_Corr25.cell(row=StationIndex.index(StationName)+2, column=daysToStart+1).value = -9999
                sheet_LE_Corr75.cell(row=StationIndex.index(StationName)+2, column=daysToStart+1).value = -9999
                sheet_H_Corr.cell(row=StationIndex.index(StationName)+2, column=daysToStart+1).value = -9999
                sheet_H_Corr25.cell(row=StationIndex.index(StationName)+2, column=daysToStart+1).value = -9999
                sheet_H_Corr75.cell(row=StationIndex.index(StationName)+2, column=daysToStart+1).value = -9999
                
                
                
            
            #---------------------------------------------------------------------------------------------------------------------------------------------------------
            #now we do the same but for the daily averages: 
            
            
            try:
                DayValues  = np.zeros([48,1])
                
                
                
                for timeOfDay in range(1,49):
                    
                    if(day==1):
                        if -9999 in DayValues:
                            DayValues = np.full([48,1],-9999)
                        else:
                            valueAtTime = sheet.cell(row=dayNumber * 48+ timeOfDay+1, column=3).value
                            DayValues[timeOfDay-1] = valueAtTime
                            
                    else:
                        
                        valueAtTime = sheet.cell(row=dayNumber * 48+timeOfDay+1, column=3).value
                        if -9999 in DayValues:
                            DayValues = np.full([48,1],-9999)
                        else:
                            DayValues[timeOfDay-1] = valueAtTime
                            
                        
                        
                    
                    
                sheet_LE_CorrDay.cell(row=1, column=dayNumber+1).value = dayNumber
                
                valueOfDay = float(sum(DayValues)/len(DayValues))
                #print(valueOfDay)
                #print(dayNumber)
                sheet_LE_CorrDay.cell(row=StationIndex.index(StationName)+2, column=daysToStart+1).value = valueOfDay
                
                #_---------------------------------------------------------------------------------------------------------------
                DayValues = np.zeros([48,1])

                for timeOfDay in range(1,49):
                    
                    if(day==1):
                        if -9999 in DayValues:
                            DayValues = np.full([48,1],-9999)
                        else:
                            valueAtTime = sheet.cell(row=dayNumber * 48+ timeOfDay+1, column=4).value
                            DayValues[timeOfDay-1] = valueAtTime
                            
                    else:
                        
                        valueAtTime = sheet.cell(row=dayNumber * 48+timeOfDay+1, column=4).value
                        if -9999 in DayValues:
                            DayValues = np.full([48,1],-9999)
                        else:
                            DayValues[timeOfDay-1] = valueAtTime
                            
                        
                        
                    
                    
                sheet_LE_Corr25Day.cell(row=1, column=dayNumber+1).value = dayNumber
                
                valueOfDay = float(sum(DayValues)/len(DayValues))
                #print(valueOfDay)
                #print(dayNumber)
                sheet_LE_Corr25Day.cell(row=StationIndex.index(StationName)+2, column=daysToStart+1).value = valueOfDay
                
                #_---------------------------------------------------------------------------------------------------------------
                DayValues = np.zeros([48,1])

                for timeOfDay in range(1,49):
                    
                    if(day==1):
                        if -9999 in DayValues:
                            DayValues = np.full([48,1],-9999)
                        else:
                            valueAtTime = sheet.cell(row=dayNumber * 48+ timeOfDay+1, column=5).value
                            DayValues[timeOfDay-1] = valueAtTime
                            
                    else:
                        
                        valueAtTime = sheet.cell(row=dayNumber * 48+timeOfDay+1, column=5).value
                        if -9999 in DayValues:
                            DayValues = np.full([48,1],-9999)
                        else:
                            DayValues[timeOfDay-1] = valueAtTime
                            
                        
                        
                    
                    
                sheet_LE_Corr75Day.cell(row=1, column=dayNumber+1).value = dayNumber
                
                valueOfDay = float(sum(DayValues)/len(DayValues))
                #print(valueOfDay)
                #print(dayNumber)
                sheet_LE_Corr75Day.cell(row=StationIndex.index(StationName)+2, column=daysToStart+1).value = valueOfDay
                
                #_---------------------------------------------------------------------------------------------------------------
                DayValues = np.zeros([48,1])

                for timeOfDay in range(1,49):
                    
                    if(day==1):
                        if -9999 in DayValues:
                            DayValues = np.full([48,1],-9999)
                        else:
                            valueAtTime = sheet.cell(row=dayNumber * 48+ timeOfDay+1, column=6).value
                            DayValues[timeOfDay-1] = valueAtTime
                            
                    else:
                        
                        valueAtTime = sheet.cell(row=dayNumber * 48+timeOfDay+1, column=6).value
                        if -9999 in DayValues:
                            DayValues = np.full([48,1],-9999)
                        else:
                            DayValues[timeOfDay-1] = valueAtTime
                            
                        
                        
                    
                    
                sheet_H_CorrDay.cell(row=1, column=dayNumber+1).value = dayNumber
                
                valueOfDay = float(sum(DayValues)/len(DayValues))
                #print(valueOfDay)
                #print(dayNumber)
                sheet_H_CorrDay.cell(row=StationIndex.index(StationName)+2, column=daysToStart+1).value = valueOfDay
                
                #_---------------------------------------------------------------------------------------------------------------
                DayValues = np.zeros([48,1])

                for timeOfDay in range(1,49):
                    
                    if(day==1):
                        if -9999 in DayValues:
                            DayValues = np.full([48,1],-9999)
                        else:
                            valueAtTime = sheet.cell(row=dayNumber * 48+ timeOfDay+1, column=7).value
                            DayValues[timeOfDay-1] = valueAtTime
                            
                    else:
                        
                        valueAtTime = sheet.cell(row=dayNumber * 48+timeOfDay+1, column=7).value
                        if -9999 in DayValues:
                            DayValues = np.full([48,1],-9999)
                        else:
                            DayValues[timeOfDay-1] = valueAtTime
                            
                        
                        
                    
                    
                sheet_H_Corr25Day.cell(row=1, column=dayNumber+1).value = dayNumber
                
                valueOfDay = float(sum(DayValues)/len(DayValues))
                #print(valueOfDay)
                #print(dayNumber)
                sheet_H_Corr25Day.cell(row=StationIndex.index(StationName)+2, column=daysToStart+1).value = valueOfDay
                
                #_---------------------------------------------------------------------------------------------------------------
                DayValues = np.zeros([48,1])

                for timeOfDay in range(1,49):
                    
                    if(day==1):
                        if -9999 in DayValues:
                            DayValues = np.full([48,1],-9999)
                        else:
                            valueAtTime = sheet.cell(row=dayNumber * 48+ timeOfDay+1, column=8).value
                            DayValues[timeOfDay-1] = valueAtTime
                            
                    else:
                        
                        valueAtTime = sheet.cell(row=dayNumber * 48+timeOfDay+1, column=8).value
                        if -9999 in DayValues:
                            DayValues = np.full([48,1],-9999)
                        else:
                            DayValues[timeOfDay-1] = valueAtTime
                            
                        
                        
                    
                    
                sheet_H_Corr75Day.cell(row=1, column=dayNumber+1).value = dayNumber
                
                valueOfDay = float(sum(DayValues)/len(DayValues))
                #print(valueOfDay)
                #print(dayNumber)
                sheet_H_Corr75Day.cell(row=StationIndex.index(StationName)+2, column=daysToStart+1).value = valueOfDay
                
                
                
            except:
                #print("data missing average")
                sheet_LE_CorrDay.cell(row=StationIndex.index(StationName)+2, column=daysToStart+1).value = -9999
                sheet_LE_Corr25Day.cell(row=StationIndex.index(StationName)+2, column=daysToStart+1).value = -9999
                sheet_LE_Corr75Day.cell(row=StationIndex.index(StationName)+2, column=daysToStart+1).value = -9999
                sheet_H_CorrDay.cell(row=StationIndex.index(StationName)+2, column=daysToStart+1).value = -9999
                sheet_H_Corr25Day.cell(row=StationIndex.index(StationName)+2, column=daysToStart+1).value = -9999
                sheet_H_Corr75Day.cell(row=StationIndex.index(StationName)+2, column=daysToStart+1).value = -9999
                
                
        
            dayNumber = dayNumber+1
            daysToStart = daysToStart + 1
                   
                   
                
                
                
        workbookCombined.save("E:\\HenrikTSEB\\FluxnetDataset\\FluxnetDataset\\Fluxnet_xlxs\\Fluxnet_combined.xlsx") 
def runItForAllDaily(): #takes the data from all the station excel files and puts what we need into FluxnetCombined.xlsx
    
    print("STARTING")
    StationIndex = np.zeros([60,1]) #for enumeration of all stations.
        
    StationIndex=[
    "DE-Tha"
    ]
        
    for station in StationIndex:
            
        try:
            print ("getting data for station: " + station)
            getDataFromFluxnetDaily(station)
            
        except:
            print("couldn't get data for station: " + station)
            continue        
def getDataFromFluxnetDaily(StationName):  #creates the excel sheets based on the arrays. Daily instead of mid day. I'm sorry for the way this is done, please forgive, easier to do it this way based on previous script than to make a loop.
    
        print("init")
        #StationName = "AT-Neu"
        StationIndex = np.zeros([60,1])         #this is used to write the stations names in the excel sheet.
        
        StationIndex=[
        "BE-Lon",
        "CH-Oe2",
        "DE-Geb",
        "DE-Kli",
        "DE-RuS",
        "DE-Seh",
        "DK-Fou",
        "FR-Gri",
        "IT-BCi",
        "IT-CA2",
        "IT-Noe",
        "DE-Hai",
        "DE-Lnf",
        "DK-Sor",
        "FR-Fon",
        "IT-CA1",
        "IT-CA3",
        "IT-Col",
        "IT-Isp",
        "IT-PT1",
        "IT-Ro1",
        "IT-Ro2",
        "FR-Pue",
        "IT-Cp2",
        "IT-Cpz",
        "CH-Dav",
        "CZ-BK1",
        "DE-Lkb",
        "DE-Obe",
        "DE-Tha",
        "FR-LBr",
        "IT-La2",
        "IT-Lav",
        "IT-Ren",
        "IT-SR2",
        "IT-SRo",
        "NL-Loo",
        "RU-Fyo",
        "AT-Neu",
        "CH-Cha",
        "CH-Fru",
        "CH-Oe1",
        "CZ-BK2",
        "DE-Gri",
        "DE-RuR",
        "DK-Eng",
        "IT-MBo",
        "IT-Tor",
        "NL-Hor",
        "BE-Bra",
        "BE-Vie",
        "CH-Lae",
        "ES-Amo",
        "ES-LgS",
        "ES-LJu",
        "ES-Ln2",
        "CZ-wet",
        "DE-Akm",
        "DE-SfN",
        "DE-Spw",
        "DE-Zrk",
        ]
    
        StationIndex.index("ES-Amo")
    
    
    
        from openpyxl.reader.excel import load_workbook
        workbookr = load_workbook(filename="E:\\HenrikTSEB\\FluxnetDataset\\FluxnetDataset\\Fluxnet_xlxs\\FLX_"+StationName+".xlsx") #these have been created manually.
        
        workbookCombined = load_workbook(filename="E:\\HenrikTSEB\\FluxnetDataset\\FluxnetDataset\\Fluxnet_xlxs\\Fluxnet_combined_Daily_Test.xlsx")
        
        
        sheet = workbookr.get_sheet_by_name('Ark1')
        sheet_LE_Corr = workbookCombined.get_sheet_by_name('LE_Corr')
        sheet_LE_Corr25 = workbookCombined.get_sheet_by_name('LE_Corr25')
        sheet_LE_Corr75 = workbookCombined.get_sheet_by_name('LE_Corr75')
        sheet_H_Corr = workbookCombined.get_sheet_by_name('H_Corr')
        sheet_H_Corr25 = workbookCombined.get_sheet_by_name('H_Corr25')
        sheet_H_Corr75 = workbookCombined.get_sheet_by_name('H_Corr75')
        sheet_LE_CorrDay = workbookCombined.get_sheet_by_name('LE_CorrDay')
        sheet_LE_Corr25Day = workbookCombined.get_sheet_by_name('LE_Corr25Day')
        sheet_LE_Corr75Day = workbookCombined.get_sheet_by_name('LE_Corr75Day')
        sheet_H_CorrDay = workbookCombined.get_sheet_by_name('H_CorrDay')
        sheet_H_Corr25Day = workbookCombined.get_sheet_by_name('H_Corr25Day')
        sheet_H_Corr75Day = workbookCombined.get_sheet_by_name('H_Corr75Day')
       
        a = date(1996, 1, 1)
        b = date(2015, 1, 1)
        dayNumber=1         #this is the day number for the station itself. So if it starts in 2002 then it will be day one 2002/01/01
        realDayNumber=1       #this is the day number between 1996 and 2015. Not using this?
        
        yearValue = sheet.cell(row=2, column=1).value
        
        
        yearValue = int(str(yearValue)[:4])   #gets the first 4 digits of the first row in the first column, aka the year that we start at for that station.
        
        
        #now we get the number of days we need to start writing at. So if the station starts in 2005 we need 5*365 (+ leap days) before we write.
        import datetime
        startDate = datetime.date(1996, 01, 01)
        stationStartDate = datetime.date(yearValue, 01, 01)
        diff = stationStartDate - startDate
        daysToStart = diff.days           #this is used to know in what cell we should start writing at for that station. So how many days after 01/01/1996 the station started working at.                                         

        
        
        for dt in rrule(DAILY, dtstart=a, until=b):
            day =  int(dt.strftime("%d"))
            year =  int(dt.strftime("%Y"))/int(a.strftime("%Y"))
            month =  int(dt.strftime("%m"))
                   
    
            #first we now write the averages for the 12:00 to 14:00 period for the 6 variables. (This is also what TSEB measures?)
                   
            try:            #FORGIVE ME-------------------------------------------
                
                if(day==1):
                    value0000M= sheet.cell(row=dayNumber * 48-46, column=4).value
                    value0030M= sheet.cell(row=dayNumber * 48-45, column=4).value
                    value0100M= sheet.cell(row=dayNumber * 48-44, column=4).value
                    value0130M= sheet.cell(row=dayNumber * 48-43, column=4).value
                    value0200M= sheet.cell(row=dayNumber * 48-42, column=4).value
                    value0230M= sheet.cell(row=dayNumber * 48-41, column=4).value
                    value0300M= sheet.cell(row=dayNumber * 48-40, column=4).value
                    value0330M= sheet.cell(row=dayNumber * 48-39, column=4).value
                    value0400M= sheet.cell(row=dayNumber * 48-38, column=4).value
                    value0430M= sheet.cell(row=dayNumber * 48-37, column=4).value
                    value0500M= sheet.cell(row=dayNumber * 48-36, column=4).value
                    value0530M= sheet.cell(row=dayNumber * 48-35, column=4).value
                    value0600M= sheet.cell(row=dayNumber * 48-34, column=4).value
                    value0630M= sheet.cell(row=dayNumber * 48-33, column=4).value
                    value0700M= sheet.cell(row=dayNumber * 48-32, column=4).value
                    value0730M= sheet.cell(row=dayNumber * 48-31, column=4).value
                    value0800M= sheet.cell(row=dayNumber * 48-30, column=4).value
                    value0830M= sheet.cell(row=dayNumber * 48-29, column=4).value
                    value0900M= sheet.cell(row=dayNumber * 48-28, column=4).value
                    value0930M= sheet.cell(row=dayNumber * 48-27, column=4).value
                    value1000M= sheet.cell(row=dayNumber * 48-26, column=4).value
                    value1030M= sheet.cell(row=dayNumber * 48-25, column=4).value
                    value1100M= sheet.cell(row=dayNumber * 48-24, column=4).value
                    value1130M= sheet.cell(row=dayNumber * 48-23, column=4).value
                    value1200M = sheet.cell(row=dayNumber * 48-22, column=4).value
                    value1230M = sheet.cell(row=dayNumber * 48-21, column=4).value
                    value1300M = sheet.cell(row=dayNumber * 48-20, column=4).value
                    value1330M = sheet.cell(row=dayNumber * 48-19, column=4).value
                    value1400M= sheet.cell(row=dayNumber * 48-18, column=4).value
                    value1430M= sheet.cell(row=dayNumber * 48-17, column=4).value
                    value1500M= sheet.cell(row=dayNumber * 48-16, column=4).value
                    value1530M= sheet.cell(row=dayNumber * 48-15, column=4).value
                    value1600M= sheet.cell(row=dayNumber * 48-14, column=4).value
                    value1630M= sheet.cell(row=dayNumber * 48-13, column=4).value
                    value1700M= sheet.cell(row=dayNumber * 48-12, column=4).value
                    value1730M= sheet.cell(row=dayNumber * 48-11, column=4).value
                    value1800M= sheet.cell(row=dayNumber * 48-10, column=4).value
                    value1830M= sheet.cell(row=dayNumber * 48-9, column=4).value
                    value1900M= sheet.cell(row=dayNumber * 48-8, column=4).value
                    value1930M= sheet.cell(row=dayNumber * 48-7, column=4).value
                    value2000M= sheet.cell(row=dayNumber * 48-6, column=4).value
                    value2030M= sheet.cell(row=dayNumber * 48-5, column=4).value
                    value2100M= sheet.cell(row=dayNumber * 48-4, column=4).value
                    value2130M= sheet.cell(row=dayNumber * 48-3, column=4).value
                    value2200M= sheet.cell(row=dayNumber * 48-2, column=4).value
                    value2230M= sheet.cell(row=dayNumber * 48-1, column=4).value
                    value2300M= sheet.cell(row=dayNumber * 48-0, column=4).value
                    value2330M= sheet.cell(row=dayNumber * 48+1, column=4).value
                else:
                    value0000M= sheet.cell(row=dayNumber * 48+2, column=4).value
                    value0030M= sheet.cell(row=dayNumber * 48+3, column=4).value
                    value0100M= sheet.cell(row=dayNumber * 48+4, column=4).value
                    value0130M= sheet.cell(row=dayNumber * 48+5, column=4).value
                    value0200M= sheet.cell(row=dayNumber * 48+6, column=4).value
                    value0230M= sheet.cell(row=dayNumber * 48+7, column=4).value
                    value0300M= sheet.cell(row=dayNumber * 48+8, column=4).value
                    value0330M= sheet.cell(row=dayNumber * 48+9, column=4).value
                    value0400M= sheet.cell(row=dayNumber * 48+10, column=4).value
                    value0430M= sheet.cell(row=dayNumber * 48+11, column=4).value
                    value0500M= sheet.cell(row=dayNumber * 48+12, column=4).value
                    value0530M= sheet.cell(row=dayNumber * 48+13, column=4).value
                    value0600M= sheet.cell(row=dayNumber * 48+14, column=4).value
                    value0630M= sheet.cell(row=dayNumber * 48+15, column=4).value
                    value0700M= sheet.cell(row=dayNumber * 48+16, column=4).value
                    value0730M= sheet.cell(row=dayNumber * 48+17, column=4).value
                    value0800M= sheet.cell(row=dayNumber * 48+18, column=4).value
                    value0830M= sheet.cell(row=dayNumber * 48+19, column=4).value
                    value0900M= sheet.cell(row=dayNumber * 48+20, column=4).value
                    value0930M= sheet.cell(row=dayNumber * 48+21, column=4).value
                    value1000M= sheet.cell(row=dayNumber * 48+22, column=4).value
                    value1030M= sheet.cell(row=dayNumber * 48+23, column=4).value
                    value1100M= sheet.cell(row=dayNumber * 48+24, column=4).value
                    value1130M= sheet.cell(row=dayNumber * 48+25, column=4).value
                    value1200M = sheet.cell(row=dayNumber * 48+26, column=4).value
                    value1230M = sheet.cell(row=dayNumber * 48+27, column=4).value
                    value1300M = sheet.cell(row=dayNumber * 48+28, column=4).value
                    value1330M = sheet.cell(row=dayNumber * 48+29, column=4).value
                    value1400M= sheet.cell(row=dayNumber * 48+30, column=4).value
                    value1430M= sheet.cell(row=dayNumber * 48+31, column=4).value
                    value1500M= sheet.cell(row=dayNumber * 48+32, column=4).value
                    value1530M= sheet.cell(row=dayNumber * 48+33, column=4).value
                    value1600M= sheet.cell(row=dayNumber * 48+34, column=4).value
                    value1630M= sheet.cell(row=dayNumber * 48+35, column=4).value
                    value1700M= sheet.cell(row=dayNumber * 48+36, column=4).value
                    value1730M= sheet.cell(row=dayNumber * 48+37, column=4).value
                    value1800M= sheet.cell(row=dayNumber * 48+38, column=4).value
                    value1830M= sheet.cell(row=dayNumber * 48+39, column=4).value
                    value1900M= sheet.cell(row=dayNumber * 48+40, column=4).value
                    value1930M= sheet.cell(row=dayNumber * 48+41, column=4).value
                    value2000M= sheet.cell(row=dayNumber * 48+42, column=4).value
                    value2030M= sheet.cell(row=dayNumber * 48+43, column=4).value
                    value2100M= sheet.cell(row=dayNumber * 48+44, column=4).value
                    value2130M= sheet.cell(row=dayNumber * 48+45, column=4).value
                    value2200M= sheet.cell(row=dayNumber * 48+46, column=4).value
                    value2230M= sheet.cell(row=dayNumber * 48+47, column=4).value
                    value2300M= sheet.cell(row=dayNumber * 48+48, column=4).value
                    value2330M= sheet.cell(row=dayNumber * 48+49, column=4).value
                    
                sheet_LE_Corr25.cell(row=1, column=dayNumber+1).value = dayNumber     
                sheet_LE_Corr25.cell(row=StationIndex.index(StationName)+2, column=daysToStart+1).value = np.mean([value0000M, value0030M, value0100M, value0130M, value0200M, value0230M, value0300M, value0330M, value0400M, value0430M, value0500M, value0530M, value0600M, value0630M, value0700M, value0730M, value0800M, value0830M, value0900M, value0930M, value1000M, value1030M, value1100M, value1130M, value1200M, value1230M, value1300M, value1330M, value1400M, value1430M, value1500M, value1530M, value1600M, value1630M, value1700M, value1730M, value1800M, value1830M, value1900M, value1930M, value2000M, value2030M, value2100M, value2130M, value2200M, value2230M, value2300M, value2330M])
                
                if(day==1):
                    if(value1200M == -9999 or value1230M == -9999 or value1300M == -9999 or value1330M == -9999): #this is done so that if we have one bad value then all the values become bad.
                        value0000= -9999
                        value0030= -9999
                        value0100= -9999
                        value0130= -9999
                        value0200= -9999
                        value0230= -9999
                        value0300= -9999
                        value0330= -9999
                        value0400= -9999
                        value0430= -9999
                        value0500= -9999
                        value0530= -9999
                        value0600= -9999
                        value0630= -9999
                        value0700= -9999
                        value0730= -9999
                        value0800= -9999
                        value0830= -9999
                        value0900= -9999
                        value0930= -9999
                        value1000= -9999
                        value1030= -9999
                        value1100= -9999
                        value1130= -9999
                        value1200 = -9999
                        value1230 = -9999
                        value1300 = -9999
                        value1330 = -9999
                        value1400= -9999
                        value1430= -9999
                        value1500= -9999
                        value1530= -9999
                        value1600= -9999
                        value1630= -9999
                        value1700= -9999
                        value1730= -9999
                        value1800= -9999
                        value1830= -9999
                        value1900= -9999
                        value1930= -9999
                        value2000= -9999
                        value2030= -9999
                        value2100= -9999
                        value2130= -9999
                        value2200= -9999
                        value2230= -9999
                        value2300= -9999
                        value2330= -9999
                    else:
                        value0000= sheet.cell(row=dayNumber * 48-46, column=3).value
                        value0030= sheet.cell(row=dayNumber * 48-45, column=3).value
                        value0100= sheet.cell(row=dayNumber * 48-44, column=3).value
                        value0130= sheet.cell(row=dayNumber * 48-43, column=3).value
                        value0200= sheet.cell(row=dayNumber * 48-42, column=3).value
                        value0230= sheet.cell(row=dayNumber * 48-41, column=3).value
                        value0300= sheet.cell(row=dayNumber * 48-40, column=3).value
                        value0330= sheet.cell(row=dayNumber * 48-39, column=3).value
                        value0400= sheet.cell(row=dayNumber * 48-38, column=3).value
                        value0430= sheet.cell(row=dayNumber * 48-37, column=3).value
                        value0500= sheet.cell(row=dayNumber * 48-36, column=3).value
                        value0530= sheet.cell(row=dayNumber * 48-35, column=3).value
                        value0600= sheet.cell(row=dayNumber * 48-34, column=3).value
                        value0630= sheet.cell(row=dayNumber * 48-33, column=3).value
                        value0700= sheet.cell(row=dayNumber * 48-32, column=3).value
                        value0730= sheet.cell(row=dayNumber * 48-31, column=3).value
                        value0800= sheet.cell(row=dayNumber * 48-30, column=3).value
                        value0830= sheet.cell(row=dayNumber * 48-29, column=3).value
                        value0900= sheet.cell(row=dayNumber * 48-28, column=3).value
                        value0930= sheet.cell(row=dayNumber * 48-27, column=3).value
                        value1000= sheet.cell(row=dayNumber * 48-26, column=3).value
                        value1030= sheet.cell(row=dayNumber * 48-25, column=3).value
                        value1100= sheet.cell(row=dayNumber * 48-24, column=3).value
                        value1130= sheet.cell(row=dayNumber * 48-23, column=3).value
                        value1200 = sheet.cell(row=dayNumber * 48-22, column=3).value
                        value1230 = sheet.cell(row=dayNumber * 48-21, column=3).value
                        value1300 = sheet.cell(row=dayNumber * 48-20, column=3).value
                        value1330 = sheet.cell(row=dayNumber * 48-19, column=3).value
                        value1400= sheet.cell(row=dayNumber * 48-18, column=3).value
                        value1430= sheet.cell(row=dayNumber * 48-17, column=3).value
                        value1500= sheet.cell(row=dayNumber * 48-16, column=3).value
                        value1530= sheet.cell(row=dayNumber * 48-15, column=3).value
                        value1600= sheet.cell(row=dayNumber * 48-14, column=3).value
                        value1630= sheet.cell(row=dayNumber * 48-13, column=3).value
                        value1700= sheet.cell(row=dayNumber * 48-12, column=3).value
                        value1730= sheet.cell(row=dayNumber * 48-11, column=3).value
                        value1800= sheet.cell(row=dayNumber * 48-10, column=3).value
                        value1830= sheet.cell(row=dayNumber * 48-9, column=3).value
                        value1900= sheet.cell(row=dayNumber * 48-8, column=3).value
                        value1930= sheet.cell(row=dayNumber * 48-7, column=3).value
                        value2000= sheet.cell(row=dayNumber * 48-6, column=3).value
                        value2030= sheet.cell(row=dayNumber * 48-5, column=3).value
                        value2100= sheet.cell(row=dayNumber * 48-4, column=3).value
                        value2130= sheet.cell(row=dayNumber * 48-3, column=3).value
                        value2200= sheet.cell(row=dayNumber * 48-2, column=3).value
                        value2230= sheet.cell(row=dayNumber * 48-1, column=3).value
                        value2300= sheet.cell(row=dayNumber * 48-0, column=3).value
                        value2330= sheet.cell(row=dayNumber * 48+1, column=3).value
                else:
                    if(value1200M == -9999 or value1230M == -9999 or value1300M == -9999 or value1330M == -9999):
                        value0000= -9999
                        value0030= -9999
                        value0100= -9999
                        value0130= -9999
                        value0200= -9999
                        value0230= -9999
                        value0300= -9999
                        value0330= -9999
                        value0400= -9999
                        value0430= -9999
                        value0500= -9999
                        value0530= -9999
                        value0600= -9999
                        value0630= -9999
                        value0700= -9999
                        value0730= -9999
                        value0800= -9999
                        value0830= -9999
                        value0900= -9999
                        value0930= -9999
                        value1000= -9999
                        value1030= -9999
                        value1100= -9999
                        value1130= -9999
                        value1200 = -9999
                        value1230 = -9999
                        value1300 = -9999
                        value1330 = -9999
                        value1400= -9999
                        value1430= -9999
                        value1500= -9999
                        value1530= -9999
                        value1600= -9999
                        value1630= -9999
                        value1700= -9999
                        value1730= -9999
                        value1800= -9999
                        value1830= -9999
                        value1900= -9999
                        value1930= -9999
                        value2000= -9999
                        value2030= -9999
                        value2100= -9999
                        value2130= -9999
                        value2200= -9999
                        value2230= -9999
                        value2300= -9999
                        value2330= -9999
                    else:
                        value0000= sheet.cell(row=dayNumber * 48+2, column=3).value
                        value0030= sheet.cell(row=dayNumber * 48+3, column=3).value
                        value0100= sheet.cell(row=dayNumber * 48+4, column=3).value
                        value0130= sheet.cell(row=dayNumber * 48+5, column=3).value
                        value0200= sheet.cell(row=dayNumber * 48+6, column=3).value
                        value0230= sheet.cell(row=dayNumber * 48+7, column=3).value
                        value0300= sheet.cell(row=dayNumber * 48+8, column=3).value
                        value0330= sheet.cell(row=dayNumber * 48+9, column=3).value
                        value0400= sheet.cell(row=dayNumber * 48+10, column=3).value
                        value0430= sheet.cell(row=dayNumber * 48+11, column=3).value
                        value0500= sheet.cell(row=dayNumber * 48+12, column=3).value
                        value0530= sheet.cell(row=dayNumber * 48+13, column=3).value
                        value0600= sheet.cell(row=dayNumber * 48+14, column=3).value
                        value0630= sheet.cell(row=dayNumber * 48+15, column=3).value
                        value0700= sheet.cell(row=dayNumber * 48+16, column=3).value
                        value0730= sheet.cell(row=dayNumber * 48+17, column=3).value
                        value0800= sheet.cell(row=dayNumber * 48+18, column=3).value
                        value0830= sheet.cell(row=dayNumber * 48+19, column=3).value
                        value0900= sheet.cell(row=dayNumber * 48+20, column=3).value
                        value0930= sheet.cell(row=dayNumber * 48+21, column=3).value
                        value1000= sheet.cell(row=dayNumber * 48+22, column=3).value
                        value1030= sheet.cell(row=dayNumber * 48+23, column=3).value
                        value1100= sheet.cell(row=dayNumber * 48+24, column=3).value
                        value1130= sheet.cell(row=dayNumber * 48+25, column=3).value
                        value1200 = sheet.cell(row=dayNumber * 48+26, column=3).value
                        value1230 = sheet.cell(row=dayNumber * 48+27, column=3).value
                        value1300 = sheet.cell(row=dayNumber * 48+28, column=3).value
                        value1330 = sheet.cell(row=dayNumber * 48+29, column=3).value
                        value1400= sheet.cell(row=dayNumber * 48+30, column=3).value
                        value1430= sheet.cell(row=dayNumber * 48+31, column=3).value
                        value1500= sheet.cell(row=dayNumber * 48+32, column=3).value
                        value1530= sheet.cell(row=dayNumber * 48+33, column=3).value
                        value1600= sheet.cell(row=dayNumber * 48+34, column=3).value
                        value1630= sheet.cell(row=dayNumber * 48+35, column=3).value
                        value1700= sheet.cell(row=dayNumber * 48+36, column=3).value
                        value1730= sheet.cell(row=dayNumber * 48+37, column=3).value
                        value1800= sheet.cell(row=dayNumber * 48+38, column=3).value
                        value1830= sheet.cell(row=dayNumber * 48+39, column=3).value
                        value1900= sheet.cell(row=dayNumber * 48+40, column=3).value
                        value1930= sheet.cell(row=dayNumber * 48+41, column=3).value
                        value2000= sheet.cell(row=dayNumber * 48+42, column=3).value
                        value2030= sheet.cell(row=dayNumber * 48+43, column=3).value
                        value2100= sheet.cell(row=dayNumber * 48+44, column=3).value
                        value2130= sheet.cell(row=dayNumber * 48+45, column=3).value
                        value2200= sheet.cell(row=dayNumber * 48+46, column=3).value
                        value2230= sheet.cell(row=dayNumber * 48+47, column=3).value
                        value2300= sheet.cell(row=dayNumber * 48+48, column=3).value
                        value2330= sheet.cell(row=dayNumber * 48+49, column=3).value
                 
                sheet_LE_Corr.cell(row=1, column=dayNumber+1).value = dayNumber  
                sheet_LE_Corr.cell(row=StationIndex.index(StationName)+2, column=daysToStart+1).value = np.mean([value0000, value0030, value0100, value0130, value0200, value0230, value0300, value0330, value0400, value0430, value0500, value0530, value0600, value0630, value0700, value0730, value0800, value0830, value0900, value0930, value1000, value1030, value1100, value1130, value1200, value1230, value1300, value1330, value1400, value1430, value1500, value1530, value1600, value1630, value1700, value1730, value1800, value1830, value1900, value1930, value2000, value2030, value2100, value2130, value2200, value2230, value2300, value2330])
                
                  
                
                if(day==1):
                    value0000= sheet.cell(row=dayNumber * 48-46, column=5).value
                    value0030= sheet.cell(row=dayNumber * 48-45, column=5).value
                    value0100= sheet.cell(row=dayNumber * 48-44, column=5).value
                    value0130= sheet.cell(row=dayNumber * 48-43, column=5).value
                    value0200= sheet.cell(row=dayNumber * 48-42, column=5).value
                    value0230= sheet.cell(row=dayNumber * 48-41, column=5).value
                    value0300= sheet.cell(row=dayNumber * 48-40, column=5).value
                    value0330= sheet.cell(row=dayNumber * 48-39, column=5).value
                    value0400= sheet.cell(row=dayNumber * 48-38, column=5).value
                    value0430= sheet.cell(row=dayNumber * 48-37, column=5).value
                    value0500= sheet.cell(row=dayNumber * 48-36, column=5).value
                    value0530= sheet.cell(row=dayNumber * 48-35, column=5).value
                    value0600= sheet.cell(row=dayNumber * 48-34, column=5).value
                    value0630= sheet.cell(row=dayNumber * 48-33, column=5).value
                    value0700= sheet.cell(row=dayNumber * 48-32, column=5).value
                    value0730= sheet.cell(row=dayNumber * 48-31, column=5).value
                    value0800= sheet.cell(row=dayNumber * 48-30, column=5).value
                    value0830= sheet.cell(row=dayNumber * 48-29, column=5).value
                    value0900= sheet.cell(row=dayNumber * 48-28, column=5).value
                    value0930= sheet.cell(row=dayNumber * 48-27, column=5).value
                    value1000= sheet.cell(row=dayNumber * 48-26, column=5).value
                    value1030= sheet.cell(row=dayNumber * 48-25, column=5).value
                    value1100= sheet.cell(row=dayNumber * 48-24, column=5).value
                    value1130= sheet.cell(row=dayNumber * 48-23, column=5).value
                    value1200 = sheet.cell(row=dayNumber * 48-22, column=5).value
                    value1230 = sheet.cell(row=dayNumber * 48-21, column=5).value
                    value1300 = sheet.cell(row=dayNumber * 48-20, column=5).value
                    value1330 = sheet.cell(row=dayNumber * 48-19, column=5).value
                    value1400= sheet.cell(row=dayNumber * 48-18, column=5).value
                    value1430= sheet.cell(row=dayNumber * 48-17, column=5).value
                    value1500= sheet.cell(row=dayNumber * 48-16, column=5).value
                    value1530= sheet.cell(row=dayNumber * 48-15, column=5).value
                    value1600= sheet.cell(row=dayNumber * 48-14, column=5).value
                    value1630= sheet.cell(row=dayNumber * 48-13, column=5).value
                    value1700= sheet.cell(row=dayNumber * 48-12, column=5).value
                    value1730= sheet.cell(row=dayNumber * 48-11, column=5).value
                    value1800= sheet.cell(row=dayNumber * 48-10, column=5).value
                    value1830= sheet.cell(row=dayNumber * 48-9, column=5).value
                    value1900= sheet.cell(row=dayNumber * 48-8, column=5).value
                    value1930= sheet.cell(row=dayNumber * 48-7, column=5).value
                    value2000= sheet.cell(row=dayNumber * 48-6, column=5).value
                    value2030= sheet.cell(row=dayNumber * 48-5, column=5).value
                    value2100= sheet.cell(row=dayNumber * 48-4, column=5).value
                    value2130= sheet.cell(row=dayNumber * 48-3, column=5).value
                    value2200= sheet.cell(row=dayNumber * 48-2, column=5).value
                    value2230= sheet.cell(row=dayNumber * 48-1, column=5).value
                    value2300= sheet.cell(row=dayNumber * 48-0, column=5).value
                    value2330= sheet.cell(row=dayNumber * 48+1, column=5).value
                else:
                    value0000= sheet.cell(row=dayNumber * 48+2, column=5).value
                    value0030= sheet.cell(row=dayNumber * 48+3, column=5).value
                    value0100= sheet.cell(row=dayNumber * 48+4, column=5).value
                    value0130= sheet.cell(row=dayNumber * 48+5, column=5).value
                    value0200= sheet.cell(row=dayNumber * 48+6, column=5).value
                    value0230= sheet.cell(row=dayNumber * 48+7, column=5).value
                    value0300= sheet.cell(row=dayNumber * 48+8, column=5).value
                    value0330= sheet.cell(row=dayNumber * 48+9, column=5).value
                    value0400= sheet.cell(row=dayNumber * 48+10, column=5).value
                    value0430= sheet.cell(row=dayNumber * 48+11, column=5).value
                    value0500= sheet.cell(row=dayNumber * 48+12, column=5).value
                    value0530= sheet.cell(row=dayNumber * 48+13, column=5).value
                    value0600= sheet.cell(row=dayNumber * 48+14, column=5).value
                    value0630= sheet.cell(row=dayNumber * 48+15, column=5).value
                    value0700= sheet.cell(row=dayNumber * 48+16, column=5).value
                    value0730= sheet.cell(row=dayNumber * 48+17, column=5).value
                    value0800= sheet.cell(row=dayNumber * 48+18, column=5).value
                    value0830= sheet.cell(row=dayNumber * 48+19, column=5).value
                    value0900= sheet.cell(row=dayNumber * 48+20, column=5).value
                    value0930= sheet.cell(row=dayNumber * 48+21, column=5).value
                    value1000= sheet.cell(row=dayNumber * 48+22, column=5).value
                    value1030= sheet.cell(row=dayNumber * 48+23, column=5).value
                    value1100= sheet.cell(row=dayNumber * 48+24, column=5).value
                    value1130= sheet.cell(row=dayNumber * 48+25, column=5).value
                    value1200 = sheet.cell(row=dayNumber * 48+26, column=5).value
                    value1230 = sheet.cell(row=dayNumber * 48+27, column=5).value
                    value1300 = sheet.cell(row=dayNumber * 48+28, column=5).value
                    value1330 = sheet.cell(row=dayNumber * 48+29, column=5).value
                    value1400= sheet.cell(row=dayNumber * 48+30, column=5).value
                    value1430= sheet.cell(row=dayNumber * 48+31, column=5).value
                    value1500= sheet.cell(row=dayNumber * 48+32, column=5).value
                    value1530= sheet.cell(row=dayNumber * 48+33, column=5).value
                    value1600= sheet.cell(row=dayNumber * 48+34, column=5).value
                    value1630= sheet.cell(row=dayNumber * 48+35, column=5).value
                    value1700= sheet.cell(row=dayNumber * 48+36, column=5).value
                    value1730= sheet.cell(row=dayNumber * 48+37, column=5).value
                    value1800= sheet.cell(row=dayNumber * 48+38, column=5).value
                    value1830= sheet.cell(row=dayNumber * 48+39, column=5).value
                    value1900= sheet.cell(row=dayNumber * 48+40, column=5).value
                    value1930= sheet.cell(row=dayNumber * 48+41, column=5).value
                    value2000= sheet.cell(row=dayNumber * 48+42, column=5).value
                    value2030= sheet.cell(row=dayNumber * 48+43, column=5).value
                    value2100= sheet.cell(row=dayNumber * 48+44, column=5).value
                    value2130= sheet.cell(row=dayNumber * 48+45, column=5).value
                    value2200= sheet.cell(row=dayNumber * 48+46, column=5).value
                    value2230= sheet.cell(row=dayNumber * 48+47, column=5).value
                    value2300= sheet.cell(row=dayNumber * 48+48, column=5).value
                    value2330= sheet.cell(row=dayNumber * 48+49, column=5).value
                 
                sheet_LE_Corr75.cell(row=1, column=dayNumber+1).value = dayNumber    
                sheet_LE_Corr75.cell(row=StationIndex.index(StationName)+2, column=daysToStart+1).value = np.mean([value0000, value0030, value0100, value0130, value0200, value0230, value0300, value0330, value0400, value0430, value0500, value0530, value0600, value0630, value0700, value0730, value0800, value0830, value0900, value0930, value1000, value1030, value1100, value1130, value1200, value1230, value1300, value1330, value1400, value1430, value1500, value1530, value1600, value1630, value1700, value1730, value1800, value1830, value1900, value1930, value2000, value2030, value2100, value2130, value2200, value2230, value2300, value2330])
                
                
                
            except:     #if something is wrong we just write -9999 (maybe we should leave it blank instead?)
                #print("data missing")
                sheet_LE_Corr.cell(row=StationIndex.index(StationName)+2, column=daysToStart+1).value = -9999
                sheet_LE_Corr25.cell(row=StationIndex.index(StationName)+2, column=daysToStart+1).value = -9999
                sheet_LE_Corr75.cell(row=StationIndex.index(StationName)+2, column=daysToStart+1).value = -9999
                sheet_H_Corr.cell(row=StationIndex.index(StationName)+2, column=daysToStart+1).value = -9999
                sheet_H_Corr25.cell(row=StationIndex.index(StationName)+2, column=daysToStart+1).value = -9999
                sheet_H_Corr75.cell(row=StationIndex.index(StationName)+2, column=daysToStart+1).value = -9999
                
                
                
            
            #---------------------------------------------------------------------------------------------------------------------------------------------------------
            #now we do the same but for the daily averages: 
            
            
            try:
                DayValues  = np.zeros([48,1])
                
                
                
                for timeOfDay in range(1,49):
                    
                    if(day==1):
                        if -9999 in DayValues:
                            DayValues = np.full([48,1],-9999)
                        else:
                            valueAtTime = sheet.cell(row=dayNumber * 48+ timeOfDay+1, column=3).value
                            DayValues[timeOfDay-1] = valueAtTime
                            
                    else:
                        
                        valueAtTime = sheet.cell(row=dayNumber * 48+timeOfDay+1, column=3).value
                        if -9999 in DayValues:
                            DayValues = np.full([48,1],-9999)
                        else:
                            DayValues[timeOfDay-1] = valueAtTime
                            
                        
                        
                    
                    
                sheet_LE_CorrDay.cell(row=1, column=dayNumber+1).value = dayNumber
                
                valueOfDay = float(sum(DayValues)/len(DayValues))
                #print(valueOfDay)
                #print(dayNumber)
                sheet_LE_CorrDay.cell(row=StationIndex.index(StationName)+2, column=daysToStart+1).value = valueOfDay
                
                #_---------------------------------------------------------------------------------------------------------------
                DayValues = np.zeros([48,1])

                for timeOfDay in range(1,49):
                    
                    if(day==1):
                        if -9999 in DayValues:
                            DayValues = np.full([48,1],-9999)
                        else:
                            valueAtTime = sheet.cell(row=dayNumber * 48+ timeOfDay+1, column=4).value
                            DayValues[timeOfDay-1] = valueAtTime
                            
                    else:
                        
                        valueAtTime = sheet.cell(row=dayNumber * 48+timeOfDay+1, column=4).value
                        if -9999 in DayValues:
                            DayValues = np.full([48,1],-9999)
                        else:
                            DayValues[timeOfDay-1] = valueAtTime
                            
                        
                        
                    
                    
                sheet_LE_Corr25Day.cell(row=1, column=dayNumber+1).value = dayNumber
                
                valueOfDay = float(sum(DayValues)/len(DayValues))
                #print(valueOfDay)
                #print(dayNumber)
                sheet_LE_Corr25Day.cell(row=StationIndex.index(StationName)+2, column=daysToStart+1).value = valueOfDay
                
                #_---------------------------------------------------------------------------------------------------------------
                DayValues = np.zeros([48,1])

                for timeOfDay in range(1,49):
                    
                    if(day==1):
                        if -9999 in DayValues:
                            DayValues = np.full([48,1],-9999)
                        else:
                            valueAtTime = sheet.cell(row=dayNumber * 48+ timeOfDay+1, column=5).value
                            DayValues[timeOfDay-1] = valueAtTime
                            
                    else:
                        
                        valueAtTime = sheet.cell(row=dayNumber * 48+timeOfDay+1, column=5).value
                        if -9999 in DayValues:
                            DayValues = np.full([48,1],-9999)
                        else:
                            DayValues[timeOfDay-1] = valueAtTime
                            
                        
                        
                    
                    
                sheet_LE_Corr75Day.cell(row=1, column=dayNumber+1).value = dayNumber
                
                valueOfDay = float(sum(DayValues)/len(DayValues))
                #print(valueOfDay)
                #print(dayNumber)
                sheet_LE_Corr75Day.cell(row=StationIndex.index(StationName)+2, column=daysToStart+1).value = valueOfDay
                
                #_---------------------------------------------------------------------------------------------------------------
                DayValues = np.zeros([48,1])

                for timeOfDay in range(1,49):
                    
                    if(day==1):
                        if -9999 in DayValues:
                            DayValues = np.full([48,1],-9999)
                        else:
                            valueAtTime = sheet.cell(row=dayNumber * 48+ timeOfDay+1, column=6).value
                            DayValues[timeOfDay-1] = valueAtTime
                            
                    else:
                        
                        valueAtTime = sheet.cell(row=dayNumber * 48+timeOfDay+1, column=6).value
                        if -9999 in DayValues:
                            DayValues = np.full([48,1],-9999)
                        else:
                            DayValues[timeOfDay-1] = valueAtTime
                            
                        
                        
                    
                    
                
                
                
            except:
                #print("data missing average")
                sheet_LE_CorrDay.cell(row=StationIndex.index(StationName)+2, column=daysToStart+1).value = -9999
                sheet_LE_Corr25Day.cell(row=StationIndex.index(StationName)+2, column=daysToStart+1).value = -9999
                sheet_LE_Corr75Day.cell(row=StationIndex.index(StationName)+2, column=daysToStart+1).value = -9999
                sheet_H_CorrDay.cell(row=StationIndex.index(StationName)+2, column=daysToStart+1).value = -9999
                sheet_H_Corr25Day.cell(row=StationIndex.index(StationName)+2, column=daysToStart+1).value = -9999
                sheet_H_Corr75Day.cell(row=StationIndex.index(StationName)+2, column=daysToStart+1).value = -9999
                
                
        
            dayNumber = dayNumber+1
            daysToStart = daysToStart + 1
                   
                   
                
                
                
        workbookCombined.save("E:\\HenrikTSEB\\FluxnetDataset\\FluxnetDataset\\Fluxnet_xlxs\\Fluxnet_combined_Daily_Test.xlsx")                 
        


